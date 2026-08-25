import json
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pytest

from devoteam_reference_ai.phase5_bm25 import BM25Index
from devoteam_reference_ai.phase5_retrieval import (
    FilterEngine,
    HybridSearchEngine,
    _query_metrics,
    build_bootstrap_probes,
    build_embeddings_resumable,
    build_retrieval_texts,
    create_expert_gold_set_workbook,
    load_phase5_config,
)


CONFIG_PATH = Path(__file__).parents[1] / "config" / "phase5_retrieval.yaml"


@pytest.fixture()
def config():
    return load_phase5_config(CONFIG_PATH)


def _chunk_rows():
    rows = []
    for index, (client, country, year, security, text) in enumerate(
        [
            ("Banque A", "Tunisie", "2022", "INTERNAL", "schéma directeur bancaire"),
            ("Télécom B", "Maroc", "2020", "INTERNAL", "audit réseau télécom"),
            ("Secret C", "Tunisie", "2024", "RESTRICTED", "audit bancaire secret"),
        ]
    ):
        rows.append(
            {
                "chunk_id": f"chunk-{index}",
                "document_id": f"doc-{index}",
                "source_file_name": f"proof-{index}.pdf",
                "chunk_text": text,
                "security_classification": security,
                "country_values_json": json.dumps([country]),
                "business_unit_values_json": json.dumps(["MENA"]),
                "client_values_json": json.dumps([client]),
                "sector_values_json": json.dumps(["Banque" if index != 1 else "Télécom"]),
                "service_nature_values_json": json.dumps(["Conseil"]),
                "offering_values_json": json.dumps(["Digital"]),
                "project_year_values_json": json.dumps([year]),
                "attestation_available_values_json": json.dumps(["Oui"]),
                "document_type": "ATTESTATION",
                "data_quality_status": "PASS",
                "document_language": "fr",
                "citation_label": f"proof-{index}.pdf — page 1",
                "citation_uri": f"https://drive.google.com/file/d/doc-{index}/view#page=1",
                "reference_rows_json": json.dumps([index + 2]),
            }
        )
    return pd.DataFrame(rows)


def test_config_pins_local_multilingual_model_and_expert_gate(config):
    assert config["embedding"]["model_id"] == "intfloat/multilingual-e5-base"
    assert len(config["embedding"]["revision"]) == 40
    assert config["embedding"]["local_execution_only"] is True
    assert config["security"]["external_embedding_api_enabled"] is False
    assert config["security"]["external_llm_enabled"] is False
    assert config["evaluation"]["expert_gold_set_required_for_promotion"] is True


def test_filter_engine_requires_authorization_and_applies_hard_filters(config):
    chunks = _chunk_rows()
    engine = FilterEngine(chunks, config)
    with pytest.raises(PermissionError):
        engine.mask(allowed_security_classifications=[])
    mask = engine.mask(
        allowed_security_classifications=["INTERNAL"],
        hard_filters={"country": "Tunisie", "year_after": 2021},
    )
    assert mask.tolist() == [True, False, False]


def test_hybrid_search_is_deterministic_and_preserves_citations(config):
    chunks = _chunk_rows()
    texts = build_retrieval_texts(chunks)
    bm25 = BM25Index.build(texts)
    embeddings = np.asarray(
        [[1.0, 0.0, 0.0], [0.0, 1.0, 0.0], [0.8, 0.0, 0.6]], dtype=np.float32
    )
    engine = HybridSearchEngine(chunks, texts, bm25, embeddings, config)
    query_vector = np.asarray([1.0, 0.0, 0.0], dtype=np.float32)
    first = engine.search_chunks(
        "banque schéma directeur",
        allowed_security_classifications=["INTERNAL"],
        mode="hybrid",
        top_k=3,
        query_vector=query_vector,
    )
    second = engine.search_chunks(
        "banque schéma directeur",
        allowed_security_classifications=["INTERNAL"],
        mode="hybrid",
        top_k=3,
        query_vector=query_vector,
    )
    assert first.chunk_id.tolist() == second.chunk_id.tolist()
    assert first.iloc[0].chunk_id == "chunk-0"
    assert "chunk-2" not in first.chunk_id.tolist()
    assert first.citation_uri.str.startswith("https://drive.google.com/").all()


def test_retrieval_context_is_bounded_and_keeps_evidence(config):
    chunks = _chunk_rows()
    chunks.loc[0, "client_values_json"] = json.dumps(["X" * 5000])
    texts = build_retrieval_texts(chunks)
    assert max(map(len, texts)) <= 1700
    assert "schéma directeur bancaire" in texts[0]


def test_query_metrics_known_example():
    metrics = _query_metrics({"a", "b"}, ["x", "a", "y", "b"])
    assert metrics["recall_at_10"] == 1.0
    assert metrics["precision_at_5"] == 0.4
    assert metrics["mrr"] == 0.5
    assert 0.0 < metrics["ndcg_at_10"] <= 1.0


def test_bootstrap_probe_builder_is_deterministic(config):
    records = []
    for index in range(60):
        records.append(
            {
                "document_id": f"doc-{index:03d}",
                "retrieval_eligible": True,
                "client_values_json": json.dumps([f"Client {index:03d}"]),
                "offering_values_json": json.dumps([f"Offre {index % 7}"]),
                "sector_values_json": json.dumps([f"Secteur {index % 9}"]),
                "service_nature_values_json": json.dumps([f"Mission {index % 11}"]),
                "country_values_json": json.dumps(["Tunisie" if index % 2 else "Maroc"]),
            }
        )
    documents = pd.DataFrame(records)
    first = build_bootstrap_probes(documents, config)
    second = build_bootstrap_probes(documents, config)
    pd.testing.assert_frame_equal(first, second)
    assert len(first) == 50
    assert first.primary_document_id.nunique() == 50
    assert first.label_source.eq("BOOTSTRAP_METADATA_NOT_EXPERT").all()


class FakeAdapter:
    device = "cpu"

    def __init__(self):
        self.calls = 0

    def encode_passages(self, texts):
        self.calls += 1
        rows = []
        for text in texts:
            seed = sum(ord(character) for character in text)
            value = np.asarray([seed % 7 + 1, seed % 5 + 1, seed % 3 + 1, 1], dtype=np.float32)
            rows.append(value / np.linalg.norm(value))
        return np.asarray(rows, dtype=np.float32)


def test_embedding_checkpoint_is_reusable(tmp_path, config):
    local = json.loads(json.dumps(config))
    local["embedding"]["dimensions"] = 4
    local["embedding"]["batch_size_cpu"] = 2
    local["embedding"]["checkpoint_every_rows"] = 3
    output = tmp_path / "embeddings.npy"
    progress = tmp_path / "progress.json"
    first_adapter = FakeAdapter()
    first = build_embeddings_resumable(
        texts=["a", "b", "c", "d", "e"],
        adapter=first_adapter,
        output_path=output,
        progress_path=progress,
        config=local,
        input_hash="abc",
        progress=lambda _: None,
    )
    assert first.shape == (5, 4)
    second_adapter = FakeAdapter()
    second = build_embeddings_resumable(
        texts=["a", "b", "c", "d", "e"],
        adapter=second_adapter,
        output_path=output,
        progress_path=progress,
        config=local,
        input_hash="abc",
        progress=lambda _: None,
    )
    np.testing.assert_allclose(first, second)
    assert second_adapter.calls == 0


def test_expert_gold_workbook_has_controlled_structure(tmp_path):
    output = tmp_path / "gold.xlsx"
    create_expert_gold_set_workbook(output, query_target=50)
    workbook = openpyxl.load_workbook(output, read_only=False, data_only=False)
    assert workbook.sheetnames == ["Instructions", "Queries", "Labels"]
    assert workbook["Queries"].max_row == 51
    assert workbook["Labels"].max_row == 501
