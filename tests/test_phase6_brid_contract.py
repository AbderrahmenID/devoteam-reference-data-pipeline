from pathlib import Path
import json
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data/opportunities/OPP-2098dd1874292d22/phase6_brid_controlled_case_v2"

def test_brid_case_outputs_exist():
    required = ["offer_profile.json", "requirements.jsonl", "eligibility_rules.json",
        "scoring_model.json", "filter_proposals.json", "matching_contract.json",
        "BRID_PHASE_6_REVIEW.xlsx", "PHASE_6_BRID_QUALITY_GATE.json",
        "PHASE_6_BRID_REPORT.md", "PHASE_6_BRID_MANIFEST.json", "SHA256SUMS.txt", "_SUCCESS.json"]
    assert all((OUTPUT / name).is_file() for name in required)

def test_source_and_status():
    profile = json.loads((OUTPUT / "offer_profile.json").read_text(encoding="utf-8"))
    assert profile["label"] == "SYNTHETIC_TEST_ONLY"
    assert profile["source_pages"] == 7
    assert profile["status"] == "READY_FOR_HUMAN_REVIEW"
    assert profile["external_llm_calls"] == 0
    assert profile["business_filters_auto_applied"] == 0

def test_portfolio_rules_are_structured():
    rules = json.loads((OUTPUT / "eligibility_rules.json").read_text(encoding="utf-8"))
    by_code = {row["code"]: row for row in rules}
    assert by_code["ELIG-SDSI-3"]["minimum_count"] == 3
    assert by_code["ELIG-SDSI-BFSI-2"]["minimum_count"] == 2
    assert by_code["ELIG-SIGNED"]["level"] == "PER_REFERENCE"
    assert all(row["human_review_required"] and not row["approved"] for row in rules)

def test_scoring_and_promotion_boundary():
    scoring = json.loads((OUTPUT / "scoring_model.json").read_text(encoding="utf-8"))
    contract = json.loads((OUTPUT / "matching_contract.json").read_text(encoding="utf-8"))
    assert sum(row["points"] for row in scoring) == 100
    assert contract["approval_status"] == "PENDING_HUMAN_REVIEW"
    assert contract["executable_for_phase5_2"] is False

def test_filters_are_visible_and_unconfirmed():
    rows = json.loads((OUTPUT / "filter_proposals.json").read_text(encoding="utf-8"))
    assert all(row["visible_to_user"] for row in rows)
    assert all(row["requires_human_confirmation"] and not row["confirmed"] for row in rows)

def test_workbook_has_review_surfaces():
    wb = openpyxl.load_workbook(OUTPUT / "BRID_PHASE_6_REVIEW.xlsx", data_only=False)
    assert wb.sheetnames == ["Overview", "Requirements", "EligibilityRules", "Scoring", "Filters"]
    for sheet in wb.sheetnames[1:]:
        assert "reviewer_decision" in [cell.value for cell in wb[sheet][1]]
