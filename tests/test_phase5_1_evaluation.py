import json
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from devoteam_reference_ai.phase5_1_evaluation import (
    QUERY_COLUMNS,
    _metric_row,
    create_query_intake_workbook,
    evaluate_rankings,
    label_agreement,
    load_phase5_1_config,
    paired_bootstrap,
    validate_query_intake,
    weighted_kappa,
)


CONFIG_PATH = Path(__file__).parents[1] / "config" / "phase5_1_evaluation.yaml"


def test_config_enforces_frozen_gold_set_and_no_auto_promotion():
    config = load_phase5_1_config(CONFIG_PATH)
    assert config["queries"]["target_count"] == 50
    assert config["labeling"]["labelers_required"] == 2
    assert config["evaluation"]["gold_set_purpose"] == "FROZEN_TEST_ONLY_NO_TUNING"
    assert config["evaluation"]["no_automatic_production_promotion"] is True
    assert config["security"]["external_llm_enabled"] is False
    assert config["security"]["phase5_output_mutation_allowed"] is False


def test_query_intake_template_is_controlled(tmp_path):
    config = load_phase5_1_config(CONFIG_PATH)
    output = tmp_path / "intake.xlsx"
    create_query_intake_workbook(output, config)
    workbook = openpyxl.load_workbook(output, read_only=False, data_only=False)
    assert workbook.sheetnames == ["Instructions", "Governance", "Queries", "TemplateManifest"]
    assert workbook["Queries"].max_row == 51
    assert [cell.value for cell in workbook["Queries"][1]] == QUERY_COLUMNS
    assert workbook["TemplateManifest"].sheet_state == "hidden"


def test_query_validation_rejects_bootstrap_style_incomplete_template(tmp_path):
    config = load_phase5_1_config(CONFIG_PATH)
    output = tmp_path / "intake.xlsx"
    create_query_intake_workbook(output, config)
    _, _, issues = validate_query_intake(output, config)
    assert issues
    assert any("query_text" in issue for issue in issues)
    assert any("not derived" in issue for issue in issues)
    assert any("Governance" in issue for issue in issues)


def test_weighted_kappa_known_cases():
    assert weighted_kappa([0, 1, 2], [0, 1, 2]) == 1.0
    value = weighted_kappa([0, 0, 2, 2], [0, 1, 1, 2])
    assert 0.0 < value < 1.0


def test_label_agreement_aligns_by_candidate_not_row_order():
    first = pd.DataFrame(
        {
            "query_id": ["Q1", "Q1"],
            "candidate_id": ["A", "B"],
            "relevance": [2, 0],
        }
    )
    second = pd.DataFrame(
        {
            "query_id": ["Q1", "Q1"],
            "candidate_id": ["B", "A"],
            "relevance": [0, 1],
        }
    )
    merged, metrics = label_agreement(first, second)
    assert len(merged) == 2
    assert metrics["raw_agreement"] == 0.5
    assert metrics["disagreements"] == 1


def test_metric_row_uses_binary_relevance_and_graded_ndcg():
    values = _metric_row({"a": 2, "b": 1, "c": 0}, ["x", "a", "c", "b"])
    assert values["recall_at_10"] == 1.0
    assert values["precision_at_5"] == 0.4
    assert values["mrr"] == 0.5
    assert 0.0 < values["ndcg_at_10"] <= 1.0


def test_no_relevant_query_is_reported_not_silently_divided():
    values = _metric_row({"a": 0, "b": 0}, ["a", "b"])
    assert values["answerable"] == 0.0
    assert np.isnan(values["recall_at_10"])
    assert values["precision_at_5"] == 0.0


def test_evaluation_is_system_aligned_and_deterministic():
    queries = pd.DataFrame(
        {
            "query_id": ["Q1", "Q2"],
            "language": ["fr", "ar"],
            "query_type": ["STANDARD", "SPARSE"],
        }
    )
    candidates = pd.DataFrame(
        {
            "query_id": ["Q1", "Q1", "Q2", "Q2"],
            "candidate_id": ["A", "B", "C", "D"],
            "document_id": ["d1", "d2", "d3", "d4"],
        }
    )
    adjudicated = pd.DataFrame(
        {
            "query_id": ["Q1", "Q1", "Q2", "Q2"],
            "candidate_id": ["A", "B", "C", "D"],
            "adjudicated_relevance": [2, 0, 1, 0],
        }
    )
    rankings = pd.DataFrame(
        [
            ("Q1", "bm25", "d2", 1),
            ("Q1", "bm25", "d1", 2),
            ("Q1", "hybrid", "d1", 1),
            ("Q1", "hybrid", "d2", 2),
            ("Q2", "bm25", "d4", 1),
            ("Q2", "bm25", "d3", 2),
            ("Q2", "hybrid", "d3", 1),
            ("Q2", "hybrid", "d4", 2),
        ],
        columns=["query_id", "retrieval_mode", "document_id", "rank"],
    )
    per_query, aggregate = evaluate_rankings(
        rankings, candidates, adjudicated, queries, ["bm25", "hybrid"]
    )
    assert aggregate["bm25"]["recall_at_10"] == 1.0
    assert aggregate["hybrid"]["recall_at_10"] == 1.0
    assert aggregate["hybrid"]["mrr"] > aggregate["bm25"]["mrr"]
    first = paired_bootstrap(per_query, ["bm25", "hybrid"], ["mrr"], 200, 42)
    second = paired_bootstrap(per_query, ["bm25", "hybrid"], ["mrr"], 200, 42)
    assert first == second
