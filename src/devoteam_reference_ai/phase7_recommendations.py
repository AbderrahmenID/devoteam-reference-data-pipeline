from __future__ import annotations

import hashlib
import json
import math
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
import yaml

from .phase5_bm25 import BM25Index, normalize_search_text
from .phase5_retrieval import FilterEngine
from .phase6_opportunity import sha256_file, verify_phase6_run


YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b")


class RecommendationError(RuntimeError):
    """Raised when recommendation controls or signed inputs are incomplete."""


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _atomic_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(value, encoding="utf-8")
    os.replace(temporary, path)


def _atomic_json(path: Path, value: Any) -> None:
    _atomic_text(path, json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n")


def _json_list(value: Any) -> list[str]:
    if isinstance(value, list):
        raw = value
    elif value is None or (isinstance(value, float) and math.isnan(value)):
        raw = []
    else:
        try:
            parsed = json.loads(str(value))
            raw = parsed if isinstance(parsed, list) else [parsed]
        except (TypeError, ValueError, json.JSONDecodeError):
            raw = [value]
    return [" ".join(str(item or "").split()) for item in raw if str(item or "").strip()]


def load_phase7_config(path: Path) -> dict:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    if int(config.get("phase", 0)) != 7:
        raise ValueError("Expected Phase 7 configuration")
    security = config.get("security", {})
    blocked = (
        "external_llm_enabled",
        "external_embedding_api_enabled",
        "raw_opportunity_logging_allowed",
        "security_filters_disableable",
    )
    if any(security.get(field) for field in blocked):
        raise ValueError("Phase 7 security defaults were weakened")
    retrieval = config.get("retrieval", {})
    if retrieval.get("mode") != "bm25_secure_baseline":
        raise ValueError("This signed Phase 7 package is the secure BM25 baseline")
    if retrieval.get("cross_encoder_enabled"):
        raise ValueError("Cross-encoder promotion requires the pending expert evaluation")
    scoring = config.get("scoring", {})
    if not math.isclose(sum(float(value) for value in scoring["component_weights"].values()), 1.0):
        raise ValueError("Recommendation component weights must sum to one")
    if config.get("promotion", {}).get("automatic_production_promotion"):
        raise ValueError("Phase 7 may not promote itself to production")
    return config


def verify_phase5_index(phase5_root: Path, config: dict) -> dict:
    manifest_path = phase5_root / "PHASE_5_MANIFEST.json"
    sums_path = phase5_root / "SHA256SUMS.txt"
    success_path = phase5_root / "_SUCCESS.json"
    for path in (manifest_path, sums_path, success_path):
        if not path.exists():
            raise FileNotFoundError(path)
    expected = config["input"]
    if sha256_file(manifest_path) != expected["expected_phase5_manifest_sha256"]:
        raise AssertionError("Phase 5 manifest hash changed")
    if sha256_file(sums_path) != expected["expected_phase5_sha256sums_sha256"]:
        raise AssertionError("Phase 5 checksum file changed")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    success = json.loads(success_path.read_text(encoding="utf-8"))
    if manifest.get("status") != "TECHNICAL_PASS" or manifest.get("qa_gate") != "PASS":
        raise AssertionError("Phase 5 technical gate is incomplete")
    if success.get("manifest_sha256") != expected["expected_phase5_manifest_sha256"]:
        raise AssertionError("Phase 5 success marker is inconsistent")
    sums = {
        name: digest
        for digest, name in (
            line.split("  ", 1)
            for line in sums_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        )
    }
    required = config["input"]["required_phase5_files"]
    for name, expected_hash in required.items():
        path = phase5_root / name
        if not path.exists() or sums.get(name) != expected_hash or sha256_file(path) != expected_hash:
            raise AssertionError(f"Signed Phase 5 index file changed: {name}")
    return {"manifest": manifest, "success": success, "sums": sums}


def _rows_from_sheet(sheet) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(value not in (None, "") for value in row)]


def load_approved_opportunity(
    phase6_root: Path,
    phase6_config: dict,
    phase7_config: dict,
) -> dict:
    verification = verify_phase6_run(phase6_root, phase6_config)
    manifest = verification["manifest"]
    requirements = [
        json.loads(line)
        for line in (phase6_root / "requirements.jsonl").read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    proposals = json.loads((phase6_root / "filter_proposals.json").read_text(encoding="utf-8"))
    if manifest["input_mode"] == "SAMPLE_REDACTED":
        if not phase7_config["review"]["synthetic_sample_auto_approval_allowed"]:
            raise RecommendationError("Synthetic development approval is disabled")
        approved_requirements = [dict(row, approval_source="SYNTHETIC_DEVELOPMENT_FIXTURE") for row in requirements]
        approved_filters = [dict(row, approval_source="SYNTHETIC_DEVELOPMENT_FIXTURE") for row in proposals]
        approval_mode = "APPROVED_SYNTHETIC_DEVELOPMENT_FIXTURE"
    else:
        import openpyxl

        workbook_path = phase6_root / "OPPORTUNITY_REVIEW.xlsx"
        workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
        requirement_reviews = {row["requirement_id"]: row for row in _rows_from_sheet(workbook["Requirements"])}
        filter_reviews = {row["filter_id"]: row for row in _rows_from_sheet(workbook["Filters"])}
        approved_requirements = []
        for requirement in requirements:
            review = requirement_reviews.get(requirement["requirement_id"])
            decision = str((review or {}).get("reviewer_decision") or "PENDING").strip().upper()
            if decision == "PENDING":
                raise RecommendationError("Every real requirement must be reviewed before retrieval")
            if decision == "REJECT":
                continue
            if decision not in {"APPROVE", "EDIT"}:
                raise RecommendationError(f"Unknown requirement decision: {decision}")
            updated = dict(requirement)
            if decision == "EDIT":
                correction = " ".join(str(review.get("reviewer_correction") or "").split())
                if not correction:
                    raise RecommendationError("Edited requirement is missing reviewer_correction")
                updated["requirement_text"] = correction
            updated["approval_source"] = "HUMAN_REVIEW_WORKBOOK"
            approved_requirements.append(updated)
        approved_filters = []
        for proposal in proposals:
            review = filter_reviews.get(proposal["filter_id"])
            decision = str((review or {}).get("reviewer_decision") or "PENDING").strip().upper()
            if decision == "PENDING":
                raise RecommendationError("Every real filter proposal must be reviewed before retrieval")
            if decision == "REJECT":
                continue
            if decision not in {"APPROVE", "EDIT"}:
                raise RecommendationError(f"Unknown filter decision: {decision}")
            updated = dict(proposal)
            if decision == "EDIT":
                value = review.get("reviewer_value")
                behavior = str(review.get("reviewer_behavior") or "").strip()
                if value in (None, "") or not behavior:
                    raise RecommendationError("Edited filter requires reviewer_value and reviewer_behavior")
                updated["value"] = value
                updated["proposed_behavior"] = behavior
            updated["approval_source"] = "HUMAN_REVIEW_WORKBOOK"
            approved_filters.append(updated)
        approval_mode = "APPROVED_HUMAN_REVIEW_WORKBOOK"
    if not approved_requirements:
        raise RecommendationError("No approved requirements remain")
    hard_filters: dict[str, Any] = {}
    soft_preferences: list[dict] = []
    exclusions: list[dict] = []
    for proposal in approved_filters:
        behavior = proposal["proposed_behavior"]
        if behavior == "HARD_CANDIDATE":
            field, value = proposal["field"], proposal["value"]
            existing = hard_filters.get(field)
            if field == "year_after":
                hard_filters[field] = max(int(existing), int(value)) if existing is not None else int(value)
                continue
            if field == "year_before":
                hard_filters[field] = min(int(existing), int(value)) if existing is not None else int(value)
                continue
            if existing is None:
                hard_filters[field] = value
            elif isinstance(existing, list):
                if value not in existing:
                    existing.append(value)
            elif existing != value:
                hard_filters[field] = [existing, value]
        elif behavior == "SOFT_PREFERENCE":
            soft_preferences.append(proposal)
        elif behavior == "EXCLUSION_CANDIDATE":
            exclusions.append(proposal)
    return {
        "phase6_manifest": manifest,
        "phase6_manifest_sha256": sha256_file(phase6_root / "PHASE_6_MANIFEST.json"),
        "requirements": approved_requirements,
        "filters": approved_filters,
        "hard_filters": hard_filters,
        "soft_preferences": soft_preferences,
        "exclusions": exclusions,
        "approval_mode": approval_mode,
    }


@dataclass
class SecureBM25Retriever:
    chunks: pd.DataFrame
    bm25: BM25Index
    phase5_config: dict

    def __post_init__(self) -> None:
        self.chunks = self.chunks.reset_index(drop=True)
        if self.bm25.document_count != len(self.chunks):
            raise AssertionError("BM25/chunk row alignment failed")
        if self.chunks["vector_row"].astype(int).tolist() != list(range(len(self.chunks))):
            raise AssertionError("Chunk vector rows are not stable and contiguous")
        self.filter_engine = FilterEngine(self.chunks, self.phase5_config)
        self.tie_ids = self.chunks["chunk_id"].astype(str).tolist()

    @classmethod
    def load(cls, phase5_root: Path, phase5_config: dict) -> "SecureBM25Retriever":
        chunks = pd.read_parquet(phase5_root / "chunk_lookup.parquet")
        bm25 = BM25Index.load(phase5_root / "bm25_index.npz", phase5_root / "bm25_vocabulary.json")
        return cls(chunks=chunks, bm25=bm25, phase5_config=phase5_config)

    def search(
        self,
        query: str,
        *,
        allowed_security_classifications: Iterable[str],
        hard_filters: dict[str, Any],
        top_k: int,
        minimum_score: float,
    ) -> pd.DataFrame:
        mask = self.filter_engine.mask(
            allowed_security_classifications=allowed_security_classifications,
            hard_filters=hard_filters,
        )
        if not mask.any():
            return pd.DataFrame()
        scores = self.bm25.score(query, allowed_mask=mask)
        rows = [index for index in np.flatnonzero(mask) if float(scores[index]) > minimum_score]
        rows.sort(key=lambda index: (-float(scores[index]), self.tie_ids[index]))
        records = []
        for rank, index in enumerate(rows[:top_k], start=1):
            record = self.chunks.iloc[index].to_dict()
            record.update({"retrieval_rank": rank, "bm25_score": float(scores[index])})
            records.append(record)
        return pd.DataFrame(records)


def _reference_ids_by_document(documents: pd.DataFrame) -> dict[str, list[str]]:
    return {
        str(row["document_id"]): _json_list(row["reference_ids_json"])
        for row in documents.to_dict(orient="records")
    }


def eligible_reference_ids(references: pd.DataFrame, hard_filters: dict[str, Any]) -> set[str]:
    """Reapply compatible business filters to canonical reference metadata.

    Page/chunk filtering remains mandatory before scoring. This second gate prevents
    a multi-reference source document from transferring one reference's metadata to
    another reference linked to the same file.
    """
    mask = pd.Series(True, index=references.index, dtype=bool)
    exact_fields = {
        "country",
        "business_unit",
        "client",
        "sector",
        "service_nature",
        "offering",
        "project_year",
        "attestation_available",
        "data_quality_status",
    }
    for field, requested in hard_filters.items():
        if field in exact_fields and field in references.columns:
            wanted = requested if isinstance(requested, list) else [requested]
            normalized = {normalize_search_text(value) for value in wanted if str(value).strip()}
            if normalized:
                mask &= references[field].map(normalize_search_text).isin(normalized)
        elif field in {"year_after", "year_before"}:
            threshold = int(requested)
            years = references["project_year"].map(
                lambda value: [int(year) for year in YEAR_RE.findall(str(value or ""))]
            )
            if field == "year_after":
                mask &= years.map(lambda values: bool(values) and any(year >= threshold for year in values))
            else:
                mask &= years.map(lambda values: bool(values) and any(year <= threshold for year in values))
    return set(references.loc[mask, "reference_id"].astype(str))


def retrieve_requirement_evidence(
    *,
    review: dict,
    retriever: SecureBM25Retriever,
    documents: pd.DataFrame,
    references: pd.DataFrame,
    config: dict,
) -> pd.DataFrame:
    reference_ids = eligible_reference_ids(references, review["hard_filters"])
    document_to_references = _reference_ids_by_document(documents)
    rows: list[dict] = []
    for requirement in review["requirements"]:
        results = retriever.search(
            requirement["requirement_text"],
            allowed_security_classifications=config["security"]["allowed_classifications_for_pilot"],
            hard_filters=review["hard_filters"],
            top_k=int(config["retrieval"]["candidate_chunks_per_requirement"]),
            minimum_score=float(config["retrieval"]["minimum_bm25_score"]),
        )
        if results.empty:
            continue
        maximum = max(float(results["bm25_score"].max()), 1e-9)
        best: dict[str, dict] = {}
        for result in results.to_dict(orient="records"):
            normalized_score = min(1.0, float(result["bm25_score"]) / maximum)
            for reference_id in document_to_references.get(str(result["document_id"]), []):
                if reference_id not in reference_ids:
                    continue
                record = {
                    "requirement_id": requirement["requirement_id"],
                    "requirement_classification": requirement["classification"],
                    "requirement_text": requirement["requirement_text"],
                    "reference_id": reference_id,
                    "document_id": str(result["document_id"]),
                    "document_type": str(result["document_type"]),
                    "data_quality_status": str(result["data_quality_status"]),
                    "chunk_id": str(result["chunk_id"]),
                    "chunk_text_sha256": str(result["chunk_text_sha256"]),
                    "source_sha256": str(result["source_sha256"]),
                    "source_file_name": str(result["source_file_name"]),
                    "page_number_1_based": int(result["page_number_1_based"]),
                    "citation_label": str(result["citation_label"]),
                    "citation_uri": str(result["citation_uri"]),
                    "evidence_excerpt": str(result["chunk_text"])[:700].strip(),
                    "bm25_score": float(result["bm25_score"]),
                    "requirement_relevance_score": normalized_score,
                    "retrieval_rank": int(result["retrieval_rank"]),
                }
                previous = best.get(reference_id)
                if (
                    previous is None
                    or record["bm25_score"] > previous["bm25_score"]
                    or (
                        record["bm25_score"] == previous["bm25_score"]
                        and record["chunk_id"] < previous["chunk_id"]
                    )
                ):
                    best[reference_id] = record
        rows.extend(best.values())
    evidence = pd.DataFrame(rows)
    if not evidence.empty:
        evidence = evidence.sort_values(
            ["requirement_id", "requirement_relevance_score", "reference_id"],
            ascending=[True, False, True],
        ).reset_index(drop=True)
    return evidence


def _recency_score(value: Any, reference_year: int) -> tuple[int | None, float]:
    years = [int(year) for year in YEAR_RE.findall(str(value or ""))]
    if not years:
        return None, 0.25
    latest = min(max(years), reference_year)
    return latest, max(0.0, 1.0 - (reference_year - latest) / 20.0)


def _soft_preference_score(reference: dict, preferences: list[dict]) -> float:
    if not preferences:
        return 1.0
    mapping = {
        "country": "country",
        "business_unit": "business_unit",
        "client": "client",
        "sector": "sector",
        "service_nature": "service_nature",
        "offering": "offering",
        "project_year": "project_year",
        "attestation_available": "attestation_available",
        "data_quality_status": "data_quality_status",
    }
    matches = 0
    applicable = 0
    for preference in preferences:
        column = mapping.get(preference["field"])
        if not column:
            continue
        applicable += 1
        if normalize_search_text(preference["value"]) in normalize_search_text(reference.get(column, "")):
            matches += 1
    return matches / applicable if applicable else 1.0


def score_references(
    *,
    review: dict,
    evidence: pd.DataFrame,
    references: pd.DataFrame,
    config: dict,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    requirement_weights = config["scoring"]["requirement_weights"]
    component_weights = config["scoring"]["component_weights"]
    coverage_threshold = float(config["scoring"]["coverage_threshold"])
    evidence_strengths = config["scoring"]["document_type_strength"]
    quality_strengths = config["scoring"]["data_quality_strength"]
    requirements = review["requirements"]
    total_weight = sum(float(requirement_weights[row["classification"]]) for row in requirements)
    reference_map = references.set_index("reference_id").to_dict(orient="index")
    evidence_lookup = {
        (str(row["reference_id"]), str(row["requirement_id"])): row
        for row in evidence.to_dict(orient="records")
    }
    candidate_ids = sorted(set(evidence["reference_id"].astype(str))) if not evidence.empty else []
    recommendation_rows = []
    coverage_rows = []
    for reference_id in candidate_ids:
        metadata = dict(reference_map[reference_id])
        relevance_sum = covered_weight = 0.0
        must_total = must_covered = 0
        reference_evidence = []
        for requirement in requirements:
            weight = float(requirement_weights[requirement["classification"]])
            hit = evidence_lookup.get((reference_id, requirement["requirement_id"]))
            score = float(hit["requirement_relevance_score"]) if hit else 0.0
            covered = bool(hit) and score >= coverage_threshold
            relevance_sum += weight * score
            if covered:
                covered_weight += weight
            if requirement["classification"] == "MUST":
                must_total += 1
                must_covered += int(covered)
            coverage_rows.append(
                {
                    "reference_id": reference_id,
                    "requirement_id": requirement["requirement_id"],
                    "classification": requirement["classification"],
                    "requirement_text": requirement["requirement_text"],
                    "covered": covered,
                    "relevance_score": score,
                    "citation_uri": hit["citation_uri"] if hit else "",
                    "chunk_id": hit["chunk_id"] if hit else "",
                }
            )
            if hit:
                reference_evidence.append(hit)
        relevance = relevance_sum / total_weight
        coverage = covered_weight / total_weight
        evidence_quality = max(
            (
                float(evidence_strengths.get(str(row["document_type"]), evidence_strengths["OTHER"]))
                * float(quality_strengths.get(str(row["data_quality_status"]), quality_strengths["OTHER"]))
                for row in reference_evidence
            ),
            default=0.0,
        )
        latest_year, recency = _recency_score(metadata.get("project_year"), int(config["scoring"]["reference_year"]))
        soft_fit = _soft_preference_score(metadata, review["soft_preferences"])
        score = (
            component_weights["relevance"] * relevance
            + component_weights["coverage"] * coverage
            + component_weights["evidence_quality"] * evidence_quality
            + component_weights["recency"] * recency
            + component_weights["soft_preference_fit"] * soft_fit
        )
        warnings = []
        if must_covered < must_total:
            warnings.append(f"MUST coverage {must_covered}/{must_total}")
        if not bool(metadata.get("evidence_available")):
            warnings.append("missing supporting evidence")
        recommendation_rows.append(
            {
                "reference_id": reference_id,
                "base_score": round(float(score), 8),
                "weighted_relevance": round(float(relevance), 8),
                "weighted_coverage": round(float(coverage), 8),
                "must_covered": must_covered,
                "must_total": must_total,
                "evidence_quality": round(float(evidence_quality), 8),
                "latest_project_year": latest_year,
                "recency_score": round(float(recency), 8),
                "soft_preference_fit": round(float(soft_fit), 8),
                "client": str(metadata.get("client") or ""),
                "country": str(metadata.get("country") or ""),
                "sector": str(metadata.get("sector") or ""),
                "offering": str(metadata.get("offering") or ""),
                "service_nature": str(metadata.get("service_nature") or ""),
                "project_year": str(metadata.get("project_year") or ""),
                "attestation_available": str(metadata.get("attestation_available") or ""),
                "evidence_available": bool(metadata.get("evidence_available")),
                "data_quality_status": str(metadata.get("data_quality_status") or ""),
                "warnings": "; ".join(warnings),
                "recommendation_basis": (
                    f"Deterministic evidence score; covered {must_covered}/{must_total} MUST requirements."
                ),
            }
        )
    recommendations = pd.DataFrame(recommendation_rows)
    if not recommendations.empty:
        recommendations = recommendations.sort_values(
            ["base_score", "reference_id"], ascending=[False, True]
        ).reset_index(drop=True)
        recommendations["base_rank"] = np.arange(1, len(recommendations) + 1)
    return recommendations, pd.DataFrame(coverage_rows)


def diversify_recommendations(recommendations: pd.DataFrame, config: dict) -> pd.DataFrame:
    if recommendations.empty:
        return recommendations.copy()
    limit = int(config["output"]["maximum_recommendations"])
    max_client = int(config["scoring"]["diversity_max_per_client"])
    selected: list[dict] = []
    deferred: list[dict] = []
    counts: Counter[str] = Counter()
    for row in recommendations.to_dict(orient="records"):
        client_key = normalize_search_text(row["client"]) or f"unknown:{row['reference_id']}"
        if counts[client_key] < max_client and len(selected) < limit:
            selected.append(row)
            counts[client_key] += 1
        else:
            deferred.append(row)
    for row in deferred:
        if len(selected) == limit:
            break
        selected.append(row)
    for rank, row in enumerate(selected, start=1):
        row["final_rank"] = rank
    return pd.DataFrame(selected)


def requirement_gap_summary(
    requirements: list[dict], evidence: pd.DataFrame, threshold: float
) -> list[dict]:
    rows = []
    for requirement in requirements:
        matches = evidence.loc[evidence["requirement_id"].eq(requirement["requirement_id"])] if not evidence.empty else pd.DataFrame()
        qualifying = matches.loc[matches["requirement_relevance_score"].ge(threshold)] if not matches.empty else pd.DataFrame()
        rows.append(
            {
                "requirement_id": requirement["requirement_id"],
                "classification": requirement["classification"],
                "requirement_text": requirement["requirement_text"],
                "references_with_evidence": int(qualifying["reference_id"].nunique()) if not qualifying.empty else 0,
                "gap": qualifying.empty,
                "status": "NO_RELIABLE_EVIDENCE_FOUND" if qualifying.empty else "EVIDENCE_FOUND",
            }
        )
    return rows


def validate_evidence(evidence: pd.DataFrame) -> float:
    if evidence.empty:
        return 0.0
    required = [
        "reference_id",
        "requirement_id",
        "document_id",
        "chunk_id",
        "chunk_text_sha256",
        "source_sha256",
        "page_number_1_based",
        "citation_label",
        "citation_uri",
        "evidence_excerpt",
    ]
    missing = set(required) - set(evidence.columns)
    if missing:
        raise AssertionError(f"Evidence fields missing: {sorted(missing)}")
    complete = evidence[required].apply(
        lambda column: column.map(lambda value: value not in (None, "") and not pd.isna(value))
    ).all(axis=1)
    coverage = float(complete.mean())
    if coverage != 1.0:
        raise AssertionError("Every evidence row must have complete provenance and citation")
    return coverage


def create_recommendation_workbook(
    path: Path,
    recommendations: pd.DataFrame,
    coverage: pd.DataFrame,
    evidence: pd.DataFrame,
    gaps: list[dict],
    manifest: dict,
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = openpyxl.Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Control", "Value"])
    for row in (
        ("Status", manifest["status"]),
        ("Opportunity ID", manifest["opportunity_id"]),
        ("Retrieval mode", manifest["retrieval_mode"]),
        ("Recommendations", manifest["recommendations"]),
        ("Requirement gaps", manifest["requirement_gaps"]),
        ("Citation coverage", manifest["citation_coverage"]),
        ("External LLM calls", manifest["external_llm_calls"]),
        ("Cross-encoder calls", manifest["cross_encoder_calls"]),
        ("Decision rule", "A human must approve the shortlist before template generation."),
    ):
        overview.append(row)
    ranked = workbook.create_sheet("Ranked References")
    ranked_headers = [
        "final_rank", "reference_id", "base_score", "must_covered", "must_total",
        "weighted_coverage", "client", "country", "sector", "offering",
        "project_year", "attestation_available", "warnings", "reviewer_decision", "reviewer_notes",
    ]
    ranked.append(ranked_headers)
    for row in recommendations.to_dict(orient="records"):
        ranked.append([row.get(header, "") for header in ranked_headers[:-2]] + ["PENDING", ""])
    coverage_sheet = workbook.create_sheet("Requirement Coverage")
    coverage_headers = [
        "reference_id", "requirement_id", "classification", "requirement_text",
        "covered", "relevance_score", "citation_uri", "chunk_id",
    ]
    coverage_sheet.append(coverage_headers)
    for row in coverage.to_dict(orient="records"):
        coverage_sheet.append([row.get(header, "") for header in coverage_headers])
    evidence_sheet = workbook.create_sheet("Evidence")
    evidence_headers = [
        "reference_id", "requirement_id", "source_file_name", "page_number_1_based",
        "evidence_excerpt", "citation_label", "citation_uri", "chunk_id",
        "source_sha256", "chunk_text_sha256",
    ]
    evidence_sheet.append(evidence_headers)
    selected_ids = set(recommendations["reference_id"].astype(str)) if not recommendations.empty else set()
    for row in evidence.loc[evidence["reference_id"].astype(str).isin(selected_ids)].to_dict(orient="records"):
        evidence_sheet.append([row.get(header, "") for header in evidence_headers])
    gap_sheet = workbook.create_sheet("Requirement Gaps")
    gap_headers = ["requirement_id", "classification", "requirement_text", "references_with_evidence", "gap", "status"]
    gap_sheet.append(gap_headers)
    for row in gaps:
        gap_sheet.append([row.get(header, "") for header in gap_headers])
    decision = DataValidation(type="list", formula1='"PENDING,SHORTLIST,REJECT"')
    ranked.add_data_validation(decision)
    decision.add(f"N2:N{max(2, ranked.max_row)}")
    widths = {
        "Overview": [28, 100],
        "Ranked References": [12, 24, 14, 14, 12, 18, 35, 18, 18, 24, 15, 25, 45, 20, 45],
        "Requirement Coverage": [24, 22, 15, 75, 12, 16, 65, 28],
        "Evidence": [24, 22, 35, 14, 85, 45, 65, 28, 68, 68],
        "Requirement Gaps": [22, 15, 80, 24, 12, 30],
    }
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B2C83")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, width in enumerate(widths[sheet.title], start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _build_report(
    manifest: dict,
    recommendations: pd.DataFrame,
    gaps: list[dict],
    evidence: pd.DataFrame,
) -> str:
    lines = [
        "# Phase 7 — Evidence-backed recommendation report",
        "",
        f"**Status:** {manifest['status']}  ",
        f"**Opportunity:** `{manifest['opportunity_id']}`  ",
        f"**Retrieval mode:** `{manifest['retrieval_mode']}`  ",
        "",
        "## Ranked references",
        "",
    ]
    for row in recommendations.head(5).to_dict(orient="records"):
        lines.append(
            f"{int(row['final_rank'])}. **{row['client'] or row['reference_id']}** — "
            f"score {row['base_score']:.3f}; MUST {int(row['must_covered'])}/{int(row['must_total'])}; "
            f"{row['country']}; {row['offering']}."
        )
        citations = evidence.loc[evidence["reference_id"].eq(row["reference_id"]), "citation_uri"].drop_duplicates().head(3).tolist()
        for citation in citations:
            lines.append(f"   - Evidence: {citation}")
        if row["warnings"]:
            lines.append(f"   - Warning: {row['warnings']}")
    lines.extend(["", "## Requirement gaps", ""])
    gap_rows = [row for row in gaps if row["gap"]]
    if gap_rows:
        for row in gap_rows:
            lines.append(f"- `{row['requirement_id']}` ({row['classification']}): {row['status']}")
    else:
        lines.append("- No complete-corpus evidence gap was detected at the configured technical threshold.")
    lines.extend(
        [
            "",
            "## Decision boundary",
            "",
            "This is a deterministic technical baseline. It does not use an LLM or cross-encoder,",
            "does not replace expert retrieval evaluation, and does not authorize external use.",
            "A business reviewer must approve the shortlist before template generation.",
            "",
        ]
    )
    return "\n".join(lines)


def run_phase7(
    *,
    phase5_root: Path,
    phase4_root: Path,
    phase6_root: Path,
    phase5_config: dict,
    phase6_config: dict,
    phase7_config: dict,
    output_root: Path,
) -> tuple[Path, dict]:
    verify_phase5_index(phase5_root, phase7_config)
    review = load_approved_opportunity(phase6_root, phase6_config, phase7_config)
    documents_path = phase4_root / "documents_catalog.parquet"
    references_path = phase4_root / "reference_catalog.parquet"
    if sha256_file(documents_path) != phase7_config["input"]["expected_documents_sha256"]:
        raise AssertionError("Phase 4 document catalogue changed")
    if sha256_file(references_path) != phase7_config["input"]["expected_references_sha256"]:
        raise AssertionError("Phase 4 reference catalogue changed")
    documents = pd.read_parquet(documents_path)
    references = pd.read_parquet(references_path)
    retriever = SecureBM25Retriever.load(phase5_root, phase5_config)
    evidence = retrieve_requirement_evidence(
        review=review,
        retriever=retriever,
        documents=documents,
        references=references,
        config=phase7_config,
    )
    if evidence.empty:
        raise RecommendationError("No evidence survived authorization, hard filters, and BM25 scoring")
    citation_coverage = validate_evidence(evidence)
    recommendations, coverage = score_references(
        review=review,
        evidence=evidence,
        references=references,
        config=phase7_config,
    )
    final_recommendations = diversify_recommendations(recommendations, phase7_config)
    gaps = requirement_gap_summary(
        review["requirements"], evidence, float(phase7_config["scoring"]["coverage_threshold"])
    )
    opportunity_id = review["phase6_manifest"]["opportunity_id"]
    run_root = output_root / opportunity_id / phase7_config["pipeline_version"]
    success_path = run_root / phase7_config["output"]["success_marker"]
    if success_path.exists():
        verify_phase7_run(run_root, phase7_config)
        return run_root, json.loads((run_root / "PHASE_7_MANIFEST.json").read_text(encoding="utf-8"))
    run_root.mkdir(parents=True, exist_ok=True)
    manifest = {
        "schema_version": 1,
        "phase": 7,
        "pipeline_version": phase7_config["pipeline_version"],
        "completed_at_utc": utc_now(),
        "status": (
            "TECHNICAL_PASS_SAMPLE_ONLY"
            if review["phase6_manifest"]["input_mode"] == "SAMPLE_REDACTED"
            else "READY_FOR_BUSINESS_SHORTLIST_REVIEW"
        ),
        "opportunity_id": opportunity_id,
        "phase6_manifest_sha256": review["phase6_manifest_sha256"],
        "approval_mode": review["approval_mode"],
        "retrieval_mode": phase7_config["retrieval"]["mode"],
        "approved_requirements": len(review["requirements"]),
        "approved_hard_filters": len(review["hard_filters"]),
        "recommendations": int(len(final_recommendations)),
        "evidence_rows": int(len(evidence)),
        "requirement_gaps": sum(int(row["gap"]) for row in gaps),
        "citation_coverage": citation_coverage,
        "external_llm_calls": 0,
        "external_embedding_api_calls": 0,
        "dense_query_embedding_calls": 0,
        "cross_encoder_calls": 0,
        "raw_opportunity_log_rows": 0,
        "security_filters_applied_before_scoring": True,
        "business_shortlist_auto_approved": False,
        "production_promotion_allowed": False,
        "phase5_1_expert_evaluation_bypassed": False,
    }
    final_recommendations.to_parquet(run_root / "recommendations.parquet", index=False)
    final_recommendations.to_json(
        run_root / "recommendations.json", orient="records", force_ascii=False, indent=2
    )
    evidence.to_parquet(run_root / "evidence_matrix.parquet", index=False)
    evidence.to_json(run_root / "evidence_matrix.jsonl", orient="records", lines=True, force_ascii=False)
    coverage.to_csv(run_root / "requirement_coverage.csv", index=False, encoding="utf-8-sig")
    _atomic_json(run_root / "requirement_gaps.json", gaps)
    manifest_path = run_root / "PHASE_7_MANIFEST.json"
    _atomic_json(manifest_path, manifest)
    create_recommendation_workbook(
        run_root / "RECOMMENDATION_REVIEW.xlsx",
        final_recommendations,
        coverage,
        evidence,
        gaps,
        manifest,
    )
    _atomic_text(run_root / "PHASE_7_REPORT.md", _build_report(manifest, final_recommendations, gaps, evidence))
    hashed = [
        path
        for path in run_root.iterdir()
        if path.is_file() and path.name not in {"SHA256SUMS.txt", phase7_config["output"]["success_marker"]}
    ]
    sums = "".join(
        f"{sha256_file(path)}  {path.name}\n" for path in sorted(hashed, key=lambda item: item.name)
    )
    sums_path = run_root / "SHA256SUMS.txt"
    _atomic_text(sums_path, sums)
    _atomic_json(
        success_path,
        {
            "status": "COMPLETE_REPRODUCIBLE_EVIDENCE_RECOMMENDATIONS",
            "created_at_utc": utc_now(),
            "pipeline_version": phase7_config["pipeline_version"],
            "opportunity_id": opportunity_id,
            "manifest_sha256": sha256_file(manifest_path),
            "sha256sums_sha256": sha256_file(sums_path),
        },
    )
    verify_phase7_run(run_root, phase7_config)
    return run_root, manifest


def verify_phase7_run(run_root: Path, config: dict) -> dict:
    success_path = run_root / config["output"]["success_marker"]
    manifest_path = run_root / "PHASE_7_MANIFEST.json"
    sums_path = run_root / "SHA256SUMS.txt"
    for path in (success_path, manifest_path, sums_path):
        if not path.exists():
            raise FileNotFoundError(path)
    success = json.loads(success_path.read_text(encoding="utf-8"))
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if success["manifest_sha256"] != sha256_file(manifest_path):
        raise AssertionError("Phase 7 manifest hash mismatch")
    if success["sha256sums_sha256"] != sha256_file(sums_path):
        raise AssertionError("Phase 7 checksum-file hash mismatch")
    for line in sums_path.read_text(encoding="utf-8").splitlines():
        expected, name = line.split("  ", 1)
        if sha256_file(run_root / name) != expected:
            raise AssertionError(f"Phase 7 output changed: {name}")
    prohibited_nonzero = (
        "external_llm_calls",
        "external_embedding_api_calls",
        "dense_query_embedding_calls",
        "cross_encoder_calls",
        "raw_opportunity_log_rows",
    )
    if any(int(manifest.get(field, -1)) != 0 for field in prohibited_nonzero):
        raise AssertionError("Phase 7 provider/security boundary failed")
    if not manifest["security_filters_applied_before_scoring"]:
        raise AssertionError("Security filtering order failed")
    if manifest["business_shortlist_auto_approved"] or manifest["production_promotion_allowed"]:
        raise AssertionError("Phase 7 human/promotion boundary failed")
    if manifest["citation_coverage"] != 1.0:
        raise AssertionError("Phase 7 citation coverage is incomplete")
    return {"success": success, "manifest": manifest}
