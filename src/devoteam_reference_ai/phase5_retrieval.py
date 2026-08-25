from __future__ import annotations

import hashlib
import importlib.metadata
import json
import math
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
import yaml

from .phase2_utils import sha256_file
from .phase4_corpus import verify_phase4
from .phase5_bm25 import BM25Index, normalize_search_text


SAFE_CONTEXT_FIELDS = (
    ("Client", "client_values_json"),
    ("Pays", "country_values_json"),
    ("Secteur", "sector_values_json"),
    ("Offre", "offering_values_json"),
    ("Prestation", "service_nature_values_json"),
    ("Année", "project_year_values_json"),
)
YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b")


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _atomic_json(path: Path, value: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8"
    )
    os.replace(temporary, path)


def load_phase5_config(path: Path) -> dict:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    if int(config.get("phase", 0)) != 5:
        raise ValueError("Expected Phase 5 configuration")
    security = config.get("security", {})
    if security.get("external_llm_enabled") or security.get("external_embedding_api_enabled"):
        raise ValueError("Phase 5 permits local embedding execution only")
    if not config.get("embedding", {}).get("local_execution_only"):
        raise ValueError("Embedding model must execute locally")
    if config.get("filters", {}).get("security_filters_disableable"):
        raise ValueError("Security filters may not be disableable")
    if not config.get("evaluation", {}).get("expert_gold_set_required_for_promotion"):
        raise ValueError("Expert labels must remain a production-promotion gate")
    return config


def _json_values(value: Any) -> list[str]:
    if isinstance(value, list):
        raw = value
    elif value is None or (isinstance(value, float) and math.isnan(value)):
        raw = []
    else:
        try:
            parsed = json.loads(str(value))
            raw = parsed if isinstance(parsed, list) else [parsed]
        except (TypeError, ValueError, json.JSONDecodeError):
            raw = [value]
    output: list[str] = []
    seen: set[str] = set()
    for item in raw:
        cleaned = " ".join(str(item or "").split()).strip()
        key = normalize_search_text(cleaned)
        if cleaned and key not in seen:
            output.append(cleaned)
            seen.add(key)
    return output


def build_retrieval_texts(chunks: pd.DataFrame) -> list[str]:
    required = {"source_file_name", "chunk_text", *(column for _, column in SAFE_CONTEXT_FIELDS)}
    missing = required - set(chunks.columns)
    if missing:
        raise AssertionError(f"Chunk corpus missing retrieval context fields: {sorted(missing)}")
    texts: list[str] = []
    for row in chunks.to_dict(orient="records"):
        context = [f"Document: {row['source_file_name']}"]
        for label, column in SAFE_CONTEXT_FIELDS:
            values = _json_values(row[column])
            if values:
                # Metadata is retrieval context, not a replacement for source
                # evidence. Bound it so the 512-token model always sees the
                # complete <=900-character evidence chunk.
                bounded = [value[:120] for value in values[:3]]
                context.append(f"{label}: {' | '.join(bounded)}"[:200])
        metadata_context = "\n".join(context)[:700].rstrip()
        text = f"{metadata_context}\n{str(row['chunk_text']).strip()}".strip()
        if not text:
            raise AssertionError("A retrieval passage is empty")
        if len(text) > 1700:
            raise AssertionError("Bounded retrieval passage unexpectedly exceeds 1,700 characters")
        texts.append(text)
    return texts


class FilterEngine:
    def __init__(self, chunks: pd.DataFrame, config: dict):
        self.chunks = chunks.reset_index(drop=True)
        self.config = config
        self.exact_columns = config["filters"]["supported_exact"]
        self.range_columns = config["filters"]["supported_ranges"]
        self.security_field = config["filters"]["security_field"]
        required = {self.security_field, *self.exact_columns.values(), *self.range_columns.values()}
        missing = required - set(chunks.columns)
        if missing:
            raise AssertionError(f"Filter columns missing from chunks: {sorted(missing)}")
        self.cache: dict[str, list[set[str]]] = {}
        for name, column in self.exact_columns.items():
            if column.endswith("_json"):
                self.cache[name] = [
                    {normalize_search_text(value) for value in _json_values(raw)}
                    for raw in self.chunks[column]
                ]
            else:
                self.cache[name] = [
                    {normalize_search_text(raw)} if str(raw or "").strip() else set()
                    for raw in self.chunks[column]
                ]
        year_column = next(iter(set(self.range_columns.values())))
        self.year_cache = [
            {int(year) for value in _json_values(raw) for year in YEAR_RE.findall(value)}
            for raw in self.chunks[year_column]
        ]

    def mask(
        self,
        *,
        allowed_security_classifications: Iterable[str] | None,
        hard_filters: dict[str, Any] | None = None,
    ) -> np.ndarray:
        allowed = {
            normalize_search_text(value)
            for value in (allowed_security_classifications or [])
            if str(value).strip()
        }
        if not allowed:
            raise PermissionError("Explicit security authorization is required before retrieval")
        mask = self.chunks[self.security_field].map(
            lambda value: normalize_search_text(value) in allowed
        ).to_numpy(dtype=bool)
        filters = hard_filters or {}
        supported = set(self.exact_columns) | set(self.range_columns)
        unknown = set(filters) - supported
        if unknown:
            raise ValueError(f"Unsupported hard filters: {sorted(unknown)}")
        for name, requested in filters.items():
            if requested is None or requested == "" or requested == []:
                continue
            if name in self.exact_columns:
                values = requested if isinstance(requested, list) else [requested]
                wanted = {normalize_search_text(value) for value in values if str(value).strip()}
                if wanted:
                    mask &= np.asarray(
                        [bool(row_values & wanted) for row_values in self.cache[name]], dtype=bool
                    )
            elif name == "year_before":
                threshold = int(requested)
                mask &= np.asarray(
                    [bool(years) and any(year <= threshold for year in years) for years in self.year_cache],
                    dtype=bool,
                )
            elif name == "year_after":
                threshold = int(requested)
                mask &= np.asarray(
                    [bool(years) and any(year >= threshold for year in years) for years in self.year_cache],
                    dtype=bool,
                )
        return mask


class E5EmbeddingAdapter:
    def __init__(self, config: dict, device: str | None = None):
        from sentence_transformers import SentenceTransformer

        settings = config["embedding"]
        self.settings = settings
        self.model = SentenceTransformer(
            settings["model_id"],
            revision=settings["revision"],
            trust_remote_code=bool(settings["trust_remote_code"]),
            device=device,
        )
        dimension = int(self.model.get_sentence_embedding_dimension())
        if dimension != int(settings["dimensions"]):
            raise AssertionError(f"Embedding dimension changed: {dimension}")
        self.device = str(self.model.device)

    def _batch_size(self) -> int:
        return int(
            self.settings["batch_size_gpu"]
            if self.device.startswith("cuda")
            else self.settings["batch_size_cpu"]
        )

    def encode_passages(self, texts: list[str]) -> np.ndarray:
        prefixed = [self.settings["passage_prefix"] + text for text in texts]
        values = self.model.encode(
            prefixed,
            batch_size=self._batch_size(),
            normalize_embeddings=bool(self.settings["normalize_embeddings"]),
            convert_to_numpy=True,
            show_progress_bar=False,
        )
        return np.asarray(values, dtype=np.float32)

    def encode_queries(self, texts: list[str]) -> np.ndarray:
        prefixed = [self.settings["query_prefix"] + text for text in texts]
        values = self.model.encode(
            prefixed,
            batch_size=self._batch_size(),
            normalize_embeddings=bool(self.settings["normalize_embeddings"]),
            convert_to_numpy=True,
            show_progress_bar=False,
        )
        return np.asarray(values, dtype=np.float32)


def build_embeddings_resumable(
    *,
    texts: list[str],
    adapter: Any,
    output_path: Path,
    progress_path: Path,
    config: dict,
    input_hash: str,
    progress=print,
) -> np.ndarray:
    dimension = int(config["embedding"]["dimensions"])
    row_count = len(texts)
    partial_path = output_path.with_name("embeddings.partial.npy")
    identity = {
        "pipeline_version": config["pipeline_version"],
        "input_chunks_sha256": input_hash,
        "model_id": config["embedding"]["model_id"],
        "model_revision": config["embedding"]["revision"],
        "rows": row_count,
        "dimensions": dimension,
    }
    if output_path.exists():
        values = np.load(output_path, mmap_mode="r")
        if values.shape != (row_count, dimension):
            raise AssertionError("Existing embedding matrix has the wrong shape")
        return values
    completed = 0
    if partial_path.exists() or progress_path.exists():
        if not partial_path.exists() or not progress_path.exists():
            raise AssertionError("Incomplete embedding checkpoint pair")
        state = json.loads(progress_path.read_text(encoding="utf-8"))
        for key, expected in identity.items():
            if state.get(key) != expected:
                raise AssertionError(f"Embedding checkpoint identity changed: {key}")
        completed = int(state.get("completed_rows", 0))
        matrix = np.lib.format.open_memmap(partial_path, mode="r+")
        if matrix.shape != (row_count, dimension):
            raise AssertionError("Partial embedding matrix has the wrong shape")
        progress(f"Resuming embeddings at row {completed}/{row_count}.")
    else:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        matrix = np.lib.format.open_memmap(
            partial_path, mode="w+", dtype=np.float32, shape=(row_count, dimension)
        )
        _atomic_json(progress_path, identity | {"completed_rows": 0, "status": "IN_PROGRESS"})
    checkpoint = int(config["embedding"]["checkpoint_every_rows"])
    batch_size = int(
        config["embedding"]["batch_size_gpu"]
        if str(getattr(adapter, "device", "cpu")).startswith("cuda")
        else config["embedding"]["batch_size_cpu"]
    )
    for start in range(completed, row_count, batch_size):
        end = min(start + batch_size, row_count)
        batch = np.asarray(adapter.encode_passages(texts[start:end]), dtype=np.float32)
        if batch.shape != (end - start, dimension):
            raise AssertionError("Embedding adapter returned the wrong shape")
        norms = np.linalg.norm(batch, axis=1)
        if not np.isfinite(batch).all() or not np.allclose(norms, 1.0, atol=2e-3):
            raise AssertionError("Passage embeddings are not finite unit vectors")
        matrix[start:end] = batch
        if end == row_count or end % checkpoint < batch_size:
            matrix.flush()
            _atomic_json(
                progress_path,
                identity | {"completed_rows": end, "status": "IN_PROGRESS", "updated_at_utc": _now()},
            )
            progress(f"Embedding checkpoint: {end}/{row_count} rows.")
    matrix.flush()
    del matrix
    os.replace(partial_path, output_path)
    _atomic_json(
        progress_path,
        identity
        | {
            "completed_rows": row_count,
            "status": "COMPLETE",
            "embedding_matrix_sha256": sha256_file(output_path),
            "completed_at_utc": _now(),
        },
    )
    return np.load(output_path, mmap_mode="r")


def write_faiss_index(embeddings: np.ndarray, path: Path) -> dict:
    import faiss

    values = np.ascontiguousarray(np.asarray(embeddings, dtype=np.float32))
    index = faiss.IndexFlatIP(values.shape[1])
    index.add(values)
    faiss.write_index(index, str(path))
    loaded = faiss.read_index(str(path))
    if int(loaded.ntotal) != values.shape[0] or int(loaded.d) != values.shape[1]:
        raise AssertionError("FAISS index verification failed")
    return {"index_type": "IndexFlatIP", "rows": int(loaded.ntotal), "dimensions": int(loaded.d)}


def _stable_top(scores: np.ndarray, mask: np.ndarray, tie_ids: list[str], limit: int) -> list[int]:
    candidates = [
        index for index in np.flatnonzero(mask) if np.isfinite(float(scores[index]))
    ]
    candidates.sort(key=lambda index: (-float(scores[index]), tie_ids[index]))
    return candidates[:limit]


@dataclass
class HybridSearchEngine:
    chunks: pd.DataFrame
    retrieval_texts: list[str]
    bm25: BM25Index
    embeddings: np.ndarray
    config: dict
    embedding_adapter: Any | None = None

    def __post_init__(self):
        self.chunks = self.chunks.reset_index(drop=True)
        self.embeddings = np.asarray(self.embeddings, dtype=np.float32)
        if len(self.chunks) != self.bm25.document_count or len(self.chunks) != len(self.embeddings):
            raise AssertionError("Hybrid retriever row alignment failed")
        if len(self.retrieval_texts) != len(self.chunks):
            raise AssertionError("Retrieval-text alignment failed")
        self.filter_engine = FilterEngine(self.chunks, self.config)
        self.tie_ids = self.chunks["chunk_id"].astype(str).tolist()

    def search_chunks(
        self,
        query: str,
        *,
        allowed_security_classifications: Iterable[str],
        hard_filters: dict[str, Any] | None = None,
        mode: str = "hybrid",
        top_k: int | None = None,
        query_vector: np.ndarray | None = None,
    ) -> pd.DataFrame:
        if mode not in {"bm25", "dense", "hybrid"}:
            raise ValueError(f"Unsupported retrieval mode: {mode}")
        top_k = int(top_k or self.config["hybrid"]["default_top_k"])
        candidate_pool = max(top_k, int(self.config["hybrid"]["candidate_pool_per_retriever"]))
        mask = self.filter_engine.mask(
            allowed_security_classifications=allowed_security_classifications,
            hard_filters=hard_filters,
        )
        if not mask.any():
            return pd.DataFrame()
        lexical_scores = self.bm25.score(query, allowed_mask=mask)
        dense_scores = np.full(len(self.chunks), -np.inf, dtype=np.float32)
        if mode in {"dense", "hybrid"}:
            if query_vector is None:
                if self.embedding_adapter is None:
                    raise ValueError("Dense retrieval requires an embedding adapter or query vector")
                query_vector = self.embedding_adapter.encode_queries([query])[0]
            query_vector = np.asarray(query_vector, dtype=np.float32).reshape(-1)
            if query_vector.shape != (self.embeddings.shape[1],):
                raise ValueError("Query vector has the wrong dimension")
            norm = float(np.linalg.norm(query_vector))
            if not np.isfinite(query_vector).all() or not math.isclose(norm, 1.0, abs_tol=2e-3):
                raise ValueError("Query vector must be a finite unit vector")
            allowed_rows = np.flatnonzero(mask)
            dense_scores[allowed_rows] = self.embeddings[allowed_rows] @ query_vector
        lexical_rows = _stable_top(lexical_scores, mask, self.tie_ids, candidate_pool)
        dense_rows = _stable_top(dense_scores, mask, self.tie_ids, candidate_pool)
        lexical_rank = {row: rank for rank, row in enumerate(lexical_rows, start=1)}
        dense_rank = {row: rank for rank, row in enumerate(dense_rows, start=1)}
        if mode == "bm25":
            ordered = lexical_rows[:top_k]
            retrieval_scores = {row: float(lexical_scores[row]) for row in ordered}
        elif mode == "dense":
            ordered = dense_rows[:top_k]
            retrieval_scores = {row: float(dense_scores[row]) for row in ordered}
        else:
            rrf_k = float(self.config["hybrid"]["rrf_k"])
            lexical_weight = float(self.config["hybrid"]["lexical_weight"])
            dense_weight = float(self.config["hybrid"]["dense_weight"])
            retrieval_scores = {}
            for row in set(lexical_rows) | set(dense_rows):
                score = 0.0
                if row in lexical_rank:
                    score += lexical_weight / (rrf_k + lexical_rank[row])
                if row in dense_rank:
                    score += dense_weight / (rrf_k + dense_rank[row])
                retrieval_scores[row] = score
            ordered = sorted(
                retrieval_scores,
                key=lambda row: (-retrieval_scores[row], self.tie_ids[row]),
            )[:top_k]
        records: list[dict] = []
        for rank, row_index in enumerate(ordered, start=1):
            record = self.chunks.iloc[row_index].to_dict()
            record.update(
                {
                    "vector_row": int(row_index),
                    "retrieval_mode": mode,
                    "rank": rank,
                    "retrieval_score": float(retrieval_scores[row_index]),
                    "bm25_score": float(lexical_scores[row_index]),
                    "dense_score": (
                        float(dense_scores[row_index]) if np.isfinite(dense_scores[row_index]) else None
                    ),
                    "bm25_rank": lexical_rank.get(row_index),
                    "dense_rank": dense_rank.get(row_index),
                }
            )
            records.append(record)
        return pd.DataFrame(records)

    def search_documents(
        self,
        query: str,
        *,
        allowed_security_classifications: Iterable[str],
        hard_filters: dict[str, Any] | None = None,
        mode: str = "hybrid",
        top_k: int = 10,
        query_vector: np.ndarray | None = None,
    ) -> pd.DataFrame:
        evidence_count = int(self.config["hybrid"]["document_evidence_chunks"])
        chunk_results = self.search_chunks(
            query,
            allowed_security_classifications=allowed_security_classifications,
            hard_filters=hard_filters,
            mode=mode,
            top_k=max(int(self.config["hybrid"]["candidate_pool_per_retriever"]), top_k * 5),
            query_vector=query_vector,
        )
        if chunk_results.empty:
            return pd.DataFrame()
        records: list[dict] = []
        for document_id, group in chunk_results.groupby("document_id", sort=False):
            ordered = group.sort_values(["retrieval_score", "chunk_id"], ascending=[False, True])
            evidence = ordered.head(evidence_count)
            best = ordered.iloc[0]
            records.append(
                {
                    "document_id": document_id,
                    "source_file_name": best["source_file_name"],
                    "document_type": best["document_type"],
                    "document_language": best["document_language"],
                    "data_quality_status": best["data_quality_status"],
                    "retrieval_mode": mode,
                    "document_score": float(best["retrieval_score"]),
                    "best_chunk_id": best["chunk_id"],
                    "evidence_chunk_ids_json": json.dumps(evidence["chunk_id"].tolist()),
                    "citation_labels_json": json.dumps(evidence["citation_label"].tolist(), ensure_ascii=False),
                    "citation_uris_json": json.dumps(evidence["citation_uri"].tolist()),
                    "reference_rows_json": best["reference_rows_json"],
                }
            )
        records.sort(key=lambda row: (-row["document_score"], row["document_id"]))
        for rank, record in enumerate(records[:top_k], start=1):
            record["rank"] = rank
        return pd.DataFrame(records[:top_k])


def _document_value_cache(documents: pd.DataFrame) -> dict[str, list[set[str]]]:
    fields = ("client", "offering", "sector", "service_nature", "country")
    return {
        field: [
            {normalize_search_text(value) for value in _json_values(raw)}
            for raw in documents[f"{field}_values_json"]
        ]
        for field in fields
    }


def build_bootstrap_probes(documents: pd.DataFrame, config: dict) -> pd.DataFrame:
    eligible = documents.loc[documents["retrieval_eligible"].eq(True)].copy().reset_index(drop=True)
    cache = _document_value_cache(eligible)
    templates = (
        ("client", "offering", "sector"),
        ("client", "service_nature", "country"),
        ("offering", "sector", "country"),
        ("client", "country"),
        ("service_nature", "country"),
    )
    candidates: dict[str, dict] = {}
    for row_index, row in eligible.iterrows():
        for fields in templates:
            selected: list[tuple[str, str]] = []
            for field in fields:
                values = _json_values(row[f"{field}_values_json"])
                if values:
                    value = values[0][:160].strip()
                    if normalize_search_text(value) not in {"oui", "non", "na", "n a"}:
                        selected.append((field, value))
            if len(selected) < 2:
                continue
            normalized_terms = [(field, normalize_search_text(value)) for field, value in selected]
            relevant: list[str] = []
            for candidate_index, candidate in eligible.iterrows():
                if all(term in cache[field][candidate_index] for field, term in normalized_terms):
                    relevant.append(str(candidate["document_id"]))
            if not relevant or len(relevant) > 12:
                continue
            query = "expérience et référence Devoteam " + " ".join(value for _, value in selected)
            key = normalize_search_text(query)
            filter_field, filter_value = selected[-1]
            candidates[key] = {
                "query_text": query,
                "query_language": "fr",
                "label_source": config["evaluation"]["bootstrap_label_source"],
                "relevant_document_ids_json": json.dumps(sorted(set(relevant))),
                "hard_filters_json": json.dumps({filter_field: filter_value}, ensure_ascii=False),
                "primary_document_id": str(row["document_id"]),
                "specificity_fields": len(selected),
                "relevant_document_count": len(set(relevant)),
            }
    rows = list(candidates.values())
    seed = str(config["evaluation"]["bootstrap_seed"])
    rows.sort(
        key=lambda row: (
            row["relevant_document_count"],
            -row["specificity_fields"],
            _sha256_text(seed + "|" + row["query_text"]),
        )
    )
    target = int(config["evaluation"]["bootstrap_probe_count"])
    selected: list[dict] = []
    used_primary: set[str] = set()
    for row in rows:
        if row["primary_document_id"] not in used_primary:
            selected.append(row)
            used_primary.add(row["primary_document_id"])
            if len(selected) == target:
                break
    if len(selected) < target:
        selected_keys = {normalize_search_text(row["query_text"]) for row in selected}
        for row in rows:
            if normalize_search_text(row["query_text"]) not in selected_keys:
                selected.append(row)
                selected_keys.add(normalize_search_text(row["query_text"]))
                if len(selected) == target:
                    break
    if len(selected) != target:
        raise AssertionError(f"Expected {target} bootstrap probes, produced {len(selected)}")
    for index, row in enumerate(selected, start=1):
        row["query_id"] = f"BOOT-{index:03d}-{_sha256_text(row['query_text'])[:8]}"
    return pd.DataFrame(selected)[
        [
            "query_id",
            "query_text",
            "query_language",
            "label_source",
            "relevant_document_ids_json",
            "hard_filters_json",
            "primary_document_id",
            "specificity_fields",
            "relevant_document_count",
        ]
    ]


def _query_metrics(relevant: set[str], ranked: list[str]) -> dict[str, float]:
    top10 = ranked[:10]
    top5 = ranked[:5]
    hits10 = len(relevant & set(top10))
    reciprocal_rank = next((1.0 / rank for rank, value in enumerate(ranked, start=1) if value in relevant), 0.0)
    dcg = sum((1.0 / math.log2(rank + 1)) for rank, value in enumerate(top10, start=1) if value in relevant)
    ideal = sum(1.0 / math.log2(rank + 1) for rank in range(1, min(len(relevant), 10) + 1))
    return {
        "recall_at_10": hits10 / max(len(relevant), 1),
        "precision_at_5": len(relevant & set(top5)) / 5.0,
        "mrr": reciprocal_rank,
        "ndcg_at_10": dcg / ideal if ideal else 0.0,
    }


def evaluate_bootstrap(
    *,
    probes: pd.DataFrame,
    query_vectors: np.ndarray,
    engine: HybridSearchEngine,
) -> tuple[pd.DataFrame, dict]:
    if query_vectors.shape[0] != len(probes):
        raise AssertionError("Bootstrap query-vector alignment failed")
    records: list[dict] = []
    metric_rows: dict[str, list[dict[str, float]]] = {mode: [] for mode in ("bm25", "dense", "hybrid")}
    latencies: list[float] = []
    filter_checks: list[bool] = []
    citation_checks: list[bool] = []
    for probe_index, probe in probes.iterrows():
        relevant = set(json.loads(probe["relevant_document_ids_json"]))
        for mode in ("bm25", "dense", "hybrid"):
            started = time.perf_counter()
            results = engine.search_documents(
                probe["query_text"],
                allowed_security_classifications=["INTERNAL", "PUBLIC"],
                mode=mode,
                top_k=10,
                query_vector=query_vectors[probe_index],
            )
            elapsed_ms = (time.perf_counter() - started) * 1000.0
            if mode == "hybrid":
                latencies.append(elapsed_ms)
            ranked = results["document_id"].astype(str).tolist() if not results.empty else []
            metric_rows[mode].append(_query_metrics(relevant, ranked))
            for result in results.to_dict(orient="records"):
                records.append(
                    {
                        "query_id": probe["query_id"],
                        "retrieval_mode": mode,
                        "document_id": result["document_id"],
                        "rank": int(result["rank"]),
                        "document_score": float(result["document_score"]),
                        "is_relevant_bootstrap": result["document_id"] in relevant,
                        "best_chunk_id": result["best_chunk_id"],
                        "citation_labels_json": result["citation_labels_json"],
                        "citation_uris_json": result["citation_uris_json"],
                    }
                )
                citation_checks.append(bool(json.loads(result["citation_uris_json"])))
        filters = json.loads(probe["hard_filters_json"])
        mask = engine.filter_engine.mask(
            allowed_security_classifications=["INTERNAL", "PUBLIC"], hard_filters=filters
        )
        filtered = engine.search_chunks(
            probe["query_text"],
            allowed_security_classifications=["INTERNAL", "PUBLIC"],
            hard_filters=filters,
            mode="hybrid",
            top_k=10,
            query_vector=query_vectors[probe_index],
        )
        filter_checks.append(
            (not filtered.empty)
            and all(mask[int(row)] for row in filtered["vector_row"].astype(int).tolist())
        )
    mode_metrics: dict[str, dict[str, float]] = {}
    for mode, values in metric_rows.items():
        mode_metrics[mode] = {
            metric: round(float(np.mean([row[metric] for row in values])), 6)
            for metric in ("recall_at_10", "precision_at_5", "mrr", "ndcg_at_10")
        }
    latency_array = np.asarray(latencies, dtype=float)
    metrics = {
        "schema_version": 1,
        "label_source": "BOOTSTRAP_METADATA_NOT_EXPERT",
        "expert_labeled": False,
        "production_quality_claim_allowed": False,
        "query_count": int(len(probes)),
        "retrieval_metrics": mode_metrics,
        "filter_correctness": float(np.mean(filter_checks)) if filter_checks else 0.0,
        "citation_coverage": float(np.mean(citation_checks)) if citation_checks else 0.0,
        "index_integrity": 1.0,
        "retrieval_core_latency_ms_p50": round(float(np.percentile(latency_array, 50)), 3),
        "retrieval_core_latency_ms_p95": round(float(np.percentile(latency_array, 95)), 3),
    }
    return pd.DataFrame(records), metrics


def create_expert_gold_set_workbook(path: Path, query_target: int = 50) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = openpyxl.Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instruction_rows = [
        ("Purpose", "Frozen expert relevance set for retrieval model promotion."),
        ("Queries", f"Provide {query_target} real or realistic Devoteam opportunities; do not copy confidential text unless approved."),
        ("Labelers", "Two domain experts label independently; disagreements are adjudicated."),
        ("Relevance 0", "Irrelevant: the reference does not address the opportunity."),
        ("Relevance 1", "Partially relevant: useful overlap, but not a strong primary reference."),
        ("Relevance 2", "Highly relevant: strong evidence for material opportunity requirements."),
        ("Promotion rule", "Only adjudicated labels may support Recall@10, Precision@5, MRR, and nDCG@10 claims."),
    ]
    for row in instruction_rows:
        instructions.append(row)
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 110

    queries = workbook.create_sheet("Queries")
    queries.append(
        [
            "query_id",
            "query_text",
            "language",
            "business_context",
            "mandatory_filters_json",
            "approved_for_evaluation",
            "notes",
        ]
    )
    for index in range(1, query_target + 1):
        queries.append([f"EXP-{index:03d}", "", "fr", "", "{}", "NO", ""])
    queries.freeze_panes = "A2"
    queries.auto_filter.ref = f"A1:G{query_target + 1}"
    widths = [15, 70, 14, 45, 45, 24, 45]
    for column, width in enumerate(widths, start=1):
        queries.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    language_validation = DataValidation(type="list", formula1='"fr,en,ar,mixed"')
    approval_validation = DataValidation(type="list", formula1='"NO,YES"')
    queries.add_data_validation(language_validation)
    queries.add_data_validation(approval_validation)
    language_validation.add(f"C2:C{query_target + 1}")
    approval_validation.add(f"F2:F{query_target + 1}")

    labels = workbook.create_sheet("Labels")
    labels.append(
        [
            "query_id",
            "document_id",
            "reference_id",
            "labeler_1_relevance",
            "labeler_2_relevance",
            "adjudicated_relevance",
            "failure_category",
            "evidence_notes",
        ]
    )
    for _ in range(query_target * 10):
        labels.append(["", "", "", "", "", "", "", ""])
    labels.freeze_panes = "A2"
    labels.auto_filter.ref = f"A1:H{query_target * 10 + 1}"
    relevance_validation = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    labels.add_data_validation(relevance_validation)
    for column in ("D", "E", "F"):
        relevance_validation.add(f"{column}2:{column}{query_target * 10 + 1}")
    for column, width in enumerate([15, 38, 38, 22, 22, 24, 28, 55], start=1):
        labels.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B2C83")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _package_versions() -> dict[str, str]:
    packages = ("numpy", "pandas", "pyarrow", "sentence-transformers", "transformers", "torch", "faiss-cpu")
    output: dict[str, str] = {}
    for package in packages:
        try:
            output[package] = importlib.metadata.version(package)
        except importlib.metadata.PackageNotFoundError:
            output[package] = "not-installed"
    return output


def _hash_outputs(run_root: Path, paths: Iterable[Path]) -> Path:
    lines = []
    for path in sorted(paths, key=lambda item: str(item.relative_to(run_root))):
        lines.append(f"{sha256_file(path)}  {path.relative_to(run_root)}\n")
    output = run_root / "SHA256SUMS.txt"
    output.write_text("".join(lines), encoding="utf-8")
    return output


def _verify_inputs(project_root: Path, config: dict) -> tuple[Path, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    run_root = (
        project_root
        / "data"
        / "canonical"
        / config["input"]["snapshot_id"]
        / config["input"]["phase4_run_name"]
    )
    phase4 = verify_phase4(run_root)
    paths = {
        "manifest": run_root / "PHASE_4_MANIFEST.json",
        "sums": run_root / "SHA256SUMS.txt",
        "chunks": run_root / "chunks.parquet",
        "documents": run_root / "documents_catalog.parquet",
        "references": run_root / "reference_catalog.parquet",
        "filters": run_root / "filter_values.json",
    }
    expected = {
        "manifest": "expected_phase4_manifest_sha256",
        "sums": "expected_phase4_sha256sums_sha256",
        "chunks": "expected_chunks_sha256",
        "documents": "expected_documents_sha256",
        "references": "expected_references_sha256",
        "filters": "expected_filter_values_sha256",
    }
    for name, key in expected.items():
        if sha256_file(paths[name]) != config["input"][key]:
            raise AssertionError(f"Pinned Phase 4 input changed: {name}")
    chunks = pd.read_parquet(paths["chunks"])
    documents = pd.read_parquet(paths["documents"])
    references = pd.read_parquet(paths["references"])
    if len(chunks) != int(config["input"]["expected_chunks"]):
        raise AssertionError("Phase 4 chunk count changed")
    if len(documents) != int(config["input"]["expected_documents"]):
        raise AssertionError("Phase 4 document count changed")
    if int(documents["retrieval_eligible"].sum()) != int(config["input"]["expected_retrieval_documents"]):
        raise AssertionError("Phase 4 retrieval-document count changed")
    if len(references) != int(config["input"]["expected_references"]):
        raise AssertionError("Phase 4 reference count changed")
    if phase4.get("status") != "PASS":
        raise AssertionError("Phase 4 is not approved")
    return run_root, chunks, documents, references


def _report(manifest: dict, metrics: dict) -> str:
    hybrid = metrics["retrieval_metrics"]["hybrid"]
    return f"""# Phase 5 — Multilingual hybrid retrieval

**Technical status:** {manifest['status']}
**Expert relevance status:** {manifest['expert_gold_set_status']}
**Production promotion:** {manifest['production_promotion_status']}

## Index

- Chunks indexed: **{manifest['chunks_indexed']}**
- Embedding model: **{manifest['embedding_model_id']}**
- Embedding dimensions: **{manifest['embedding_dimensions']}**
- BM25 terms: **{manifest['bm25_terms']}**
- FAISS type: **IndexFlatIP**

## Bootstrap diagnostics — not expert quality evidence

- Queries: **{metrics['query_count']}**
- Hybrid Recall@10: **{hybrid['recall_at_10']:.3f}**
- Hybrid Precision@5: **{hybrid['precision_at_5']:.3f}**
- Hybrid MRR: **{hybrid['mrr']:.3f}**
- Hybrid nDCG@10: **{hybrid['ndcg_at_10']:.3f}**
- Deterministic filter correctness: **{metrics['filter_correctness']:.3f}**
- Citation coverage: **{metrics['citation_coverage']:.3f}**

These probes validate plumbing and metadata-grounded lookup only. They are not
expert labels and cannot be used to claim the Phase 0 production targets.

## Promotion blocker

Two Devoteam experts must complete and adjudicate
`EXPERT_GOLD_SET_TEMPLATE.xlsx` using real or approved realistic opportunities.
Only then may BM25, dense, hybrid, and future reranker variants be compared for
production promotion.
"""


def run_phase5(project_root: Path, config_path: Path, progress=print, embedding_adapter=None) -> dict:
    config = load_phase5_config(config_path)
    phase4_root, chunks, documents, references = _verify_inputs(project_root, config)
    run_root = (
        project_root
        / config["output"]["root"]
        / config["input"]["snapshot_id"]
        / config["output"]["run_name"]
    )
    success_path = run_root / config["output"]["success_marker"]
    if success_path.exists():
        verified = verify_phase5(run_root)
        progress("Existing Phase 5 index verified; nothing was rebuilt.")
        return verified | {"run_root": str(run_root), "resumed": True}
    run_root.mkdir(parents=True, exist_ok=True)
    reports_root = run_root / "reports"
    evaluation_root = run_root / "evaluation"
    reports_root.mkdir(parents=True, exist_ok=True)
    evaluation_root.mkdir(parents=True, exist_ok=True)

    progress("Building aligned retrieval passages and BM25 index...")
    retrieval_texts = build_retrieval_texts(chunks)
    lookup = chunks.copy().reset_index(drop=True)
    lookup.insert(0, "vector_row", np.arange(len(lookup), dtype=np.int32))
    lookup["retrieval_text_sha256"] = [_sha256_text(text) for text in retrieval_texts]
    lookup_path = run_root / "chunk_lookup.parquet"
    lookup.to_parquet(lookup_path, index=False)
    bm25 = BM25Index.build(
        retrieval_texts,
        k1=float(config["bm25"]["k1"]),
        b=float(config["bm25"]["b"]),
    )
    bm25_path = run_root / "bm25_index.npz"
    vocabulary_path = run_root / "bm25_vocabulary.json"
    bm25.save(bm25_path, vocabulary_path)
    BM25Index.load(bm25_path, vocabulary_path)

    adapter = embedding_adapter or E5EmbeddingAdapter(config)
    progress(
        f"Encoding {len(retrieval_texts)} passages locally with {config['embedding']['model_id']} "
        f"on {getattr(adapter, 'device', 'configured device')}..."
    )
    embedding_path = run_root / "embeddings.npy"
    embedding_progress_path = run_root / "embedding_progress.json"
    embeddings = build_embeddings_resumable(
        texts=retrieval_texts,
        adapter=adapter,
        output_path=embedding_path,
        progress_path=embedding_progress_path,
        config=config,
        input_hash=config["input"]["expected_chunks_sha256"],
        progress=progress,
    )
    progress("Writing and verifying exact FAISS inner-product index...")
    faiss_path = run_root / "faiss.index"
    faiss_metadata = write_faiss_index(embeddings, faiss_path)

    progress("Running 50 metadata-grounded bootstrap probes and filter checks...")
    probes = build_bootstrap_probes(documents, config)
    query_vectors = np.asarray(adapter.encode_queries(probes["query_text"].tolist()), dtype=np.float32)
    if query_vectors.shape != (len(probes), int(config["embedding"]["dimensions"])):
        raise AssertionError("Query embedding matrix has the wrong shape")
    if not np.allclose(np.linalg.norm(query_vectors, axis=1), 1.0, atol=2e-3):
        raise AssertionError("Query embeddings are not normalized")
    engine = HybridSearchEngine(
        chunks=lookup,
        retrieval_texts=retrieval_texts,
        bm25=bm25,
        embeddings=embeddings,
        config=config,
        embedding_adapter=adapter,
    )
    bootstrap_results, metrics = evaluate_bootstrap(
        probes=probes, query_vectors=query_vectors, engine=engine
    )
    thresholds = config["evaluation"]["technical_thresholds"]
    if metrics["filter_correctness"] < float(thresholds["filter_correctness"]):
        raise AssertionError("Deterministic filter gate failed")
    if metrics["citation_coverage"] < float(thresholds["citation_coverage"]):
        raise AssertionError("Citation-coverage gate failed")
    if metrics["index_integrity"] < float(thresholds["index_integrity"]):
        raise AssertionError("Index-integrity gate failed")
    hybrid_recall = metrics["retrieval_metrics"]["hybrid"]["recall_at_10"]
    if hybrid_recall < float(thresholds["bootstrap_hybrid_recall_at_10"]):
        raise AssertionError(
            f"Bootstrap hybrid Recall@10 {hybrid_recall:.3f} is below the technical threshold"
        )
    probes_path = evaluation_root / "bootstrap_queries.parquet"
    results_path = evaluation_root / "bootstrap_results.parquet"
    metrics_path = evaluation_root / "bootstrap_metrics.json"
    probes.to_parquet(probes_path, index=False)
    bootstrap_results.to_parquet(results_path, index=False)
    _atomic_json(metrics_path, metrics)

    gold_path = evaluation_root / "EXPERT_GOLD_SET_TEMPLATE.xlsx"
    create_expert_gold_set_workbook(
        gold_path, query_target=int(config["evaluation"]["expert_query_target"])
    )
    runtime_path = run_root / "retrieval_runtime.json"
    runtime = {
        "schema_version": 1,
        "pipeline_version": config["pipeline_version"],
        "embedding": {
            "provider": config["embedding"]["provider"],
            "model_id": config["embedding"]["model_id"],
            "revision": config["embedding"]["revision"],
            "license": config["embedding"]["license"],
            "dimensions": int(config["embedding"]["dimensions"]),
            "device": str(getattr(adapter, "device", "unknown")),
            "local_execution_only": True,
        },
        "bm25": {
            "algorithm": config["bm25"]["algorithm"],
            "tokenizer_version": config["bm25"]["tokenizer_version"],
            "terms": len(bm25.vocabulary),
        },
        "faiss": faiss_metadata,
        "hybrid": config["hybrid"],
        "filters": config["filters"],
        "package_versions": _package_versions(),
    }
    _atomic_json(runtime_path, runtime)
    manifest = {
        "schema_version": 1,
        "phase": 5,
        "pipeline_version": config["pipeline_version"],
        "snapshot_id": config["input"]["snapshot_id"],
        "phase4_manifest_sha256": config["input"]["expected_phase4_manifest_sha256"],
        "chunks_sha256": config["input"]["expected_chunks_sha256"],
        "status": "TECHNICAL_PASS",
        "qa_gate": "PASS",
        "expert_gold_set_status": "PENDING_TWO_EXPERT_LABELERS",
        "production_promotion_status": "BLOCKED_PENDING_EXPERT_GOLD_SET",
        "bootstrap_label_source": config["evaluation"]["bootstrap_label_source"],
        "bootstrap_queries": int(len(probes)),
        "bootstrap_hybrid_recall_at_10": hybrid_recall,
        "filter_correctness": metrics["filter_correctness"],
        "citation_coverage": metrics["citation_coverage"],
        "chunks_indexed": int(len(chunks)),
        "documents_retrieval_eligible": int(documents["retrieval_eligible"].sum()),
        "references_catalogued": int(len(references)),
        "bm25_terms": int(len(bm25.vocabulary)),
        "embedding_model_id": config["embedding"]["model_id"],
        "embedding_model_revision": config["embedding"]["revision"],
        "embedding_dimensions": int(config["embedding"]["dimensions"]),
        "embedding_matrix_sha256": sha256_file(embedding_path),
        "faiss_index_sha256": sha256_file(faiss_path),
        "raw_user_query_log_rows": 0,
        "external_embedding_api_calls": 0,
        "external_llm_calls": 0,
        "source_snapshot_mutation_calls": 0,
        "phase4_output_mutation_calls": 0,
        "reranker_calls": 0,
        "completed_at_utc": _now(),
    }
    manifest_path = run_root / "PHASE_5_MANIFEST.json"
    _atomic_json(manifest_path, manifest)
    report_path = reports_root / "PHASE_5_RETRIEVAL_REPORT.md"
    report_path.write_text(_report(manifest, metrics), encoding="utf-8")
    sums_path = _hash_outputs(
        run_root,
        [
            lookup_path,
            bm25_path,
            vocabulary_path,
            embedding_path,
            embedding_progress_path,
            faiss_path,
            probes_path,
            results_path,
            metrics_path,
            gold_path,
            runtime_path,
            manifest_path,
            report_path,
        ],
    )
    _atomic_json(
        success_path,
        {
            "status": "COMPLETE_REPRODUCIBLE_HYBRID_INDEX",
            "snapshot_id": config["input"]["snapshot_id"],
            "pipeline_version": config["pipeline_version"],
            "manifest_sha256": sha256_file(manifest_path),
            "sha256sums_sha256": sha256_file(sums_path),
            "created_at_utc": _now(),
        },
    )
    return manifest | {"run_root": str(run_root), "resumed": False}


def verify_phase5(run_root: Path) -> dict:
    success_path = run_root / "_SUCCESS.json"
    manifest_path = run_root / "PHASE_5_MANIFEST.json"
    sums_path = run_root / "SHA256SUMS.txt"
    if not all(path.exists() for path in (success_path, manifest_path, sums_path)):
        raise AssertionError("Incomplete Phase 5 output")
    success = json.loads(success_path.read_text(encoding="utf-8"))
    if sha256_file(manifest_path) != success["manifest_sha256"]:
        raise AssertionError("Phase 5 manifest hash mismatch")
    if sha256_file(sums_path) != success["sha256sums_sha256"]:
        raise AssertionError("Phase 5 checksum-list hash mismatch")
    for line in sums_path.read_text(encoding="utf-8").splitlines():
        expected, relative = line.split("  ", 1)
        if sha256_file(run_root / relative) != expected:
            raise AssertionError(f"Phase 5 output hash mismatch: {relative}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("status") != "TECHNICAL_PASS" or manifest.get("qa_gate") != "PASS":
        raise AssertionError("Phase 5 technical gate is not PASS")
    if manifest.get("production_promotion_status") != "BLOCKED_PENDING_EXPERT_GOLD_SET":
        raise AssertionError("Phase 5 improperly bypassed expert evaluation")
    required_zero = (
        "raw_user_query_log_rows",
        "external_embedding_api_calls",
        "external_llm_calls",
        "source_snapshot_mutation_calls",
        "phase4_output_mutation_calls",
        "reranker_calls",
    )
    if any(manifest.get(field) != 0 for field in required_zero):
        raise AssertionError("Phase 5 security assertion failed")
    if manifest.get("chunks_indexed") != 1185 or manifest.get("embedding_dimensions") != 768:
        raise AssertionError("Phase 5 pinned index dimensions changed")
    lookup = pd.read_parquet(run_root / "chunk_lookup.parquet")
    if len(lookup) != 1185 or not lookup["chunk_id"].is_unique:
        raise AssertionError("Phase 5 chunk lookup is invalid")
    embeddings = np.load(run_root / "embeddings.npy", mmap_mode="r")
    if embeddings.shape != (1185, 768):
        raise AssertionError("Phase 5 embedding matrix shape changed")
    if not np.allclose(np.linalg.norm(embeddings, axis=1), 1.0, atol=2e-3):
        raise AssertionError("Phase 5 embedding matrix is not normalized")
    bm25 = BM25Index.load(run_root / "bm25_index.npz", run_root / "bm25_vocabulary.json")
    if bm25.document_count != 1185:
        raise AssertionError("Phase 5 BM25 row count changed")
    import faiss

    faiss_index = faiss.read_index(str(run_root / "faiss.index"))
    if int(faiss_index.ntotal) != 1185 or int(faiss_index.d) != 768:
        raise AssertionError("Phase 5 FAISS index is invalid")
    metrics = json.loads((run_root / "evaluation" / "bootstrap_metrics.json").read_text(encoding="utf-8"))
    if metrics.get("filter_correctness") != 1.0 or metrics.get("citation_coverage") != 1.0:
        raise AssertionError("Phase 5 deterministic retrieval checks failed")
    if metrics["retrieval_metrics"]["hybrid"]["recall_at_10"] < 0.85:
        raise AssertionError("Phase 5 bootstrap hybrid gate failed")
    return manifest
