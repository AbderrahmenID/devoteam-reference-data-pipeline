from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .phase2_policy import is_blocked_header, may_follow_links
from .phase2_utils import normalize_text, parse_drive_id


URL_RE = re.compile(r"https?://[^\s\"')]+", re.IGNORECASE)


def _candidate_strings(cell) -> list[tuple[str, str]]:
    candidates: list[tuple[str, str]] = []
    if cell.hyperlink and cell.hyperlink.target:
        candidates.append(("hyperlink", str(cell.hyperlink.target)))
    value = cell.value
    if isinstance(value, str):
        for url in URL_RE.findall(value):
            candidates.append(("cell_value", url))
    return candidates


def _find_sheet(workbook, candidates: list[str]):
    wanted = {normalize_text(value) for value in candidates}
    for sheet in workbook.worksheets:
        if normalize_text(sheet.title) in wanted:
            return sheet
    if len(workbook.worksheets) == 1:
        return workbook.worksheets[0]
    raise RuntimeError(
        f"No configured source sheet found. Available: {workbook.sheetnames}"
    )


def _find_header_row(sheet, scan_rows: int) -> tuple[int, dict[int, str]]:
    best: tuple[int, int, dict[int, str]] | None = None
    expected = {"client", "pays", "offre", "annee", "attestation", "ref"}
    for row_number in range(1, min(scan_rows, sheet.max_row) + 1):
        headers = {
            cell.column: str(cell.value).strip()
            for cell in sheet[row_number]
            if cell.value not in (None, "")
        }
        normalized = {normalize_text(value) for value in headers.values()}
        score = sum(any(token in header for header in normalized) for token in expected)
        if best is None or score > best[0]:
            best = (score, row_number, headers)
    if best is None or best[0] < 3:
        raise RuntimeError("Could not identify workbook header row safely")
    return best[1], best[2]


def discover_drive_links(workbook_path: Path, config: dict) -> dict:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    sheet = _find_sheet(
        workbook, config["source"]["master_workbook"]["sheet_candidates"]
    )
    header_row, headers = _find_header_row(
        sheet, int(config["source"]["master_workbook"]["header_scan_rows"])
    )
    blocked_headers = [
        header for header in headers.values() if is_blocked_header(header, config)
    ]
    permitted_columns = {
        column: header
        for column, header in headers.items()
        if may_follow_links(header, config)
    }
    # Intentionally do not access cell values from blocked columns.
    deduplicated: dict[tuple[int, int, str], dict] = {}
    for row_number in range(header_row + 1, sheet.max_row + 1):
        for column, header in permitted_columns.items():
            cell = sheet.cell(row=row_number, column=column)
            for channel, raw_url in _candidate_strings(cell):
                target_id = parse_drive_id(raw_url)
                if not target_id:
                    continue
                key = (row_number, column, target_id)
                record = deduplicated.setdefault(
                    key,
                    {
                        "sheet": sheet.title,
                        "row_number": row_number,
                        "column_number": column,
                        "cell_coordinate": cell.coordinate,
                        "source_header": header,
                        "target_file_id": target_id,
                        "channels": [],
                    },
                )
                if channel not in record["channels"]:
                    record["channels"].append(channel)
    records = sorted(
        deduplicated.values(),
        key=lambda row: (row["row_number"], row["column_number"], row["target_file_id"]),
    )
    for record in records:
        record["channels"] = ",".join(sorted(record["channels"]))
    workbook.close()
    return {
        "sheet": sheet.title,
        "header_row": header_row,
        "row_count": max(sheet.max_row - header_row, 0),
        "headers": list(headers.values()),
        "blocked_headers_present": blocked_headers,
        "records": records,
    }
