from __future__ import annotations

import hashlib
import json
import os
import re
import unicodedata
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import yaml

from .phase2_utils import sha256_file
from .phase3_1_repair import verify_phase3_1


WORD_RE = re.compile(r"\w+", re.UNICODE)
YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b")
DEHYPHENATE_RE = re.compile(r"(?<=[A-Za-zÀ-ÖØ-öø-ÿ])-\n(?=[a-zà-öø-ÿ])")
PUNCTUATION_ONLY_RE = re.compile(r"^[\W_]+$", re.UNICODE)
MULTISPACE_RE = re.compile(r"[ \t]+")
BOUNDARY_PATTERNS = ("\n\n", "\n", ". ", "! ", "? ", "؟ ", "; ", ": ", ", ")

FRENCH_STOPWORDS = {
    "de", "la", "le", "les", "des", "du", "un", "une", "et", "en", "pour",
    "dans", "sur", "par", "avec", "au", "aux", "est", "sont", "que", "qui",
    "ce", "cette", "ces", "nous", "notre", "leur", "leurs", "référence", "contrat",
}
ENGLISH_STOPWORDS = {
    "the", "of", "and", "to", "in", "for", "on", "with", "is", "are", "that",
    "this", "these", "by", "from", "our", "their", "project", "contract", "reference",
}

DOCUMENT_TYPE_RULES = [
    ("ATTESTATION", ("attestation", "certificate", "certificat", "lettre de référence", "شهادة", "إفادة")),
    ("MINUTES", ("procès-verbal", "proces-verbal", "p.v.", " pv ", "minutes of", "محضر")),
    ("SPECIFICATION", ("cahier des charges", "terms of reference", " cdc ", "spécification")),
    ("CONTRACT", ("contrat", "contract", "avenant", "convention", "marché", "marche", "accord", "عقد", "اتفاقية")),
]

OUTPUT_METADATA_FIELDS = [
    "reference_number",
    "country",
    "business_unit",
    "company_domain",
    "client",
    "sector",
    "service_nature",
    "funding",
    "offering",
    "project_year",
    "attestation_available",
]


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _atomic_json(path: Path, value: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    os.replace(temporary, path)


def load_phase4_config(path: Path) -> dict:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    if int(config.get("phase", 0)) != 4:
        raise ValueError("Expected Phase 4 configuration")
    if config.get("security", {}).get("external_llm_enabled"):
        raise ValueError("External LLM calls are forbidden during Phase 4")
    if config.get("security", {}).get("raw_text_allowed_in_outputs"):
        raise ValueError("Raw page text must remain forbidden during Phase 4")
    return config


def _normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9#]+", " ", text.casefold()).strip()


def _clean_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = unicodedata.normalize("NFKC", str(value))
    text = MULTISPACE_RE.sub(" ", text.replace("\r", " ").replace("\n", " ")).strip()
    return text


def _unique(values: Iterable[Any]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        cleaned = _clean_scalar(value)
        key = cleaned.casefold()
        if cleaned and key not in seen:
            output.append(cleaned)
            seen.add(key)
    return output


def _json_list(values: Iterable[Any]) -> str:
    return json.dumps(_unique(values), ensure_ascii=False)


def clean_redacted_text(value: Any, config: dict) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = "".join(
        ch for ch in text if ch in "\n\t" or unicodedata.category(ch)[0] != "C"
    )
    if config["cleaning"].get("dehyphenate_line_breaks", True):
        text = DEHYPHENATE_RE.sub("", text)
    lines: list[str] = []
    previous_blank = True
    for raw_line in text.splitlines():
        line = MULTISPACE_RE.sub(" ", raw_line).strip()
        if line and config["cleaning"].get("remove_punctuation_only_lines", True):
            if PUNCTUATION_ONLY_RE.fullmatch(line):
                continue
        if line:
            lines.append(line)
            previous_blank = False
        elif config["cleaning"].get("preserve_paragraph_breaks", True) and not previous_blank:
            lines.append("")
            previous_blank = True
    return "\n".join(lines).strip()


def detect_language(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", text)
    arabic = sum("ARABIC" in unicodedata.name(ch, "") for ch in normalized)
    latin = sum("LATIN" in unicodedata.name(ch, "") for ch in normalized)
    letters = max(arabic + latin, 1)
    words = [word.casefold() for word in WORD_RE.findall(normalized)]
    french = sum(word in FRENCH_STOPWORDS for word in words)
    english = sum(word in ENGLISH_STOPWORDS for word in words)
    latin_language = "fr" if french >= english else "en"
    arabic_ratio = arabic / letters
    latin_ratio = latin / letters
    if arabic_ratio >= 0.18 and latin_ratio >= 0.18:
        return f"{latin_language}-ar"
    if arabic_ratio >= 0.55:
        return "ar"
    if latin_ratio >= 0.40:
        return latin_language
    return "und"


def classify_document(filename: str, text: str) -> tuple[str, str]:
    haystack = f" {unicodedata.normalize('NFKC', filename).casefold()} {text[:6000].casefold()} "
    for document_type, keywords in DOCUMENT_TYPE_RULES:
        for keyword in keywords:
            if keyword in haystack:
                return document_type, f"keyword:{keyword.strip()}"
    return "OTHER_EVIDENCE", "fallback:no_keyword"


def chunk_page_text(text: str, config: dict) -> list[tuple[int, int, str]]:
    settings = config["chunking"]
    maximum = int(settings["max_characters"])
    overlap = int(settings["overlap_characters"])
    minimum = int(settings["minimum_characters"])
    window = int(settings["boundary_window_characters"])
    if maximum <= overlap or minimum <= 0:
        raise ValueError("Invalid Phase 4 chunking configuration")
    length = len(text)
    if not text.strip():
        return []
    chunks: list[tuple[int, int, str]] = []
    start = 0
    while start < length:
        target_end = min(start + maximum, length)
        end = target_end
        if target_end < length:
            search_start = max(start + minimum, target_end - window)
            best = -1
            best_width = 0
            for pattern in BOUNDARY_PATTERNS:
                position = text.rfind(pattern, search_start, target_end + 1)
                if position > best:
                    best = position
                    best_width = len(pattern)
            if best >= search_start:
                end = best + best_width
        raw = text[start:end]
        leading = len(raw) - len(raw.lstrip())
        trailing = len(raw.rstrip())
        actual_start = start + leading
        actual_end = start + trailing
        chunk = text[actual_start:actual_end]
        if chunk:
            chunks.append((actual_start, actual_end, chunk))
        if end >= length:
            break
        next_start = max(end - overlap, start + 1)
        for pattern in ("\n\n", "\n", " "):
            boundary = text.find(pattern, next_start, min(end + 1, length))
            if boundary >= 0:
                next_start = boundary + len(pattern)
                break
        start = next_start
    if len(chunks) >= 2 and len(chunks[-1][2]) < minimum:
        prior_start, _, _ = chunks[-2]
        tail_end = chunks[-1][1]
        merged = text[prior_start:tail_end].strip()
        merged_start = prior_start + (len(text[prior_start:tail_end]) - len(text[prior_start:tail_end].lstrip()))
        # A short tail may be merged only while respecting the configured
        # maximum.  Retrieval quality never justifies an oversized chunk.
        if len(merged) <= maximum:
            chunks[-2] = (merged_start, tail_end, merged)
            chunks.pop()
    if any(len(chunk) > maximum for _, _, chunk in chunks):
        raise AssertionError("Chunk exceeds configured maximum")
    if any(text[start:end] != chunk for start, end, chunk in chunks):
        raise AssertionError("Chunk provenance offsets do not reproduce chunk text")
    return chunks


def _workbook_headers(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        normalized = _normalize_header(sheet.cell(1, column).value)
        if normalized and normalized not in headers:
            headers[normalized] = column
    return headers


def build_reference_catalog(
    *,
    master_path: Path,
    evidence_links_path: Path,
    evidence_targets_path: Path,
    document_ids: set[str],
    config: dict,
) -> pd.DataFrame:
    import openpyxl

    links = pd.read_csv(evidence_links_path, dtype=str, keep_default_na=False)
    targets = pd.read_csv(evidence_targets_path, dtype=str, keep_default_na=False)
    target_map: dict[str, dict] = {}
    for row in targets.to_dict(orient="records"):
        canonical = row.get("resolved_file_id") or row["target_file_id"]
        target_map[row["target_file_id"]] = {
            "canonical_document_id": canonical,
            "snapshot_status": row.get("snapshot_status", ""),
            "target_name": row.get("name", ""),
        }

    workbook = openpyxl.load_workbook(master_path, read_only=False, data_only=True)
    sheet_name = config["master_metadata"]["sheet"]
    if sheet_name not in workbook.sheetnames:
        raise AssertionError(f"Master metadata sheet missing: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = _workbook_headers(sheet)
    field_columns: dict[str, int] = {}
    for output_name, source_header in config["master_metadata"]["fields"].items():
        normalized = _normalize_header(source_header)
        if normalized not in headers:
            raise AssertionError(f"Required master header missing: {source_header}")
        field_columns[output_name] = headers[normalized]

    forbidden = {
        _normalize_header(header) for header in config["master_metadata"]["forbidden_headers"]
    }
    if any(_normalize_header(config["master_metadata"]["fields"][field]) in forbidden for field in field_columns):
        raise AssertionError("A forbidden workbook header entered the allowlist")

    selected = links.loc[
        links["sheet"].eq(sheet_name)
        & links["source_header"].eq(config["master_metadata"]["link_source_header"])
    ].copy()
    records: list[dict] = []
    for link in selected.to_dict(orient="records"):
        row_number = int(link["row_number"])
        target = target_map.get(
            link["target_file_id"],
            {
                "canonical_document_id": link["target_file_id"],
                "snapshot_status": "UNKNOWN",
                "target_name": "",
            },
        )
        canonical_id = target["canonical_document_id"]
        available = canonical_id in document_ids and target["snapshot_status"] != "UNAVAILABLE"
        record = {
            "reference_id": _sha256_text(
                f"{sheet_name}|{row_number}|{link['target_file_id']}"
            ),
            "sheet": sheet_name,
            "row_number": row_number,
            "cell_coordinate": link["cell_coordinate"],
            "target_file_id": link["target_file_id"],
            "canonical_document_id": canonical_id,
            "snapshot_status": target["snapshot_status"],
            "evidence_available": bool(available),
            "target_name": target["target_name"],
        }
        for output_name, column in field_columns.items():
            value = _clean_scalar(sheet.cell(row_number, column).value)
            if output_name == "project_year":
                years = YEAR_RE.findall(value)
                value = ", ".join(_unique(years)) if years else value
            record[output_name] = value
        records.append(record)
    catalog = pd.DataFrame(records).sort_values(
        ["row_number", "target_file_id"], kind="stable"
    ).reset_index(drop=True)
    if catalog.empty:
        raise AssertionError("Reference catalogue is empty")
    return catalog


def _aggregate_reference_metadata(reference_catalog: pd.DataFrame) -> dict[str, dict]:
    aggregated: dict[str, dict] = {}
    available = reference_catalog.loc[reference_catalog["evidence_available"]].copy()
    for document_id, group in available.groupby("canonical_document_id", sort=True):
        values = {
            f"{field}_values": _unique(group[field].tolist())
            for field in OUTPUT_METADATA_FIELDS
        }
        values["reference_rows"] = sorted({int(value) for value in group["row_number"]})
        values["reference_ids"] = sorted(group["reference_id"].astype(str).unique().tolist())
        aggregated[str(document_id)] = values
    return aggregated


def build_corpus_frames(
    *,
    pages: pd.DataFrame,
    master_path: Path,
    evidence_links_path: Path,
    evidence_targets_path: Path,
    config: dict,
) -> dict[str, Any]:
    expected_columns = {
        "document_id", "page_number_1_based", "retrieval_eligible", "text_redacted",
        "source_file_name", "source_mime_type", "source_relative_path", "source_sha256",
        "qa_status", "qa_reasons", "extraction_method", "ocr_confidence",
        "security_classification", "snapshot_id",
    }
    missing_columns = expected_columns - set(pages.columns)
    if missing_columns:
        raise AssertionError(f"Curated pages missing columns: {sorted(missing_columns)}")
    if pages.duplicated(["document_id", "page_number_1_based"]).any():
        raise AssertionError("Duplicate page keys in curated input")

    document_ids = set(pages["document_id"].astype(str))
    reference_catalog = build_reference_catalog(
        master_path=master_path,
        evidence_links_path=evidence_links_path,
        evidence_targets_path=evidence_targets_path,
        document_ids=document_ids,
        config=config,
    )
    reference_metadata = _aggregate_reference_metadata(reference_catalog)

    eligible_input = pages.loc[pages["retrieval_eligible"].eq(True)].copy()
    canonical_records: list[dict] = []
    for row in json.loads(eligible_input.to_json(orient="records", force_ascii=False)):
        cleaned = clean_redacted_text(row[config["cleaning"]["source_text_column"]], config)
        if len(cleaned) < int(config["cleaning"]["minimum_page_characters"]):
            raise AssertionError(
                f"Eligible page became too short after cleaning: {row['document_id']} p.{row['page_number_1_based']}"
            )
        text_hash = _sha256_text(cleaned)
        page_id = _sha256_text(
            f"{row['source_sha256']}|{int(row['page_number_1_based'])}|{text_hash}|{config['pipeline_version']}"
        )
        canonical_records.append(
            {
                "page_id": page_id,
                "snapshot_id": row["snapshot_id"],
                "pipeline_version": config["pipeline_version"],
                "document_id": row["document_id"],
                "source_file_name": row["source_file_name"],
                "source_mime_type": row["source_mime_type"],
                "source_relative_path": row["source_relative_path"],
                "source_sha256": row["source_sha256"],
                "page_number_1_based": int(row["page_number_1_based"]),
                "extraction_method": row["extraction_method"],
                "ocr_confidence": row.get("ocr_confidence"),
                "phase3_qa_status": row["qa_status"],
                "security_classification": row["security_classification"],
                "page_language": detect_language(cleaned),
                "page_text_sha256": text_hash,
                "page_character_count": len(cleaned),
                "page_word_count": len(WORD_RE.findall(cleaned)),
                "page_text": cleaned,
                "citation_label": f"{row['source_file_name']} — page {int(row['page_number_1_based'])}",
                "citation_uri": f"https://drive.google.com/file/d/{row['document_id']}/view#page={int(row['page_number_1_based'])}",
            }
        )
    canonical_pages = pd.DataFrame(canonical_records).sort_values(
        ["document_id", "page_number_1_based"], kind="stable"
    ).reset_index(drop=True)

    document_records: list[dict] = []
    for (document_id, source_name), group in pages.groupby(
        ["document_id", "source_file_name"], sort=True
    ):
        eligible_group = canonical_pages.loc[canonical_pages["document_id"].eq(document_id)]
        combined_text = "\n\n".join(eligible_group["page_text"].tolist())
        document_type, document_type_rule = classify_document(source_name, combined_text)
        metadata = reference_metadata.get(str(document_id), {})
        page_count = int(len(group))
        eligible_count = int(group["retrieval_eligible"].sum())
        excluded_count = page_count - eligible_count
        if eligible_count == 0:
            quality = "NO_ELIGIBLE_PAGES"
        elif excluded_count:
            quality = "PARTIAL_PAGE_EXCLUSION"
        else:
            quality = "PASS"
        record = {
            "document_id": document_id,
            "source_file_name": source_name,
            "source_mime_type": group["source_mime_type"].iloc[0],
            "source_relative_path": group["source_relative_path"].iloc[0],
            "source_sha256": group["source_sha256"].iloc[0],
            "security_classification": group["security_classification"].iloc[0],
            "document_type": document_type,
            "document_type_rule": document_type_rule,
            "document_language": detect_language(combined_text),
            "page_count": page_count,
            "eligible_page_count": eligible_count,
            "excluded_page_count": excluded_count,
            "retrieval_eligible": eligible_count > 0,
            "data_quality_status": quality,
            "evidence_available": True,
            "reference_rows_json": json.dumps(metadata.get("reference_rows", [])),
            "reference_ids_json": json.dumps(metadata.get("reference_ids", [])),
        }
        for field in OUTPUT_METADATA_FIELDS:
            record[f"{field}_values_json"] = json.dumps(
                metadata.get(f"{field}_values", []), ensure_ascii=False
            )
        document_records.append(record)
    documents = pd.DataFrame(document_records).sort_values("document_id").reset_index(drop=True)

    document_lookup = documents.set_index("document_id").to_dict(orient="index")
    chunks: list[dict] = []
    for page in canonical_pages.to_dict(orient="records"):
        document = document_lookup[page["document_id"]]
        for chunk_index, (start, end, chunk_text) in enumerate(
            chunk_page_text(page["page_text"], config), start=1
        ):
            chunk_hash = _sha256_text(chunk_text)
            chunk_id = _sha256_text(
                f"{page['page_id']}|{start}|{end}|{chunk_hash}|{config['pipeline_version']}"
            )
            record = {
                "chunk_id": chunk_id,
                "page_id": page["page_id"],
                "snapshot_id": page["snapshot_id"],
                "pipeline_version": config["pipeline_version"],
                "document_id": page["document_id"],
                "source_file_name": page["source_file_name"],
                "source_mime_type": page["source_mime_type"],
                "source_relative_path": page["source_relative_path"],
                "source_sha256": page["source_sha256"],
                "page_number_1_based": page["page_number_1_based"],
                "chunk_index_in_page": chunk_index,
                "character_start": start,
                "character_end": end,
                "chunk_character_count": len(chunk_text),
                "chunk_word_count": len(WORD_RE.findall(chunk_text)),
                "chunk_text_sha256": chunk_hash,
                "chunk_text": chunk_text,
                "page_language": page["page_language"],
                "document_language": document["document_language"],
                "document_type": document["document_type"],
                "data_quality_status": document["data_quality_status"],
                "security_classification": page["security_classification"],
                "citation_label": page["citation_label"],
                "citation_uri": page["citation_uri"],
                "reference_rows_json": document["reference_rows_json"],
            }
            for field in OUTPUT_METADATA_FIELDS:
                record[f"{field}_values_json"] = document[f"{field}_values_json"]
            chunks.append(record)
    chunks_frame = pd.DataFrame(chunks).sort_values(
        ["document_id", "page_number_1_based", "chunk_index_in_page"], kind="stable"
    ).reset_index(drop=True)
    if chunks_frame.empty:
        raise AssertionError("Phase 4 produced no chunks")
    duplicate_counts = chunks_frame["chunk_text_sha256"].value_counts().to_dict()
    chunks_frame["exact_duplicate_count"] = chunks_frame["chunk_text_sha256"].map(duplicate_counts)
    chunks_frame["duplicate_group_id"] = chunks_frame.apply(
        lambda row: row["chunk_text_sha256"] if row["exact_duplicate_count"] > 1 else "",
        axis=1,
    )

    quality_by_document = documents.set_index("document_id")["data_quality_status"].to_dict()
    reference_catalog["data_quality_status"] = reference_catalog.apply(
        lambda row: (
            quality_by_document.get(row["canonical_document_id"], "MISSING_EVIDENCE")
            if row["evidence_available"]
            else "MISSING_EVIDENCE"
        ),
        axis=1,
    )
    eligibility = documents.set_index("document_id")["retrieval_eligible"].to_dict()
    reference_catalog["document_retrieval_eligible"] = reference_catalog[
        "canonical_document_id"
    ].map(lambda document_id: bool(eligibility.get(document_id, False)))

    excluded_columns = [
        "document_id", "source_file_name", "page_number_1_based", "qa_status", "qa_reasons",
        "phase3_1_repair_status", "source_sha256", "source_relative_path",
    ]
    excluded = pages.loc[~pages["retrieval_eligible"].eq(True), excluded_columns].copy()
    excluded = excluded.sort_values(["document_id", "page_number_1_based"]).reset_index(drop=True)

    forbidden_tokens = {
        _normalize_header(value)
        for value in config["master_metadata"]["forbidden_headers"]
    } | {"text raw"}
    for frame_name, frame in {
        "canonical_pages": canonical_pages,
        "chunks": chunks_frame,
        "documents": documents,
        "reference_catalog": reference_catalog,
        "excluded_pages": excluded,
    }.items():
        normalized_columns = {_normalize_header(column) for column in frame.columns}
        if normalized_columns & forbidden_tokens:
            raise AssertionError(f"Forbidden output field in {frame_name}")

    return {
        "canonical_pages": canonical_pages,
        "chunks": chunks_frame,
        "documents": documents,
        "reference_catalog": reference_catalog,
        "excluded_pages": excluded,
    }


def _filter_values(frames: dict[str, Any]) -> dict:
    documents = frames["documents"]
    reference_catalog = frames["reference_catalog"]
    values: dict[str, dict[str, int]] = {}
    for field in OUTPUT_METADATA_FIELDS:
        counter: Counter[str] = Counter()
        for raw in documents[f"{field}_values_json"]:
            counter.update(json.loads(raw))
        values[field] = dict(sorted(counter.items(), key=lambda item: (-item[1], item[0].casefold())))
    for field in ("document_type", "document_language", "data_quality_status"):
        counter = Counter(documents[field].astype(str).tolist())
        values[field] = dict(sorted(counter.items()))
    values["evidence_available"] = {
        str(key).lower(): int(value)
        for key, value in reference_catalog["evidence_available"].value_counts().to_dict().items()
    }
    return values


def _statistics(frames: dict[str, Any], pages: pd.DataFrame) -> dict:
    canonical = frames["canonical_pages"]
    chunks = frames["chunks"]
    documents = frames["documents"]
    references = frames["reference_catalog"]
    duplicate_chunks = int((chunks["exact_duplicate_count"] > 1).sum())
    return {
        "documents_total": int(len(documents)),
        "documents_retrieval_eligible": int(documents["retrieval_eligible"].sum()),
        "documents_without_eligible_pages": int((~documents["retrieval_eligible"]).sum()),
        "pages_input_total": int(len(pages)),
        "pages_canonical": int(len(canonical)),
        "pages_excluded": int(len(frames["excluded_pages"])),
        "chunks_total": int(len(chunks)),
        "chunk_characters_total": int(chunks["chunk_character_count"].sum()),
        "chunk_characters_mean": round(float(chunks["chunk_character_count"].mean()), 2),
        "chunk_characters_max": int(chunks["chunk_character_count"].max()),
        "duplicate_chunks": duplicate_chunks,
        "duplicate_groups": int(chunks.loc[chunks["exact_duplicate_count"] > 1, "duplicate_group_id"].nunique()),
        "references_total": int(len(references)),
        "missing_evidence_references": int((~references["evidence_available"]).sum()),
        "page_languages": dict(sorted(Counter(canonical["page_language"]).items())),
        "document_languages": dict(sorted(Counter(documents["document_language"]).items())),
        "document_types": dict(sorted(Counter(documents["document_type"]).items())),
        "document_quality": dict(sorted(Counter(documents["data_quality_status"]).items())),
    }


def _hash_outputs(run_root: Path, paths: Iterable[Path]) -> Path:
    lines = []
    for path in sorted(paths, key=lambda item: str(item.relative_to(run_root))):
        lines.append(f"{sha256_file(path)}  {path.relative_to(run_root)}\n")
    sums = run_root / "SHA256SUMS.txt"
    sums.write_text("".join(lines), encoding="utf-8")
    return sums


def _build_report(summary: dict, statistics: dict) -> str:
    return f"""# Phase 4 — Canonical retrieval corpus

**Status:** {summary['status']}
**QA gate:** {summary['qa_gate']}
**Completed (UTC):** {summary['completed_at_utc']}

## Corpus

- Downloaded evidence documents catalogued: **{statistics['documents_total']}**
- Documents with retrieval content: **{statistics['documents_retrieval_eligible']}**
- Canonical pages: **{statistics['pages_canonical']}**
- Excluded pages preserved: **{statistics['pages_excluded']}**
- Stable page-constrained chunks: **{statistics['chunks_total']}**
- Workbook reference rows catalogued: **{statistics['references_total']}**
- References with missing evidence: **{statistics['missing_evidence_references']}**

## Safety and provenance

- Corpus source: deterministically redacted Phase 3.1 text only
- Raw OCR text persisted by Phase 4: **no**
- Forbidden workbook fields ingested: **0**
- Source mutations: **0**
- Phase 3/3.1 mutations: **0**
- OCR, embeddings, or external LLM calls: **0**

## Downstream gate

Phase 5 must index `chunks.parquet` only and preserve every provenance field.
Excluded pages and missing evidence remain filterable catalogue facts but cannot
support generated claims.
"""


def _verify_input_paths(project_root: Path, config: dict) -> tuple[Path, Path, Path, Path, Path, pd.DataFrame]:
    snapshot_id = config["input"]["snapshot_id"]
    snapshot_root = project_root / "data" / "snapshots" / snapshot_id
    phase3_1_root = (
        project_root
        / "data"
        / "extracted"
        / snapshot_id
        / config["input"]["phase3_1_run_name"]
    )
    verified = verify_phase3_1(phase3_1_root)
    if sha256_file(phase3_1_root / "PHASE_3_1_MANIFEST.json") != config["input"]["expected_phase3_1_manifest_sha256"]:
        raise AssertionError("Pinned Phase 3.1 manifest hash mismatch")
    pages_path = phase3_1_root / "pages_curated.parquet"
    if sha256_file(pages_path) != config["input"]["expected_curated_pages_sha256"]:
        raise AssertionError("Pinned curated-pages hash mismatch")
    master_candidates = sorted((snapshot_root / "raw" / "master").glob("*.xlsx"))
    if len(master_candidates) != 1:
        raise AssertionError(f"Expected one snapshotted master workbook, found {len(master_candidates)}")
    master_path = master_candidates[0]
    links_path = snapshot_root / "manifests" / "evidence_links.csv"
    targets_path = snapshot_root / "manifests" / "evidence_targets.csv"
    if sha256_file(master_path) != config["input"]["expected_master_workbook_sha256"]:
        raise AssertionError("Pinned master-workbook hash mismatch")
    if sha256_file(links_path) != config["input"]["expected_evidence_links_sha256"]:
        raise AssertionError("Pinned evidence-links hash mismatch")
    if sha256_file(targets_path) != config["input"]["expected_evidence_targets_sha256"]:
        raise AssertionError("Pinned evidence-targets hash mismatch")
    pages = pd.read_parquet(pages_path)
    if len(pages) != int(config["input"]["expected_total_pages"]):
        raise AssertionError("Curated input page count changed")
    if int(pages["retrieval_eligible"].sum()) != int(config["input"]["expected_eligible_pages"]):
        raise AssertionError("Eligible page count changed")
    if verified["status"] != "PASS":
        raise AssertionError("Phase 3.1 technical gate is not PASS")
    return snapshot_root, phase3_1_root, master_path, links_path, targets_path, pages


def run_phase4(project_root: Path, config_path: Path, progress=print) -> dict:
    config = load_phase4_config(config_path)
    snapshot_root, phase3_1_root, master_path, links_path, targets_path, pages = _verify_input_paths(
        project_root, config
    )
    run_root = (
        project_root
        / config["output"]["root"]
        / config["input"]["snapshot_id"]
        / config["output"]["run_name"]
    )
    success_path = run_root / config["output"]["success_marker"]
    if success_path.exists():
        verified = verify_phase4(run_root)
        progress("Existing completed Phase 4 corpus verified; nothing was rebuilt.")
        return verified | {"run_root": str(run_root), "resumed": True}
    run_root.mkdir(parents=True, exist_ok=True)
    reports = run_root / "reports"
    reports.mkdir(parents=True, exist_ok=True)

    progress("Building allowlisted workbook reference catalogue...")
    frames = build_corpus_frames(
        pages=pages,
        master_path=master_path,
        evidence_links_path=links_path,
        evidence_targets_path=targets_path,
        config=config,
    )
    progress("Writing canonical pages, document catalogue, and stable chunks...")
    paths = {
        "canonical_pages": run_root / "canonical_pages.parquet",
        "chunks": run_root / "chunks.parquet",
        "documents": run_root / "documents_catalog.parquet",
        "references": run_root / "reference_catalog.parquet",
        "excluded": run_root / "excluded_pages.csv",
    }
    frames["canonical_pages"].to_parquet(paths["canonical_pages"], index=False)
    frames["chunks"].to_parquet(paths["chunks"], index=False)
    frames["documents"].to_parquet(paths["documents"], index=False)
    frames["reference_catalog"].to_parquet(paths["references"], index=False)
    frames["excluded_pages"].to_csv(paths["excluded"], index=False, encoding="utf-8-sig")
    jsonl_path = run_root / "chunks.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as handle:
        for row in frames["chunks"].to_dict(orient="records"):
            handle.write(json.dumps(row, ensure_ascii=False, default=str) + "\n")

    filters = _filter_values(frames)
    filter_path = run_root / "filter_values.json"
    _atomic_json(filter_path, filters)
    statistics = _statistics(frames, pages)
    if statistics["pages_canonical"] != int(config["input"]["expected_eligible_pages"]):
        raise AssertionError("Canonical page count does not match the hard gate")
    if statistics["pages_excluded"] != int(config["input"]["expected_excluded_pages"]):
        raise AssertionError("Excluded page count does not match the hard gate")
    if statistics["documents_total"] != int(config["input"]["expected_documents"]):
        raise AssertionError("Document count does not match the hard gate")
    if statistics["documents_retrieval_eligible"] != int(config["input"]["expected_retrieval_documents"]):
        raise AssertionError("Retrieval-document count does not match the hard gate")
    if statistics["references_total"] != int(config["input"]["expected_reference_rows"]):
        raise AssertionError("Reference-row count does not match the hard gate")
    if statistics["missing_evidence_references"] != int(config["input"]["expected_missing_evidence_targets"]):
        raise AssertionError("Missing-evidence count does not match the hard gate")
    if statistics["chunks_total"] != int(config["input"]["expected_chunks"]):
        raise AssertionError("Deterministic chunk count does not match the hard gate")
    if statistics["chunk_characters_max"] > int(config["chunking"]["max_characters"]):
        raise AssertionError("A chunk exceeds the configured maximum")
    statistics_path = run_root / "corpus_statistics.json"
    _atomic_json(statistics_path, statistics)
    summary = {
        "schema_version": 1,
        "phase": 4,
        "pipeline_version": config["pipeline_version"],
        "snapshot_id": config["input"]["snapshot_id"],
        "phase3_1_manifest_sha256": config["input"]["expected_phase3_1_manifest_sha256"],
        "curated_pages_sha256": config["input"]["expected_curated_pages_sha256"],
        "status": "PASS",
        "qa_gate": "PASS",
        "documents_total": statistics["documents_total"],
        "documents_retrieval_eligible": statistics["documents_retrieval_eligible"],
        "pages_canonical": statistics["pages_canonical"],
        "pages_excluded": statistics["pages_excluded"],
        "chunks_total": statistics["chunks_total"],
        "references_total": statistics["references_total"],
        "missing_evidence_references": statistics["missing_evidence_references"],
        "raw_text_output_columns": 0,
        "forbidden_workbook_fields_ingested": 0,
        "source_snapshot_mutation_calls": 0,
        "phase3_output_mutation_calls": 0,
        "ocr_calls": 0,
        "embedding_calls": 0,
        "external_llm_calls": 0,
        "completed_at_utc": _now(),
    }
    manifest_path = run_root / "PHASE_4_MANIFEST.json"
    _atomic_json(manifest_path, summary)
    report_path = reports / "PHASE_4_CORPUS_REPORT.md"
    report_path.write_text(_build_report(summary, statistics), encoding="utf-8")
    sums_path = _hash_outputs(
        run_root,
        [
            *paths.values(),
            jsonl_path,
            filter_path,
            statistics_path,
            manifest_path,
            report_path,
        ],
    )
    _atomic_json(
        success_path,
        {
            "status": "COMPLETE_REPRODUCIBLE_CANONICAL_CORPUS",
            "snapshot_id": config["input"]["snapshot_id"],
            "pipeline_version": config["pipeline_version"],
            "manifest_sha256": sha256_file(manifest_path),
            "sha256sums_sha256": sha256_file(sums_path),
            "created_at_utc": _now(),
        },
    )
    return summary | {"run_root": str(run_root), "resumed": False}


def verify_phase4(run_root: Path) -> dict:
    success_path = run_root / "_SUCCESS.json"
    manifest_path = run_root / "PHASE_4_MANIFEST.json"
    sums_path = run_root / "SHA256SUMS.txt"
    if not all(path.exists() for path in (success_path, manifest_path, sums_path)):
        raise AssertionError("Incomplete Phase 4 output")
    success = json.loads(success_path.read_text(encoding="utf-8"))
    if sha256_file(manifest_path) != success["manifest_sha256"]:
        raise AssertionError("Phase 4 manifest hash mismatch")
    if sha256_file(sums_path) != success["sha256sums_sha256"]:
        raise AssertionError("Phase 4 checksum-list hash mismatch")
    for line in sums_path.read_text(encoding="utf-8").splitlines():
        expected, relative = line.split("  ", 1)
        if sha256_file(run_root / relative) != expected:
            raise AssertionError(f"Phase 4 output hash mismatch: {relative}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    required_zero = (
        "raw_text_output_columns",
        "forbidden_workbook_fields_ingested",
        "source_snapshot_mutation_calls",
        "phase3_output_mutation_calls",
        "ocr_calls",
        "embedding_calls",
        "external_llm_calls",
    )
    if manifest.get("status") != "PASS" or manifest.get("qa_gate") != "PASS":
        raise AssertionError("Phase 4 gate is not PASS")
    if any(manifest.get(field) != 0 for field in required_zero):
        raise AssertionError("Phase 4 security assertion failed")
    if manifest.get("pages_canonical") != 389 or manifest.get("pages_excluded") != 19:
        raise AssertionError("Phase 4 pinned page counts changed")
    if manifest.get("documents_total") != 134 or manifest.get("documents_retrieval_eligible") != 132:
        raise AssertionError("Phase 4 pinned document counts changed")
    if manifest.get("chunks_total") != 1185 or manifest.get("references_total") != 161:
        raise AssertionError("Phase 4 pinned corpus counts changed")
    if manifest.get("missing_evidence_references") != 21:
        raise AssertionError("Phase 4 pinned missing-evidence count changed")
    return manifest
