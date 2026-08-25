from __future__ import annotations

import copy
import hashlib
import json
import math
import os
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

import numpy as np
import openpyxl
import pandas as pd
import yaml
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .phase2_utils import sha256_file
from .phase5_2_matching import (
    HardenedMatchingEngine,
    citation_metrics,
    filter_reference_contract,
    json_values,
    load_phase5_2_config,
    safe_spreadsheet_frame,
    safe_spreadsheet_value,
    verify_pinned_inputs,
)


class Phase7BridError(RuntimeError):
    """Raised when an approved BRID control or signed input is incomplete."""


class PinnedE5QueryAdapter:
    """Load the exact Phase 5 local E5 query encoder without pipeline imports."""

    def __init__(self, phase5_config_path: Path, device: str | None = None):
        from sentence_transformers import SentenceTransformer

        phase5_config = yaml.safe_load(
            phase5_config_path.read_text(encoding="utf-8")
        )
        if int(phase5_config.get("phase", 0)) != 5:
            raise ValueError("Expected Phase 5 configuration")
        settings = phase5_config.get("embedding", {})
        security = phase5_config.get("security", {})
        if security.get("external_llm_enabled") or security.get(
            "external_embedding_api_enabled"
        ):
            raise ValueError("The Phase 5 model must execute locally")
        if not settings.get("local_execution_only"):
            raise ValueError("The Phase 5 embedding model is not local-only")
        self.settings = settings
        self.model = SentenceTransformer(
            settings["model_id"],
            revision=settings["revision"],
            trust_remote_code=bool(settings["trust_remote_code"]),
            device=device,
        )
        dimension = int(self.model.get_sentence_embedding_dimension())
        if dimension != int(settings["dimensions"]):
            raise AssertionError(f"Embedding dimension changed: {dimension}")
        self.device = str(self.model.device)

    def encode_queries(self, texts: list[str]) -> np.ndarray:
        prefixed = [self.settings["query_prefix"] + text for text in texts]
        values = self.model.encode(
            prefixed,
            batch_size=int(
                self.settings["batch_size_gpu"]
                if self.device.startswith("cuda")
                else self.settings["batch_size_cpu"]
            ),
            normalize_embeddings=bool(
                self.settings["normalize_embeddings"]
            ),
            convert_to_numpy=True,
            show_progress_bar=False,
        )
        return np.asarray(values, dtype=np.float32)


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _atomic_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(value, encoding="utf-8")
    os.replace(temporary, path)


def _atomic_json(path: Path, value: Any) -> None:
    _atomic_text(
        path,
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
    )


def _clean(value: Any) -> str:
    return " ".join(str(value or "").split()).strip()


def _key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _clean(value).casefold())
    return "".join(char for char in text if not unicodedata.combining(char))


def _true(value: Any) -> bool:
    if value is True:
        return True
    if isinstance(value, (int, np.integer)) and not isinstance(value, bool):
        return int(value) == 1
    return _clean(value).upper() in {"TRUE", "YES", "Y", "1"}


def _sheet_rows(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [_clean(value) for value in values[0]]
    if not all(headers):
        raise Phase7BridError(f"Blank header detected in sheet {sheet.title}")
    return [
        dict(zip(headers, row))
        for row in values[1:]
        if any(value not in (None, "") for value in row)
    ]


def _parse_json_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [_clean(item) for item in value if _clean(item)]
    if value in (None, ""):
        return []
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError:
        return [_clean(value)]
    if isinstance(parsed, list):
        return [_clean(item) for item in parsed if _clean(item)]
    return [_clean(parsed)]


def load_phase7_brid_config(path: Path) -> dict[str, Any]:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    if int(config.get("phase", 0)) != 7:
        raise ValueError("Expected Phase 7 BRID configuration")
    if int(config.get("schema_version", 0)) != 2:
        raise ValueError("Expected the Phase 7 BRID v2 schema")
    security = config.get("security", {})
    blocked = (
        "external_llm_enabled",
        "external_embedding_api_enabled",
        "raw_offer_logging_allowed",
        "security_filters_disableable",
        "source_mutation_allowed",
    )
    if any(bool(security.get(field)) for field in blocked):
        raise ValueError("Phase 7 BRID security defaults were weakened")
    if not security.get("authorization_before_scoring_required"):
        raise ValueError("Authorization must precede scoring")
    retrieval = config.get("retrieval", {})
    if retrieval.get("mode") != "hybrid":
        raise ValueError("The controlled BRID run must use the hybrid engine")
    if retrieval.get("cross_encoder_enabled"):
        raise ValueError("Cross-encoder promotion remains blocked")
    if not math.isclose(
        float(retrieval["lexical_weight"]) + float(retrieval["dense_weight"]),
        1.0,
    ):
        raise ValueError("Phase 7 fusion weights must sum to one")
    scoring = config.get("scoring", {})
    if int(scoring["tender_points_total"]) != 100:
        raise ValueError("The approved BRID scoring model must total 100 points")
    if scoring.get("points_awarded_automatically"):
        raise ValueError("Tender points may not be awarded automatically")
    promotion = config.get("promotion", {})
    if promotion.get("business_shortlist_auto_approval"):
        raise ValueError("Business shortlist auto-approval is forbidden")
    if promotion.get("production_promotion_allowed"):
        raise ValueError("Production promotion remains blocked")
    return config


def _review_decision(
    row: dict[str, Any],
    *,
    identifier: str,
    config: dict[str, Any],
) -> str:
    decision = _clean(row.get("reviewer_decision") or config["review"]["pending_decision"]).upper()
    accepted = {str(value).upper() for value in config["review"]["accepted_decisions"]}
    rejected = str(config["review"]["rejected_decision"]).upper()
    pending = str(config["review"]["pending_decision"]).upper()
    if decision == pending:
        raise Phase7BridError(f"Pending review decision: {identifier}")
    if decision not in accepted | {rejected}:
        raise Phase7BridError(f"Unknown review decision for {identifier}: {decision}")
    if decision == "EDIT" and not _clean(row.get("reviewer_correction")):
        raise Phase7BridError(f"Edited row has no reviewer correction: {identifier}")
    return decision


def _approved_rows(
    rows: Sequence[dict[str, Any]],
    *,
    id_field: str,
    config: dict[str, Any],
    require_confirmed: bool = False,
) -> list[dict[str, Any]]:
    approved: list[dict[str, Any]] = []
    for row in rows:
        identifier = _clean(row.get(id_field))
        if not identifier:
            raise Phase7BridError(f"Missing {id_field}")
        decision = _review_decision(row, identifier=identifier, config=config)
        if require_confirmed and not _true(row.get("confirmed")):
            raise Phase7BridError(f"Filter is not confirmed: {identifier}")
        if decision == str(config["review"]["rejected_decision"]).upper():
            continue
        if "approved" in row and not _true(row.get("approved")):
            raise Phase7BridError(f"Accepted row is not marked approved: {identifier}")
        updated = dict(row)
        updated["reviewer_decision"] = decision
        updated["approval_source"] = "PINNED_HUMAN_REVIEW_WORKBOOK"
        approved.append(updated)
    return approved


def load_approved_brid_review(
    workbook_path: Path,
    config: dict[str, Any],
) -> dict[str, Any]:
    expected_hash = str(config["input"]["expected_review_workbook_sha256"])
    actual_hash = sha256_file(workbook_path)
    if actual_hash != expected_hash:
        raise AssertionError("Approved BRID workbook hash changed")
    workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    required_sheets = list(config["review"]["required_sheets"])
    if workbook.sheetnames != required_sheets:
        raise Phase7BridError(
            f"Unexpected BRID workbook sheets: {workbook.sheetnames}"
        )
    overview = {
        _clean(row.get("Control")): row.get("Value")
        for row in _sheet_rows(workbook["Overview"])
    }
    if _clean(overview.get("Label")) != config["input"]["required_label"]:
        raise Phase7BridError("The BRID workbook is not labeled SYNTHETIC_TEST_ONLY")
    if _clean(overview.get("Status")) != config["input"]["required_review_status"]:
        raise Phase7BridError("The BRID human review gate is not approved")
    if _clean(overview.get("Opportunity ID")) != config["input"]["opportunity_id"]:
        raise Phase7BridError("The workbook opportunity ID is inconsistent")
    if int(overview.get("Technical threshold")) != int(
        config["scoring"]["expected_technical_threshold"]
    ):
        raise Phase7BridError("The BRID technical threshold changed")

    requirements = _approved_rows(
        _sheet_rows(workbook["Requirements"]),
        id_field="requirement_id",
        config=config,
    )
    rules = _approved_rows(
        _sheet_rows(workbook["EligibilityRules"]),
        id_field="rule_id",
        config=config,
    )
    criteria = _approved_rows(
        _sheet_rows(workbook["Scoring"]),
        id_field="criterion_id",
        config=config,
    )
    filters = _approved_rows(
        _sheet_rows(workbook["Filters"]),
        id_field="filter_id",
        config=config,
        require_confirmed=bool(config["review"]["require_confirmed_filters"]),
    )
    expected_counts = {
        "requirements": int(config["input"]["expected_requirements"]),
        "rules": int(config["input"]["expected_eligibility_rules"]),
        "criteria": int(config["input"]["expected_scoring_criteria"]),
        "filters": int(config["input"]["expected_filters"]),
    }
    observed_counts = {
        "requirements": len(requirements),
        "rules": len(rules),
        "criteria": len(criteria),
        "filters": len(filters),
    }
    if observed_counts != expected_counts:
        raise Phase7BridError(
            f"Approved workbook row counts changed: {observed_counts}"
        )

    normalized_requirements: list[dict[str, Any]] = []
    for row in requirements:
        route = _clean(row.get("routing")).upper()
        priority = _clean(row.get("priority")).upper()
        if route not in {
            *config["review"]["matching_routes"],
            *config["review"]["policy_routes"],
            *config["review"]["non_retrieval_routes"],
        }:
            raise Phase7BridError(f"Unknown requirement route: {route}")
        source_text = _clean(row.get("requirement_text"))
        if row["reviewer_decision"] == "EDIT":
            source_text = _clean(row.get("reviewer_correction"))
        query_terms = _parse_json_list(row.get("query_terms"))
        match_query = source_text
        if config["retrieval"]["query_terms_appended"] and query_terms:
            match_query = f"{source_text} {' '.join(query_terms)}"
        normalized_requirements.append(
            {
                **row,
                "routing": route,
                "classification": priority,
                "source_requirement_text": source_text,
                "requirement_text": match_query,
                "query_terms": query_terms,
            }
        )

    normalized_rules: list[dict[str, Any]] = []
    for row in rules:
        normalized_rules.append(
            {
                **row,
                "level": _clean(row.get("level")).upper(),
                "minimum_count": int(row.get("minimum_count") or 0),
                "year_start_inclusive": (
                    int(row["year_start_inclusive"])
                    if row.get("year_start_inclusive") not in (None, "")
                    else None
                ),
                "year_end_inclusive": (
                    int(row["year_end_inclusive"])
                    if row.get("year_end_inclusive") not in (None, "")
                    else None
                ),
                "required_capability": _clean(row.get("required_capability")).upper(),
                "required_sector_or_region": _clean(
                    row.get("required_sector_or_region")
                ).upper(),
                "signed_attestation_required": _true(
                    row.get("signed_attestation_required")
                ),
            }
        )

    total_points = sum(int(row.get("points") or 0) for row in criteria)
    if total_points != int(config["scoring"]["tender_points_total"]):
        raise Phase7BridError(f"Approved scoring model totals {total_points}, not 100")
    normalized_criteria = [
        {
            **row,
            "points": int(row.get("points") or 0),
            "evaluation_route": _clean(row.get("evaluation_route")).upper(),
        }
        for row in criteria
    ]

    forbidden_global = {
        str(value) for value in config["eligibility"]["global_portfolio_filters_forbidden"]
    }
    normalized_filters: list[dict[str, Any]] = []
    for row in filters:
        field = _clean(row.get("field"))
        behavior = _clean(row.get("proposed_behavior")).upper()
        layer = _clean(row.get("application_layer")).upper()
        if field in forbidden_global and behavior == "HARD_ELIGIBILITY":
            raise Phase7BridError(
                f"Portfolio constraint was incorrectly promoted to a global hard filter: {field}"
            )
        normalized_filters.append(
            {
                **row,
                "field": field,
                "value": (
                    _parse_json_list(row.get("value"))
                    if _clean(row.get("value")).startswith("[")
                    else row.get("value")
                ),
                "proposed_behavior": behavior,
                "application_layer": layer,
            }
        )

    signed_types = list(config["eligibility"]["signed_attestation_evidence_types"])
    signed_filter = [
        row
        for row in normalized_filters
        if row["field"] == "evidence_type"
        and row["proposed_behavior"] == "HARD_ELIGIBILITY"
    ]
    if len(signed_filter) != 1:
        raise Phase7BridError("Exactly one signed-attestation hard gate is required")
    user_facets = [
        row
        for row in normalized_filters
        if row["proposed_behavior"] == config["eligibility"]["user_filter_behavior"]
        and row["application_layer"]
        == config["eligibility"]["user_filter_application_layer"]
    ]
    portfolio_filters = [
        row
        for row in normalized_filters
        if row["proposed_behavior"] == "PORTFOLIO_CONSTRAINT"
    ]
    if {row["field"] for row in portfolio_filters} != {
        "year_start",
        "sector_code",
        "region",
    }:
        raise Phase7BridError("The corrected BRID portfolio constraints changed")

    matching_routes = {str(value) for value in config["review"]["matching_routes"]}
    policy_routes = {str(value) for value in config["review"]["policy_routes"]}
    match_requirements = [
        row for row in normalized_requirements if row["routing"] in matching_routes
    ]
    policy_requirements = [
        row for row in normalized_requirements if row["routing"] in policy_routes
    ]
    non_retrieval_requirements = [
        row
        for row in normalized_requirements
        if row not in match_requirements and row not in policy_requirements
    ]
    if not match_requirements:
        raise Phase7BridError("No approved matching requirements remain")
    return {
        "overview": overview,
        "workbook_sha256": actual_hash,
        "requirements": normalized_requirements,
        "match_requirements": match_requirements,
        "policy_requirements": policy_requirements,
        "non_retrieval_requirements": non_retrieval_requirements,
        "eligibility_rules": normalized_rules,
        "scoring_criteria": normalized_criteria,
        "filters": normalized_filters,
        "hard_filters": {
            "evidence_available": True,
            "evidence_type": signed_types,
        },
        "portfolio_filters": portfolio_filters,
        "user_facets": user_facets,
        "approval_mode": "PINNED_HUMAN_REVIEW_WORKBOOK",
    }


def verify_phase5_2_inputs(
    project_root: Path,
    config: dict[str, Any],
) -> dict[str, Any]:
    phase5_2_config = load_phase5_2_config(
        project_root / "config" / "phase5_2_matching_hardening.yaml"
    )
    pinned = verify_pinned_inputs(project_root, phase5_2_config)
    settings = config["input"]
    phase5_2_root = (
        project_root
        / "data"
        / "indexes"
        / settings["snapshot_id"]
        / settings["phase5_2_run_name"]
    )
    manifest_path = phase5_2_root / "PHASE_5_2_MANIFEST.json"
    sums_path = phase5_2_root / "SHA256SUMS.txt"
    success_path = phase5_2_root / "_SUCCESS.json"
    quality_path = phase5_2_root / "PHASE_5_2_QUALITY_GATE.json"
    contract_path = phase5_2_root / "reference_contract.parquet"
    for path in (
        manifest_path,
        sums_path,
        success_path,
        quality_path,
        contract_path,
    ):
        if not path.is_file():
            raise FileNotFoundError(path)
    expected_hashes = {
        manifest_path: settings["expected_phase5_2_manifest_sha256"],
        sums_path: settings["expected_phase5_2_sha256sums_sha256"],
        contract_path: settings["expected_reference_contract_sha256"],
    }
    for path, expected in expected_hashes.items():
        if sha256_file(path) != expected:
            raise AssertionError(f"Pinned Phase 5.2 input changed: {path.name}")
    sums = {
        name: digest
        for digest, name in (
            line.split("  ", 1)
            for line in sums_path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        )
    }
    if sums.get(contract_path.name) != settings["expected_reference_contract_sha256"]:
        raise AssertionError("Reference contract is not signed by Phase 5.2")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    quality = json.loads(quality_path.read_text(encoding="utf-8"))
    success = json.loads(success_path.read_text(encoding="utf-8"))
    if manifest.get("qa_gate") != "PASS" or quality.get("technical_gate") != "PASS":
        raise AssertionError("Phase 5.2 technical gate is incomplete")
    if manifest.get("production_promotion_status") != (
        "BLOCKED_PENDING_PHASE_5_1_EXPERT_EVALUATION"
    ):
        raise AssertionError("Phase 5.1 promotion gate was weakened")
    if manifest.get("citation_correctness_status") != "PENDING_HUMAN_AUDIT":
        raise AssertionError("Citation correctness must remain a human gate")
    if success.get("manifest_sha256") != settings["expected_phase5_2_manifest_sha256"]:
        raise AssertionError("Phase 5.2 success marker is inconsistent")
    return {
        "phase5_2_config": phase5_2_config,
        "pinned": pinned,
        "phase5_2_root": phase5_2_root,
        "manifest": manifest,
        "quality_gate": quality,
        "success": success,
        "reference_contract_path": contract_path,
    }


def _engine_config(
    phase5_2_config: dict[str, Any],
    phase7_config: dict[str, Any],
) -> dict[str, Any]:
    engine_config = copy.deepcopy(phase5_2_config)
    retrieval = phase7_config["retrieval"]
    engine_config["retrieval"]["default_mode"] = retrieval["mode"]
    engine_config["retrieval"]["fusion"]["lexical_weight"] = float(
        retrieval["lexical_weight"]
    )
    engine_config["retrieval"]["fusion"]["dense_weight"] = float(
        retrieval["dense_weight"]
    )
    engine_config["retrieval"]["candidate_chunks_per_requirement"] = int(
        retrieval["candidate_references"]
    )
    engine_config["scoring"]["maximum_recommendations"] = int(
        retrieval["maximum_recommendations"]
    )
    # A tender portfolio may cover MUST requirements across several references.
    # The individual-reference MUST gate is replaced by the explicit portfolio
    # count and union-coverage gates below.
    engine_config["eligibility"]["must_requirement_gate"] = False
    return engine_config


def _tag_set(value: Any) -> set[str]:
    return {str(item).upper() for item in json_values(value)}


def annotate_portfolio_capabilities(frame: pd.DataFrame) -> pd.DataFrame:
    annotated = frame.copy()
    capabilities = annotated["capability_tags_json"].map(_tag_set)
    engagements = annotated["engagement_tags_json"].map(_tag_set)
    offering = annotated["offering_code"].astype(str).str.upper()
    annotated["is_signed_attestation"] = (
        annotated["evidence_available"].eq(True)
        & annotated["evidence_type"].astype(str).isin(
            {"ATTESTATION_ORIGINAL", "ATTESTATION_SCAN"}
        )
    )
    annotated["is_sdsi"] = offering.isin({"IT_STRATEGY", "IT_STRATEGY_AMOA"}) | (
        capabilities.map(lambda values: "IT_GOVERNANCE" in values)
        & capabilities.map(lambda values: "ROADMAP" in values)
    )
    annotated["is_bfsi"] = annotated["sector_code"].astype(str).isin(
        {"BANKING", "INSURANCE"}
    )
    annotated["is_pca"] = offering.eq("BUSINESS_CONTINUITY") | capabilities.map(
        lambda values: "BUSINESS_CONTINUITY" in values
    )
    annotated["is_security"] = offering.eq("CYBERSECURITY") | capabilities.map(
        lambda values: "CYBERSECURITY" in values
    )
    annotated["is_sdsi_implementation_amoa"] = offering.eq(
        "IT_STRATEGY_AMOA"
    ) | (
        annotated["is_sdsi"]
        & engagements.map(lambda values: {"AMOA", "IMPLEMENTATION"} <= values)
    )
    annotated["is_west_or_north_africa"] = annotated["subregion"].map(_key).isin(
        {_key("Afrique de l'Ouest"), _key("Afrique du Nord")}
    )
    return annotated


def _rule_mask(frame: pd.DataFrame, rule: dict[str, Any]) -> pd.Series:
    capability = rule["required_capability"]
    if capability == "SIGNED_CLIENT_ATTESTATION":
        mask = frame["is_signed_attestation"].eq(True)
    elif capability == "SDSI":
        mask = frame["is_sdsi"].eq(True)
    elif capability == "PCA_PCI_PRA":
        mask = frame["is_pca"].eq(True)
    elif capability == "SECURITY_PSSI_ISO27001":
        mask = frame["is_security"].eq(True)
    elif capability == "SDSI_IMPLEMENTATION_AMOA":
        mask = frame["is_sdsi_implementation_amoa"].eq(True)
    elif capability == "ANY":
        mask = pd.Series(True, index=frame.index, dtype=bool)
    else:
        raise Phase7BridError(f"Unsupported portfolio capability: {capability}")
    scope = rule.get("required_sector_or_region")
    if scope == "BANKING_FINANCE_INSURANCE":
        mask &= frame["is_bfsi"].eq(True)
    elif scope == "WEST_OR_NORTH_AFRICA":
        mask &= frame["is_west_or_north_africa"].eq(True)
    elif scope:
        raise Phase7BridError(f"Unsupported sector/region rule: {scope}")
    if rule.get("signed_attestation_required"):
        mask &= frame["is_signed_attestation"].eq(True)
    if rule.get("year_start_inclusive") is not None:
        mask &= frame["year_end"].notna() & frame["year_end"].ge(
            int(rule["year_start_inclusive"])
        )
    if rule.get("year_end_inclusive") is not None:
        mask &= frame["year_start"].notna() & frame["year_start"].le(
            int(rule["year_end_inclusive"])
        )
    return mask


def evaluate_portfolio_rules(
    selected: pd.DataFrame,
    rules: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for rule in rules:
        mask = _rule_mask(selected, rule) if not selected.empty else pd.Series(dtype=bool)
        qualifying = (
            selected.loc[mask, "reference_id"].astype(str).tolist()
            if not selected.empty
            else []
        )
        if rule["level"] == "PER_REFERENCE":
            passed = bool(len(selected)) and bool(selected["is_signed_attestation"].all())
            observed = int(selected["is_signed_attestation"].sum())
            required = int(len(selected))
        elif rule["level"] == "PORTFOLIO_COUNT":
            observed = len(qualifying)
            required = int(rule["minimum_count"])
            passed = observed >= required
        else:
            raise Phase7BridError(f"Unsupported rule level: {rule['level']}")
        rows.append(
            {
                "rule_id": rule["rule_id"],
                "code": rule["code"],
                "level": rule["level"],
                "required_count": required,
                "observed_count": observed,
                "passed": passed,
                "qualifying_reference_ids": qualifying,
                "description": _clean(rule.get("description")),
                "source_page": rule.get("source_page"),
            }
        )
    return rows


def _coverage_lookup(coverage: pd.DataFrame) -> dict[str, set[str]]:
    output: dict[str, set[str]] = defaultdict(set)
    if coverage.empty:
        return output
    for row in coverage.loc[coverage["covered"].eq(True)].to_dict(orient="records"):
        output[str(row["reference_id"])].add(str(row["requirement_id"]))
    return output


def select_portfolio(
    candidates: pd.DataFrame,
    coverage: pd.DataFrame,
    rules: Sequence[dict[str, Any]],
    config: dict[str, Any],
) -> pd.DataFrame:
    if candidates.empty:
        return candidates.copy()
    maximum = int(config["eligibility"]["maximum_portfolio_references"])
    minimum = int(config["eligibility"]["minimum_portfolio_references"])
    annotated = annotate_portfolio_capabilities(candidates)
    if "final_rank" not in annotated:
        annotated = annotated.sort_values(
            ["final_score", "reference_id"], ascending=[False, True]
        ).reset_index(drop=True)
        annotated["final_rank"] = np.arange(1, len(annotated) + 1)
    annotated = annotated.sort_values(
        ["final_score", "reference_id"], ascending=[False, True]
    ).reset_index(drop=True)
    selected_ids: list[str] = []
    reasons: dict[str, list[str]] = defaultdict(list)

    portfolio_rules = [row for row in rules if row["level"] == "PORTFOLIO_COUNT"]
    while len(selected_ids) < maximum:
        current = annotated.loc[
            annotated["reference_id"].astype(str).isin(selected_ids)
        ]
        deficits: dict[str, int] = {}
        for rule in portfolio_rules:
            observed = int(_rule_mask(current, rule).sum()) if not current.empty else 0
            deficits[rule["code"]] = max(0, int(rule["minimum_count"]) - observed)
        if not any(deficits.values()):
            break
        choices: list[tuple[float, float, str, list[str]]] = []
        for _, candidate in annotated.iterrows():
            reference_id = str(candidate["reference_id"])
            if reference_id in selected_ids:
                continue
            single = candidate.to_frame().T
            contributions = [
                rule["code"]
                for rule in portfolio_rules
                if deficits[rule["code"]] > 0 and bool(_rule_mask(single, rule).iloc[0])
            ]
            if contributions:
                weighted = sum(
                    1.0 / max(1, int(rule["minimum_count"]))
                    for rule in portfolio_rules
                    if rule["code"] in contributions
                )
                choices.append(
                    (
                        weighted,
                        float(candidate["final_score"]),
                        reference_id,
                        contributions,
                    )
                )
        if not choices:
            break
        choices.sort(key=lambda item: (-item[0], -item[1], item[2]))
        _, _, reference_id, contributions = choices[0]
        selected_ids.append(reference_id)
        reasons[reference_id].append(
            "PORTFOLIO_RULE:" + ",".join(sorted(contributions))
        )

    covered_by_reference = _coverage_lookup(coverage)
    requirement_priority = {
        str(row["requirement_id"]): (
            2.0 if str(row["classification"]).upper() == "MUST" else 1.0
        )
        for row in coverage.drop_duplicates("requirement_id").to_dict(orient="records")
    }
    covered = set().union(*(covered_by_reference.get(item, set()) for item in selected_ids))
    all_requirements = set(requirement_priority)
    while len(selected_ids) < maximum and covered != all_requirements:
        choices = []
        for _, candidate in annotated.iterrows():
            reference_id = str(candidate["reference_id"])
            if reference_id in selected_ids:
                continue
            new_requirements = covered_by_reference.get(reference_id, set()) - covered
            gain = sum(requirement_priority[item] for item in new_requirements)
            if gain > 0:
                choices.append(
                    (
                        gain,
                        float(candidate["final_score"]),
                        reference_id,
                        sorted(new_requirements),
                    )
                )
        if not choices:
            break
        choices.sort(key=lambda item: (-item[0], -item[1], item[2]))
        _, _, reference_id, new_requirements = choices[0]
        selected_ids.append(reference_id)
        covered.update(new_requirements)
        reasons[reference_id].append(
            "REQUIREMENT_COVERAGE:" + ",".join(new_requirements)
        )

    for _, candidate in annotated.iterrows():
        if len(selected_ids) >= max(minimum, maximum):
            break
        reference_id = str(candidate["reference_id"])
        if reference_id not in selected_ids:
            selected_ids.append(reference_id)
            reasons[reference_id].append("RANKED_PORTFOLIO_COMPLETION")

    selected = annotated.loc[
        annotated["reference_id"].astype(str).isin(selected_ids)
    ].copy()
    order = {reference_id: index for index, reference_id in enumerate(selected_ids, 1)}
    selected["portfolio_rank"] = selected["reference_id"].astype(str).map(order)
    selected["selection_reason"] = selected["reference_id"].astype(str).map(
        lambda value: ";".join(reasons[value])
    )
    selected = selected.sort_values("portfolio_rank").reset_index(drop=True)
    return selected


def requirement_portfolio_coverage(
    selected: pd.DataFrame,
    coverage: pd.DataFrame,
    requirements: Sequence[dict[str, Any]],
) -> list[dict[str, Any]]:
    selected_ids = set(selected["reference_id"].astype(str)) if not selected.empty else set()
    rows: list[dict[str, Any]] = []
    for requirement in requirements:
        relevant = coverage.loc[
            coverage["requirement_id"].astype(str).eq(str(requirement["requirement_id"]))
            & coverage["reference_id"].astype(str).isin(selected_ids)
            & coverage["covered"].eq(True)
        ]
        identifiers = sorted(set(relevant["reference_id"].astype(str)))
        rows.append(
            {
                "requirement_id": requirement["requirement_id"],
                "code": requirement["code"],
                "classification": requirement["classification"],
                "requirement_text": requirement["source_requirement_text"],
                "supporting_reference_count": len(identifiers),
                "supporting_reference_ids": identifiers,
                "portfolio_covered": bool(identifiers),
                "source_page": requirement.get("source_page"),
            }
        )
    return rows


def evaluate_scoring_criteria(
    selected: pd.DataFrame,
    criteria: Sequence[dict[str, Any]],
    rule_results: Sequence[dict[str, Any]],
    requirement_results: Sequence[dict[str, Any]],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    rules = {row["code"]: row for row in rule_results}
    requirements = {row["code"]: row for row in requirement_results}
    human_routes = set(config["scoring"]["human_only_routes"])
    output: list[dict[str, Any]] = []
    for criterion in criteria:
        code = criterion["code"]
        route = criterion["evaluation_route"]
        support_ids: set[str] = set()
        status = "READY_FOR_HUMAN_SCORING"
        if route in human_routes:
            status = "OUTSIDE_REFERENCE_MATCHING_SCOPE"
        elif code == "SCORE-GENERAL":
            for row in requirement_results:
                support_ids.update(row["supporting_reference_ids"])
            if not all(row["portfolio_covered"] for row in requirement_results):
                status = "EVIDENCE_GAP"
        elif code == "SCORE-SDSI-BFSI":
            support_ids.update(
                rules.get("ELIG-SDSI-BFSI-2", {}).get(
                    "qualifying_reference_ids", []
                )
            )
            if not rules.get("ELIG-SDSI-BFSI-2", {}).get("passed", False):
                status = "EVIDENCE_GAP"
        elif code == "SCORE-CONT-SEC":
            for rule_code in ("ELIG-PCA-2", "ELIG-SEC-2"):
                support_ids.update(
                    rules.get(rule_code, {}).get("qualifying_reference_ids", [])
                )
            if not all(
                rules.get(rule_code, {}).get("passed", False)
                for rule_code in ("ELIG-PCA-2", "ELIG-SEC-2")
            ):
                status = "EVIDENCE_GAP"
        elif code == "SCORE-REGION":
            support_ids.update(
                rules.get("ELIG-AFRICA-2", {}).get(
                    "qualifying_reference_ids", []
                )
            )
            if not rules.get("ELIG-AFRICA-2", {}).get("passed", False):
                status = "EVIDENCE_GAP"
        output.append(
            {
                "criterion_id": criterion["criterion_id"],
                "code": code,
                "criterion": criterion["criterion"],
                "points_available": int(criterion["points"]),
                "evaluation_route": route,
                "evidence_status": status,
                "supporting_reference_count": len(support_ids),
                "supporting_reference_ids": sorted(support_ids),
                "points_awarded": None,
                "human_scoring_required": True,
                "source_page": criterion.get("source_page"),
            }
        )
    return output


def _write_sheet(
    workbook: openpyxl.Workbook,
    name: str,
    rows: Sequence[dict[str, Any]] | pd.DataFrame,
) -> openpyxl.worksheet.worksheet.Worksheet:
    sheet = workbook.create_sheet(name)
    frame = rows.copy() if isinstance(rows, pd.DataFrame) else pd.DataFrame(list(rows))
    frame = safe_spreadsheet_frame(frame)
    headers = list(frame.columns)
    if not headers:
        headers = ["status"]
        frame = pd.DataFrame([{"status": "NO_ROWS"}])
    sheet.append(headers)
    for row in frame.itertuples(index=False, name=None):
        sheet.append([safe_spreadsheet_value(value) for value in row])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        observed = [
            len(str(sheet.cell(row=row, column=index).value or ""))
            for row in range(1, min(sheet.max_row, 200) + 1)
        ]
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(index)
        ].width = min(max(max(observed, default=len(header)) + 2, 12), 55)
    return sheet


def create_phase7_review_workbook(
    path: Path,
    *,
    manifest: dict[str, Any],
    selected: pd.DataFrame,
    candidates: pd.DataFrame,
    rule_results: Sequence[dict[str, Any]],
    requirement_results: Sequence[dict[str, Any]],
    evidence: pd.DataFrame,
    scoring_results: Sequence[dict[str, Any]],
    user_facets: Sequence[dict[str, Any]],
) -> None:
    workbook = openpyxl.Workbook()
    default = workbook.active
    workbook.remove(default)
    overview = [
        {"Control": "Label", "Value": manifest["label"]},
        {"Control": "Status", "Value": manifest["status"]},
        {"Control": "Opportunity ID", "Value": manifest["opportunity_id"]},
        {"Control": "Retrieval mode", "Value": manifest["retrieval_mode"]},
        {"Control": "Fusion", "Value": manifest["fusion_label"]},
        {"Control": "Candidate references", "Value": manifest["candidate_references"]},
        {"Control": "Portfolio references", "Value": manifest["portfolio_references"]},
        {"Control": "Portfolio eligibility", "Value": manifest["portfolio_eligibility"]},
        {"Control": "Tender points", "Value": manifest["tender_points_total"]},
        {
            "Control": "Tender score status",
            "Value": manifest["technical_threshold_status"],
        },
        {
            "Control": "Citation correctness",
            "Value": manifest["citation_correctness_status"],
        },
        {
            "Control": "Instruction",
            "Value": "Audit citations, then mark each portfolio row SHORTLIST or REJECT.",
        },
    ]
    _write_sheet(workbook, "Overview", overview)

    portfolio_columns = [
        "portfolio_rank",
        "reference_id",
        "client_raw",
        "country_code",
        "subregion",
        "sector_code",
        "offering_code",
        "project_year_raw",
        "evidence_type",
        "final_score",
        "must_covered",
        "must_total",
        "selection_reason",
    ]
    portfolio = selected[[column for column in portfolio_columns if column in selected]].copy()
    portfolio["reviewer_decision"] = "PENDING"
    portfolio["reviewer_notes"] = ""
    portfolio_sheet = _write_sheet(workbook, "Recommended Portfolio", portfolio)
    decision_column = portfolio.columns.get_loc("reviewer_decision") + 1
    validation = DataValidation(
        type="list",
        formula1='"PENDING,SHORTLIST,REJECT"',
        allow_blank=False,
    )
    portfolio_sheet.add_data_validation(validation)
    if portfolio_sheet.max_row >= 2:
        validation.add(
            f"{openpyxl.utils.get_column_letter(decision_column)}2:"
            f"{openpyxl.utils.get_column_letter(decision_column)}{portfolio_sheet.max_row}"
        )

    candidate_columns = [
        "final_rank",
        "reference_id",
        "client_raw",
        "country_code",
        "subregion",
        "sector_code",
        "offering_code",
        "project_year_raw",
        "evidence_type",
        "final_score",
        "weighted_relevance",
        "weighted_coverage",
        "must_covered",
        "must_total",
    ]
    _write_sheet(
        workbook,
        "Candidate Pool",
        candidates[[column for column in candidate_columns if column in candidates]],
    )
    _write_sheet(workbook, "Eligibility Rules", rule_results)
    _write_sheet(workbook, "Requirement Coverage", requirement_results)

    selected_ids = set(selected["reference_id"].astype(str)) if not selected.empty else set()
    selected_evidence = evidence.loc[
        evidence["reference_id"].astype(str).isin(selected_ids)
    ].copy()
    evidence_columns = [
        "reference_id",
        "requirement_id",
        "source_file_name",
        "page_number_1_based",
        "citation_label",
        "citation_uri",
        "chunk_id",
        "rerank_score",
        "term_coverage",
        "chunk_text",
    ]
    _write_sheet(
        workbook,
        "Evidence",
        selected_evidence[
            [column for column in evidence_columns if column in selected_evidence]
        ],
    )
    _write_sheet(workbook, "Scoring Criteria", scoring_results)
    facet_rows = [
        {
            "filter_id": row["filter_id"],
            "field": row["field"],
            "value": row["value"],
            "status": "AVAILABLE_NOT_AUTO_APPLIED",
            "reviewer_correction": row.get("reviewer_correction", ""),
        }
        for row in user_facets
    ]
    _write_sheet(workbook, "User Facets", facet_rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _build_report(
    manifest: dict[str, Any],
    selected: pd.DataFrame,
    rule_results: Sequence[dict[str, Any]],
    requirement_results: Sequence[dict[str, Any]],
    scoring_results: Sequence[dict[str, Any]],
) -> str:
    lines = [
        "# Phase 7 — BRID hybrid matching report",
        "",
        f"**Label:** `{manifest['label']}`  ",
        f"**Status:** `{manifest['status']}`  ",
        f"**Opportunity:** `{manifest['opportunity_id']}`  ",
        f"**Retrieval:** `{manifest['retrieval_mode']}` "
        f"(`{manifest['fusion_label']}`)  ",
        "",
        "## Recommended portfolio",
        "",
    ]
    for row in selected.to_dict(orient="records"):
        lines.append(
            f"{int(row['portfolio_rank'])}. **{row.get('client_raw') or row['reference_id']}** "
            f"— {row.get('country_code', '')}; {row.get('offering_code', '')}; "
            f"score {float(row.get('final_score', 0)):.3f}; "
            f"{row.get('evidence_type', '')}."
        )
    lines.extend(["", "## Eligibility rules", ""])
    for row in rule_results:
        lines.append(
            f"- `{row['code']}`: {'PASS' if row['passed'] else 'GAP'} "
            f"({row['observed_count']}/{row['required_count']})."
        )
    lines.extend(["", "## Requirement coverage", ""])
    for row in requirement_results:
        lines.append(
            f"- `{row['code']}` ({row['classification']}): "
            f"{'COVERED' if row['portfolio_covered'] else 'GAP'}; "
            f"{row['supporting_reference_count']} supporting reference(s)."
        )
    lines.extend(["", "## Approved 100-point model", ""])
    for row in scoring_results:
        lines.append(
            f"- `{row['code']}` — {row['points_available']} points: "
            f"{row['evidence_status']}; no points auto-awarded."
        )
    lines.extend(
        [
            "",
            "## Decision boundary",
            "",
            "This is a synthetic controlled-development result. It does not award tender",
            "points, approve a business shortlist, validate citation correctness, bypass",
            "Phase 5.1 expert relevance evaluation, or authorize production use.",
            "",
        ]
    )
    return "\n".join(lines)


def _hash_outputs(run_root: Path, success_marker: str) -> Path:
    paths = [
        path
        for path in run_root.iterdir()
        if path.is_file() and path.name not in {"SHA256SUMS.txt", success_marker}
    ]
    text = "".join(
        f"{sha256_file(path)}  {path.name}\n"
        for path in sorted(paths, key=lambda item: item.name)
    )
    sums_path = run_root / "SHA256SUMS.txt"
    _atomic_text(sums_path, text)
    return sums_path


def run_phase7_brid(
    *,
    project_root: Path,
    config_path: Path,
    embedding_adapter: Any | None,
    mode_override: str | None = None,
    output_root: Path | None = None,
) -> tuple[Path, dict[str, Any]]:
    config = load_phase7_brid_config(config_path)
    settings = config["input"]
    phase6_root = (
        project_root
        / "data"
        / "opportunities"
        / settings["opportunity_id"]
        / settings["phase6_run_name"]
    )
    workbook_path = phase6_root / settings["review_workbook_name"]
    review = load_approved_brid_review(workbook_path, config)
    verified = verify_phase5_2_inputs(project_root, config)
    engine_config = _engine_config(verified["phase5_2_config"], config)
    reference_contract = pd.read_parquet(verified["reference_contract_path"])
    phase5_root = verified["pinned"]["phase5_root"]
    mode = str(mode_override or config["retrieval"]["mode"])
    if mode not in {"bm25", "hybrid"}:
        raise ValueError("Phase 7 BRID supports bm25 dry-run or hybrid execution")
    if mode == "hybrid" and embedding_adapter is None:
        raise Phase7BridError("Hybrid execution requires the pinned local E5 adapter")

    hard_filtered_contract, filter_audit = filter_reference_contract(
        reference_contract,
        allowed_security_classifications=config["security"][
            "allowed_classifications_for_controlled_test"
        ],
        hard_filters=review["hard_filters"],
        exclusions=[],
        config=engine_config,
        for_shortlist=True,
    )
    if hard_filtered_contract.empty:
        raise Phase7BridError("No signed-attestation reference survived authorization")
    if not hard_filtered_contract["evidence_type"].isin(
        config["eligibility"]["signed_attestation_evidence_types"]
    ).all():
        raise AssertionError("Unsigned evidence survived the hard gate")

    engine = HardenedMatchingEngine.load(
        phase5_root,
        reference_contract,
        engine_config,
        embedding_adapter=embedding_adapter,
    )
    match = engine.match(
        review["match_requirements"],
        allowed_security_classifications=config["security"][
            "allowed_classifications_for_controlled_test"
        ],
        hard_filters=review["hard_filters"],
        soft_preferences=[],
        exclusions=[],
        mode=mode,
        policy_requirements=review["policy_requirements"],
    )
    candidates = annotate_portfolio_capabilities(match.recommendations)
    if candidates.empty:
        raise Phase7BridError("No evidence-backed candidate reference was retrieved")
    selected = select_portfolio(
        candidates,
        match.coverage,
        review["eligibility_rules"],
        config,
    )
    rule_results = evaluate_portfolio_rules(
        selected,
        review["eligibility_rules"],
    )
    requirement_results = requirement_portfolio_coverage(
        selected,
        match.coverage,
        review["match_requirements"],
    )
    scoring_results = evaluate_scoring_criteria(
        selected,
        review["scoring_criteria"],
        rule_results,
        requirement_results,
        config,
    )
    selected_ids = set(selected["reference_id"].astype(str))
    selected_evidence = match.evidence.loc[
        match.evidence["reference_id"].astype(str).isin(selected_ids)
    ].copy()
    citations = citation_metrics(selected_evidence)
    all_rules_pass = all(row["passed"] for row in rule_results)
    all_must_covered = all(
        row["portfolio_covered"]
        for row in requirement_results
        if row["classification"] == "MUST"
    )
    reference_points = sum(
        int(row["points"])
        for row in review["scoring_criteria"]
        if row["evaluation_route"] in config["scoring"]["reference_routes"]
    )
    human_points = sum(
        int(row["points"])
        for row in review["scoring_criteria"]
        if row["evaluation_route"] in config["scoring"]["human_only_routes"]
    )
    if reference_points + human_points != int(config["scoring"]["tender_points_total"]):
        raise AssertionError("Scoring routes do not reconcile to 100 points")

    run_root = (
        output_root
        if output_root is not None
        else (
            project_root
            / config["output"]["root"]
            / settings["opportunity_id"]
            / config["output"]["run_name"]
        )
    )
    success_path = run_root / config["output"]["success_marker"]
    if success_path.exists():
        verification = verify_phase7_brid_run(run_root, config)
        return run_root, verification["manifest"]
    run_root.mkdir(parents=True, exist_ok=True)

    technical_checks = {
        "approved_workbook_pinned": True,
        "all_requirements_rules_scoring_filters_reviewed": True,
        "phase5_2_signed_inputs_verified": True,
        "authorization_before_scoring": bool(
            filter_audit["authorization_applied_before_scoring"]
        ),
        "signed_attestation_prefilter": bool(
            hard_filtered_contract["evidence_type"]
            .isin(config["eligibility"]["signed_attestation_evidence_types"])
            .all()
        ),
        "portfolio_constraints_not_global_hard_filters": not any(
            field in review["hard_filters"]
            for field in config["eligibility"][
                "global_portfolio_filters_forbidden"
            ]
        ),
        "all_portfolio_rules_evaluated": len(rule_results)
        == int(config["input"]["expected_eligibility_rules"]),
        "approved_scoring_model_consumed": sum(
            row["points_available"] for row in scoring_results
        )
        == 100,
        "tender_points_not_auto_awarded": all(
            row["points_awarded"] is None for row in scoring_results
        ),
        "citation_completeness": citations["citation_completeness"] == 1.0,
        "citation_integrity": citations["citation_integrity"] == 1.0,
        "business_shortlist_not_auto_approved": True,
        "phase5_1_gate_not_bypassed": True,
    }
    if not all(technical_checks.values()):
        failed = [name for name, passed in technical_checks.items() if not passed]
        raise AssertionError(f"Phase 7 BRID technical gates failed: {failed}")
    status = (
        "TECHNICAL_DRY_RUN_BM25_ONLY"
        if mode == "bm25"
        else (
            "TECHNICAL_PASS_READY_FOR_EVIDENCE_AUDIT"
            if all_rules_pass and all_must_covered
            else "TECHNICAL_PASS_WITH_PORTFOLIO_GAPS"
        )
    )
    manifest = {
        "schema_version": 2,
        "phase": 7,
        "pipeline_version": config["pipeline_version"],
        "completed_at_utc": utc_now(),
        "label": config["input"]["required_label"],
        "status": status,
        "opportunity_id": settings["opportunity_id"],
        "approved_workbook_sha256": review["workbook_sha256"],
        "phase5_2_manifest_sha256": config["input"][
            "expected_phase5_2_manifest_sha256"
        ],
        "reference_contract_sha256": config["input"][
            "expected_reference_contract_sha256"
        ],
        "approval_mode": review["approval_mode"],
        "retrieval_mode": mode,
        "fusion_label": (
            config["retrieval"]["development_fusion_label"]
            if mode == "hybrid"
            else "LOCAL_MECHANICAL_DRY_RUN"
        ),
        "lexical_weight": (
            float(config["retrieval"]["lexical_weight"]) if mode == "hybrid" else 1.0
        ),
        "dense_weight": (
            float(config["retrieval"]["dense_weight"]) if mode == "hybrid" else 0.0
        ),
        "match_requirements": len(review["match_requirements"]),
        "policy_requirements": len(review["policy_requirements"]),
        "non_retrieval_requirements": len(review["non_retrieval_requirements"]),
        "eligibility_rules": len(review["eligibility_rules"]),
        "user_facets_available_not_auto_applied": len(review["user_facets"]),
        "authorized_signed_reference_pool": int(len(hard_filtered_contract)),
        "candidate_references": int(len(candidates)),
        "portfolio_references": int(len(selected)),
        "portfolio_rules_passed": sum(int(row["passed"]) for row in rule_results),
        "portfolio_rules_total": len(rule_results),
        "portfolio_eligibility": "PASS" if all_rules_pass else "GAP",
        "must_requirements_covered": sum(
            int(row["portfolio_covered"])
            for row in requirement_results
            if row["classification"] == "MUST"
        ),
        "must_requirements_total": sum(
            int(row["classification"] == "MUST")
            for row in requirement_results
        ),
        "citation_completeness": citations["citation_completeness"],
        "citation_integrity": citations["citation_integrity"],
        "citation_support_proxy": citations["citation_support_proxy"],
        "citation_correctness_status": citations[
            "citation_correctness_status"
        ],
        "tender_points_total": 100,
        "reference_evidence_points_scope": reference_points,
        "human_only_points_scope": human_points,
        "points_awarded_automatically": 0,
        "technical_threshold": int(
            config["scoring"]["expected_technical_threshold"]
        ),
        "technical_threshold_status": "PENDING_HUMAN_SCORING",
        "local_query_embedding_calls": (
            len(review["match_requirements"]) if mode == "hybrid" else 0
        ),
        "external_embedding_api_calls": 0,
        "external_llm_calls": 0,
        "cross_encoder_calls": 0,
        "raw_offer_log_rows": 0,
        "source_mutation_calls": 0,
        "security_filters_applied_before_scoring": True,
        "portfolio_constraints_applied_after_reference_retrieval": True,
        "business_shortlist_auto_approved": False,
        "production_promotion_allowed": False,
        "phase5_1_expert_evaluation_bypassed": False,
    }
    quality_gate = {
        "schema_version": 1,
        "phase": 7,
        "pipeline_version": config["pipeline_version"],
        "technical_gate": "PASS",
        "portfolio_eligibility": manifest["portfolio_eligibility"],
        "must_requirement_union_coverage": (
            "PASS" if all_must_covered else "GAP"
        ),
        "checks": technical_checks,
        "citation_correctness_status": citations[
            "citation_correctness_status"
        ],
        "technical_threshold_status": manifest["technical_threshold_status"],
        "production_promotion": "BLOCKED_PENDING_PHASE_5_1_AND_AUTHORIZED_PILOT",
    }

    candidates.to_parquet(run_root / "candidate_references.parquet", index=False)
    candidates.to_json(
        run_root / "candidate_references.json",
        orient="records",
        force_ascii=False,
        indent=2,
    )
    selected.to_parquet(run_root / "recommended_portfolio.parquet", index=False)
    selected.to_json(
        run_root / "recommended_portfolio.json",
        orient="records",
        force_ascii=False,
        indent=2,
    )
    match.evidence.to_parquet(run_root / "evidence_matrix.parquet", index=False)
    match.evidence.to_json(
        run_root / "evidence_matrix.jsonl",
        orient="records",
        lines=True,
        force_ascii=False,
    )
    match.coverage.to_csv(
        run_root / "candidate_requirement_coverage.csv",
        index=False,
        encoding="utf-8-sig",
    )
    pd.DataFrame(rule_results).to_csv(
        run_root / "portfolio_eligibility_rules.csv",
        index=False,
        encoding="utf-8-sig",
    )
    _atomic_json(run_root / "portfolio_eligibility_rules.json", rule_results)
    pd.DataFrame(requirement_results).to_csv(
        run_root / "portfolio_requirement_coverage.csv",
        index=False,
        encoding="utf-8-sig",
    )
    _atomic_json(
        run_root / "portfolio_requirement_coverage.json",
        requirement_results,
    )
    _atomic_json(run_root / "scoring_criteria_review.json", scoring_results)
    _atomic_json(
        run_root / "user_facets_not_auto_applied.json",
        review["user_facets"],
    )
    _atomic_json(
        run_root / "filter_audit.json",
        {
            **filter_audit,
            "portfolio_filters_not_globally_applied": review[
                "portfolio_filters"
            ],
            "user_facets_not_auto_applied": review["user_facets"],
        },
    )
    _atomic_json(run_root / config["output"]["manifest_name"], manifest)
    _atomic_json(run_root / config["output"]["quality_gate_name"], quality_gate)
    create_phase7_review_workbook(
        run_root / config["output"]["review_workbook_name"],
        manifest=manifest,
        selected=selected,
        candidates=candidates,
        rule_results=rule_results,
        requirement_results=requirement_results,
        evidence=match.evidence,
        scoring_results=scoring_results,
        user_facets=review["user_facets"],
    )
    _atomic_text(
        run_root / config["output"]["report_name"],
        _build_report(
            manifest,
            selected,
            rule_results,
            requirement_results,
            scoring_results,
        ),
    )
    sums_path = _hash_outputs(run_root, config["output"]["success_marker"])
    manifest_path = run_root / config["output"]["manifest_name"]
    _atomic_json(
        success_path,
        {
            "status": "COMPLETE_REPRODUCIBLE_PHASE_7_BRID",
            "created_at_utc": utc_now(),
            "pipeline_version": config["pipeline_version"],
            "opportunity_id": settings["opportunity_id"],
            "manifest_sha256": sha256_file(manifest_path),
            "sha256sums_sha256": sha256_file(sums_path),
        },
    )
    verify_phase7_brid_run(run_root, config)
    return run_root, manifest


def verify_phase7_brid_run(
    run_root: Path,
    config: dict[str, Any],
) -> dict[str, Any]:
    success_path = run_root / config["output"]["success_marker"]
    manifest_path = run_root / config["output"]["manifest_name"]
    quality_path = run_root / config["output"]["quality_gate_name"]
    sums_path = run_root / "SHA256SUMS.txt"
    workbook_path = run_root / config["output"]["review_workbook_name"]
    for path in (
        success_path,
        manifest_path,
        quality_path,
        sums_path,
        workbook_path,
    ):
        if not path.is_file():
            raise FileNotFoundError(path)
    success = json.loads(success_path.read_text(encoding="utf-8"))
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    quality = json.loads(quality_path.read_text(encoding="utf-8"))
    if success["manifest_sha256"] != sha256_file(manifest_path):
        raise AssertionError("Phase 7 BRID manifest hash mismatch")
    if success["sha256sums_sha256"] != sha256_file(sums_path):
        raise AssertionError("Phase 7 BRID checksum-file hash mismatch")
    for line in sums_path.read_text(encoding="utf-8").splitlines():
        expected, name = line.split("  ", 1)
        if sha256_file(run_root / name) != expected:
            raise AssertionError(f"Phase 7 BRID output changed: {name}")
    prohibited_nonzero = (
        "external_embedding_api_calls",
        "external_llm_calls",
        "cross_encoder_calls",
        "raw_offer_log_rows",
        "source_mutation_calls",
        "points_awarded_automatically",
    )
    if any(int(manifest.get(field, -1)) != 0 for field in prohibited_nonzero):
        raise AssertionError("Phase 7 BRID provider/security/scoring boundary failed")
    if not manifest["security_filters_applied_before_scoring"]:
        raise AssertionError("Authorization/filtering order failed")
    if not manifest["portfolio_constraints_applied_after_reference_retrieval"]:
        raise AssertionError("Portfolio constraint layer failed")
    if manifest["business_shortlist_auto_approved"]:
        raise AssertionError("Business shortlist was auto-approved")
    if manifest["production_promotion_allowed"]:
        raise AssertionError("Production promotion was incorrectly enabled")
    if manifest["phase5_1_expert_evaluation_bypassed"]:
        raise AssertionError("Phase 5.1 gate was bypassed")
    if manifest["technical_threshold_status"] != "PENDING_HUMAN_SCORING":
        raise AssertionError("Tender threshold was automatically claimed")
    if quality["technical_gate"] != "PASS":
        raise AssertionError("Phase 7 BRID technical gate is not PASS")
    workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise AssertionError("Executable spreadsheet formula detected")
    return {
        "success": success,
        "manifest": manifest,
        "quality_gate": quality,
    }
