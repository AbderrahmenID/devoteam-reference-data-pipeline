from __future__ import annotations

import copy
import hashlib
import json
import math
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

import numpy as np
import pandas as pd
import yaml

from .phase2_utils import sha256_file
from .phase5_bm25 import BM25Index, normalize_search_text


YEAR_RE = re.compile(r"\b(?:19|20)\d{2}\b")
HASH_RE = re.compile(r"^[0-9a-f]{64}$")
TOKEN_RE = re.compile(r"[\wÀ-ÿ]+", re.UNICODE)
SPREADSHEET_FORMULA_PREFIXES = ("=", "+", "-", "@")
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
STOPWORDS = {
    "a",
    "au",
    "aux",
    "avec",
    "ce",
    "ces",
    "dans",
    "de",
    "des",
    "du",
    "elle",
    "en",
    "et",
    "il",
    "la",
    "le",
    "les",
    "leur",
    "leurs",
    "mais",
    "ou",
    "où",
    "par",
    "pour",
    "que",
    "qui",
    "sur",
    "un",
    "une",
    "the",
    "and",
    "for",
    "from",
    "into",
    "of",
    "to",
    "with",
    "يجب",
    "من",
    "في",
    "على",
    "إلى",
}


class MatchingHardeningError(RuntimeError):
    """Raised when a signed input or a mandatory matching control fails."""


@dataclass
class MatchResult:
    recommendations: pd.DataFrame
    ineligible_candidates: pd.DataFrame
    evidence: pd.DataFrame
    coverage: pd.DataFrame
    facets: dict[str, dict[str, int]]
    filter_audit: dict[str, Any]
    citation_metrics: dict[str, Any]
    policy_requirements: list[dict[str, Any]]


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _atomic_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(value, encoding="utf-8")
    os.replace(temporary, path)


def _atomic_json(path: Path, value: Any) -> None:
    _atomic_text(path, json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n")


def normalize_key(value: Any) -> str:
    """Return a stable accent/case/punctuation-insensitive business key."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = text.casefold().replace("’", "'")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def json_values(value: Any) -> list[str]:
    if isinstance(value, list):
        raw = value
    elif value is None or (isinstance(value, float) and math.isnan(value)):
        raw = []
    else:
        try:
            parsed = json.loads(str(value))
            raw = parsed if isinstance(parsed, list) else [parsed]
        except (TypeError, ValueError, json.JSONDecodeError):
            raw = [value]
    output: list[str] = []
    seen: set[str] = set()
    for item in raw:
        cleaned = " ".join(str(item or "").split())
        key = normalize_key(cleaned)
        if cleaned and key not in seen:
            output.append(cleaned)
            seen.add(key)
    return output


def safe_spreadsheet_value(value: Any) -> Any:
    """Serialize untrusted text without allowing Excel/Sheets formula execution."""
    if value is None or isinstance(value, (bool, int, float, np.integer, np.floating)):
        return value
    text = CONTROL_RE.sub(" ", str(value)).replace("\r", " ").replace("\n", " ")
    if text.lstrip().startswith(SPREADSHEET_FORMULA_PREFIXES):
        return "'" + text
    return text


def safe_spreadsheet_frame(frame: pd.DataFrame) -> pd.DataFrame:
    safe = frame.copy()
    for column in safe.columns:
        if safe[column].dtype == object:
            safe[column] = safe[column].map(safe_spreadsheet_value)
    return safe


def parse_year_interval(value: Any, config: dict) -> dict[str, Any]:
    raw = " ".join(str(value or "").split())
    normalized = normalize_key(raw)
    minimum = int(config["normalization"]["year_minimum"])
    maximum = int(config["normalization"]["year_maximum"])
    years = sorted(
        {
            int(year)
            for year in YEAR_RE.findall(raw)
            if minimum <= int(year) <= maximum
        }
    )
    ongoing = any(
        normalize_key(token) in normalized
        for token in config["normalization"].get("ongoing_tokens", [])
    )
    start = years[0] if years else None
    end = maximum if ongoing else (years[-1] if years else None)
    status = "ONGOING" if ongoing else ("COMPLETED" if years else "UNKNOWN")
    return {
        "year_start": start,
        "year_end": end,
        "ongoing": ongoing,
        "project_status": status,
    }


def _normalized_mapping(mapping: dict[str, Any]) -> dict[str, Any]:
    return {normalize_key(key): value for key, value in mapping.items()}


def canonical_country(value: Any, config: dict) -> dict[str, str]:
    aliases = _normalized_mapping(config["normalization"]["country_aliases"])
    code = aliases.get(normalize_key(value), "")
    metadata = config["normalization"]["countries"].get(code, {})
    return {
        "country_code": code,
        "country_name": str(metadata.get("name") or ""),
        "region": str(metadata.get("region") or ""),
        "subregion": str(metadata.get("subregion") or ""),
    }


def canonical_taxonomy_value(value: Any, mapping: dict[str, str]) -> str:
    normalized = _normalized_mapping(mapping)
    return str(normalized.get(normalize_key(value), "UNKNOWN"))


def detect_tags(value: Any, taxonomy: dict[str, Sequence[str]]) -> list[str]:
    normalized = normalize_key(value)
    output = []
    for tag, aliases in taxonomy.items():
        if any(normalize_key(alias) in normalized for alias in aliases):
            output.append(str(tag))
    return sorted(set(output))


def _document_map(documents: pd.DataFrame) -> dict[str, dict[str, Any]]:
    return {
        str(row["document_id"]): row
        for row in documents.to_dict(orient="records")
    }


def build_reference_contract(
    references: pd.DataFrame,
    documents: pd.DataFrame,
    config: dict,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build a normalized additive contract without modifying Phase 4."""
    required_reference = {
        "reference_id",
        "sheet",
        "row_number",
        "canonical_document_id",
        "country",
        "client",
        "sector",
        "offering",
        "service_nature",
        "project_year",
        "attestation_available",
        "evidence_available",
        "document_retrieval_eligible",
        "data_quality_status",
        "reference_number",
    }
    required_documents = {
        "document_id",
        "source_file_name",
        "source_sha256",
        "security_classification",
        "document_type",
        "document_language",
        "retrieval_eligible",
    }
    missing_reference = required_reference - set(references.columns)
    missing_documents = required_documents - set(documents.columns)
    if missing_reference or missing_documents:
        raise AssertionError(
            f"Reference-contract inputs incomplete: references={sorted(missing_reference)}, "
            f"documents={sorted(missing_documents)}"
        )
    document_map = _document_map(documents)
    evidence_types = config["normalization"]["evidence_types"]
    evidence_strengths = config["normalization"]["evidence_strength"]
    sector_aliases = config["normalization"]["sector_aliases"]
    offering_aliases = config["normalization"]["offering_aliases"]
    capability_taxonomy = config["normalization"]["capability_taxonomy"]
    technology_taxonomy = config["normalization"]["technology_taxonomy"]
    engagement_taxonomy = config["normalization"]["engagement_taxonomy"]
    ordered = references.copy().sort_values(
        ["sheet", "row_number", "reference_id"], kind="stable"
    ).reset_index(drop=True)
    records: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    for index, row in enumerate(ordered.to_dict(orient="records"), start=1):
        reference_id = str(row["reference_id"])
        document_id = str(row.get("canonical_document_id") or "")
        document = document_map.get(document_id, {})
        country = canonical_country(row.get("country"), config)
        year = parse_year_interval(row.get("project_year"), config)
        evidence_type = canonical_taxonomy_value(row.get("attestation_available"), evidence_types)
        if evidence_type == "UNKNOWN":
            evidence_type = "OTHER"
        evidence_strength = float(evidence_strengths.get(evidence_type, evidence_strengths["OTHER"]))
        combined_text = " | ".join(
            str(row.get(field) or "")
            for field in ("service_nature", "offering", "sector", "client")
        )
        capabilities = detect_tags(combined_text, capability_taxonomy)
        technologies = detect_tags(combined_text, technology_taxonomy)
        engagements = detect_tags(combined_text, engagement_taxonomy)
        sector_code = canonical_taxonomy_value(row.get("sector"), sector_aliases)
        offering_code = canonical_taxonomy_value(row.get("offering"), offering_aliases)
        # Two catalogue rows backed by the same evidence document must not occupy
        # two shortlist positions. If evidence is missing, fall back to a strict
        # business signature instead of collapsing unrelated references.
        duplicate_signature = (
            f"document:{document_id}"
            if document_id
            else "metadata:"
            + "|".join(
                [
                    normalize_key(row.get("client")),
                    str(year["year_start"] or ""),
                    str(year["year_end"] or ""),
                    offering_code,
                    normalize_key(row.get("service_nature")),
                ]
            )
        )
        record = {
            "reference_id": reference_id,
            "display_reference_id": f"REF-{index:04d}",
            "reference_number_raw": str(row.get("reference_number") or ""),
            "sheet": str(row.get("sheet") or ""),
            "row_number": int(row.get("row_number") or 0),
            "canonical_document_id": document_id,
            "source_file_name": str(document.get("source_file_name") or row.get("target_name") or ""),
            "source_sha256": str(document.get("source_sha256") or ""),
            "document_type": str(document.get("document_type") or ""),
            "document_language": str(document.get("document_language") or ""),
            "security_classification": str(
                document.get("security_classification") or config["security"]["classification"]
            ),
            "country_raw": str(row.get("country") or ""),
            **country,
            "client_raw": str(row.get("client") or ""),
            "client_key": normalize_key(row.get("client")),
            "sector_raw": str(row.get("sector") or ""),
            "sector_code": sector_code,
            "offering_raw": str(row.get("offering") or ""),
            "offering_code": offering_code,
            "service_nature_raw": str(row.get("service_nature") or ""),
            "project_year_raw": str(row.get("project_year") or ""),
            **year,
            "evidence_type": evidence_type,
            "evidence_strength": evidence_strength,
            "evidence_available": bool(row.get("evidence_available")),
            "document_retrieval_eligible": bool(row.get("document_retrieval_eligible")),
            "document_catalog_retrieval_eligible": bool(document.get("retrieval_eligible")),
            "data_quality_status": str(row.get("data_quality_status") or ""),
            "capability_tags_json": json.dumps(capabilities, ensure_ascii=False),
            "technology_tags_json": json.dumps(technologies, ensure_ascii=False),
            "engagement_tags_json": json.dumps(engagements, ensure_ascii=False),
            "duplicate_signature": duplicate_signature,
            "duplicate_group_id": sha256_text(duplicate_signature)[:16],
        }
        records.append(record)
        if not reference_id:
            issues.append(
                _issue(record, "MISSING_REFERENCE_ID", "BLOCKER", "Stable reference ID is blank.")
            )
        if str(row.get("reference_number") or "").startswith("#"):
            issues.append(
                _issue(
                    record,
                    "INVALID_SOURCE_REFERENCE_NUMBER",
                    "WARNING",
                    f"Preserved raw spreadsheet value: {row.get('reference_number')}",
                )
            )
        if row.get("country") and not country["country_code"]:
            issues.append(
                _issue(
                    record,
                    "UNMAPPED_COUNTRY",
                    "ERROR",
                    f"No ISO mapping for: {row.get('country')}",
                )
            )
        if row.get("project_year") and year["year_start"] is None:
            issues.append(
                _issue(
                    record,
                    "UNPARSEABLE_PROJECT_YEAR",
                    "ERROR",
                    f"Could not parse: {row.get('project_year')}",
                )
            )
        if not row.get("project_year"):
            issues.append(
                _issue(record, "MISSING_PROJECT_YEAR", "WARNING", "Project year is blank.")
            )
        if not bool(row.get("evidence_available")):
            issues.append(
                _issue(
                    record,
                    "MISSING_SUPPORTING_EVIDENCE",
                    "WARNING",
                    "Reference cannot enter the controlled shortlist.",
                )
            )
        if not document:
            issues.append(
                _issue(
                    record,
                    "MISSING_CANONICAL_DOCUMENT",
                    "ERROR",
                    "Reference is not linked to a canonical document.",
                )
            )
    contract = pd.DataFrame(records)
    group_sizes = contract["duplicate_group_id"].value_counts()
    contract["duplicate_group_size"] = contract["duplicate_group_id"].map(group_sizes).astype(int)
    allowed_quality = set(config["eligibility"]["allowed_data_quality_statuses"])
    contract["base_shortlist_eligible"] = (
        contract["reference_id"].ne("")
        & contract["evidence_available"].eq(True)
        & contract["document_retrieval_eligible"].eq(True)
        & contract["document_catalog_retrieval_eligible"].eq(True)
        & contract["data_quality_status"].isin(allowed_quality)
        & contract["source_sha256"].map(lambda value: bool(HASH_RE.fullmatch(str(value))))
    )
    return contract, pd.DataFrame(issues)


def _issue(record: dict[str, Any], code: str, severity: str, detail: str) -> dict[str, Any]:
    return {
        "reference_id": record["reference_id"],
        "display_reference_id": record["display_reference_id"],
        "issue_code": code,
        "severity": severity,
        "detail": detail,
    }


def validate_query(value: Any, config: dict) -> str:
    text = CONTROL_RE.sub(" ", str(value or ""))
    text = " ".join(text.split())
    if not text:
        raise ValueError("Retrieval query is blank")
    maximum = int(config["retrieval"]["query_maximum_characters"])
    if len(text) > maximum:
        raise ValueError(f"Retrieval query exceeds {maximum} characters")
    return text


def query_tokens(value: Any) -> set[str]:
    return {
        normalize_key(token)
        for token in TOKEN_RE.findall(str(value or ""))
        if len(normalize_key(token)) >= 2 and normalize_key(token) not in STOPWORDS
    }


def term_coverage(query: Any, evidence: Any) -> float:
    wanted = query_tokens(query)
    if not wanted:
        return 0.0
    present = query_tokens(evidence)
    return len(wanted & present) / len(wanted)


def _canonical_filter_value(field: str, value: Any, config: dict) -> Any:
    if field == "country_code":
        raw = str(value or "").upper()
        if raw in config["normalization"]["countries"]:
            return raw
        return canonical_country(value, config)["country_code"]
    if field == "sector_code":
        raw = str(value or "").upper()
        if raw in set(config["normalization"]["sector_aliases"].values()):
            return raw
        return canonical_taxonomy_value(value, config["normalization"]["sector_aliases"])
    if field == "offering_code":
        raw = str(value or "").upper()
        if raw in set(config["normalization"]["offering_aliases"].values()):
            return raw
        return canonical_taxonomy_value(value, config["normalization"]["offering_aliases"])
    if field in {
        "technology_tags",
        "capability_tags",
        "engagement_tags",
        "evidence_type",
        "project_status",
        "data_quality_status",
        "security_classification",
    }:
        return str(value or "").upper()
    if field in {"region", "subregion", "client_key"}:
        return normalize_key(value)
    return value


def _json_tag_mask(series: pd.Series, wanted: set[str]) -> pd.Series:
    return series.map(
        lambda raw: bool({str(value).upper() for value in json_values(raw)} & wanted)
    )


def _positive_filter_mask(
    frame: pd.DataFrame,
    field: str,
    requested: Any,
    config: dict,
) -> pd.Series:
    aliases = {
        "technology_tags": "technology_tags_json",
        "capability_tags": "capability_tags_json",
        "engagement_tags": "engagement_tags_json",
    }
    if field in aliases:
        values = requested if isinstance(requested, list) else [requested]
        wanted = {
            str(_canonical_filter_value(field, value, config)).upper()
            for value in values
            if str(value or "").strip()
        }
        return _json_tag_mask(frame[aliases[field]], wanted)
    if field in {
        "country_code",
        "sector_code",
        "offering_code",
        "evidence_type",
        "project_status",
        "data_quality_status",
        "security_classification",
    }:
        values = requested if isinstance(requested, list) else [requested]
        wanted = {
            str(_canonical_filter_value(field, value, config))
            for value in values
            if str(value or "").strip()
        }
        return frame[field].astype(str).isin(wanted)
    if field in {"region", "subregion", "client_key"}:
        values = requested if isinstance(requested, list) else [requested]
        wanted = {
            str(_canonical_filter_value(field, value, config))
            for value in values
            if str(value or "").strip()
        }
        return frame[field].map(normalize_key).isin(wanted)
    if field == "evidence_available":
        return frame[field].eq(bool(requested))
    if field == "evidence_min_strength":
        return frame["evidence_strength"].ge(float(requested))
    raise ValueError(f"Unsupported exact filter: {field}")


def _exclusion_mask(
    frame: pd.DataFrame,
    exclusion: dict[str, Any],
    config: dict,
) -> pd.Series:
    field = str(exclusion.get("field") or "")
    operator = str(exclusion.get("operator") or "eq").lower()
    value = exclusion.get("value")
    if operator in {"eq", "in"}:
        return _positive_filter_mask(frame, field, value, config)
    if field not in frame.columns:
        raise ValueError(f"Unsupported exclusion field: {field}")
    if operator == "contains":
        key = normalize_key(value)
        return frame[field].map(normalize_key).str.contains(re.escape(key), regex=True)
    if operator in {"lt", "lte", "gt", "gte"}:
        values = pd.to_numeric(frame[field], errors="coerce")
        threshold = float(value)
        return {
            "lt": values.lt(threshold),
            "lte": values.le(threshold),
            "gt": values.gt(threshold),
            "gte": values.ge(threshold),
        }[operator]
    if operator == "is_false":
        return frame[field].eq(False)
    if operator == "is_true":
        return frame[field].eq(True)
    raise ValueError(f"Unsupported exclusion operator: {operator}")


def filter_reference_contract(
    contract: pd.DataFrame,
    *,
    allowed_security_classifications: Iterable[str],
    hard_filters: dict[str, Any] | None,
    exclusions: Sequence[dict[str, Any]] | None,
    config: dict,
    for_shortlist: bool = True,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    allowed = {str(value).upper() for value in allowed_security_classifications if str(value).strip()}
    if not allowed:
        raise PermissionError("Explicit security authorization is required before matching")
    mask = contract["security_classification"].astype(str).str.upper().isin(allowed)
    authorized_count = int(mask.sum())
    if for_shortlist:
        mask &= contract["base_shortlist_eligible"].eq(True)
    filters = hard_filters or {}
    supported = {
        "country_code",
        "region",
        "subregion",
        "client_key",
        "sector_code",
        "offering_code",
        "technology_tags",
        "capability_tags",
        "engagement_tags",
        "evidence_type",
        "evidence_available",
        "evidence_min_strength",
        "project_status",
        "data_quality_status",
        "security_classification",
        "year_from",
        "year_to",
    }
    unknown = set(filters) - supported
    if unknown:
        raise ValueError(f"Unsupported hard filters: {sorted(unknown)}")
    for field, requested in filters.items():
        if requested in (None, "", []):
            continue
        if field == "year_from":
            mask &= contract["year_end"].notna() & contract["year_end"].ge(int(requested))
        elif field == "year_to":
            mask &= contract["year_start"].notna() & contract["year_start"].le(int(requested))
        else:
            mask &= _positive_filter_mask(contract, field, requested, config)
    before_exclusions = int(mask.sum())
    excluded_ids: set[str] = set()
    for exclusion in exclusions or []:
        excluded = mask & _exclusion_mask(contract, exclusion, config)
        excluded_ids.update(contract.loc[excluded, "reference_id"].astype(str))
        mask &= ~excluded
    result = contract.loc[mask].copy().reset_index(drop=True)
    audit = {
        "authorization_applied_before_scoring": True,
        "input_references": int(len(contract)),
        "authorized_references": authorized_count,
        "references_after_hard_filters": before_exclusions,
        "references_after_exclusions": int(len(result)),
        "excluded_reference_ids": sorted(excluded_ids),
        "hard_filters": filters,
        "exclusions": list(exclusions or []),
    }
    return result, audit


def facet_counts(frame: pd.DataFrame) -> dict[str, dict[str, int]]:
    facets: dict[str, dict[str, int]] = {}
    for field in (
        "country_code",
        "region",
        "subregion",
        "sector_code",
        "offering_code",
        "evidence_type",
        "project_status",
        "document_language",
    ):
        counts = frame[field].fillna("").astype(str).value_counts()
        facets[field] = {
            key: int(value)
            for key, value in counts.items()
            if key
        }
    for output, column in (
        ("technology_tags", "technology_tags_json"),
        ("capability_tags", "capability_tags_json"),
        ("engagement_tags", "engagement_tags_json"),
    ):
        counts: dict[str, int] = {}
        for raw in frame[column]:
            for value in json_values(raw):
                counts[value] = counts.get(value, 0) + 1
        facets[output] = dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))
    return facets


def split_policy_requirements(
    requirements: Sequence[dict[str, Any]],
    config: dict,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    patterns = [normalize_key(value) for value in config["eligibility"]["evidence_policy_patterns"]]
    content: list[dict[str, Any]] = []
    policies: list[dict[str, Any]] = []
    for requirement in requirements:
        normalized = normalize_key(requirement.get("requirement_text"))
        if any(pattern in normalized for pattern in patterns):
            policies.append(
                {
                    **requirement,
                    "requirement_kind": "ELIGIBILITY_POLICY",
                    "policy": "SUPPORTING_EVIDENCE_REQUIRED",
                }
            )
        else:
            content.append({**requirement, "requirement_kind": "CONTENT"})
    return content, policies


def compile_phase6_controls(
    requirements: Sequence[dict[str, Any]],
    proposals: Sequence[dict[str, Any]],
    config: dict,
) -> dict[str, Any]:
    content, policies = split_policy_requirements(requirements, config)
    hard_filters: dict[str, Any] = {}
    soft_preferences: list[dict[str, Any]] = []
    exclusions: list[dict[str, Any]] = []
    field_map = {
        "country": "country_code",
        "sector": "sector_code",
        "offering": "offering_code",
        "client": "client_key",
        "year_after": "year_from",
        "year_before": "year_to",
        "attestation_available": "evidence_type",
    }
    for proposal in proposals:
        behavior = str(proposal.get("proposed_behavior") or "")
        raw_field = str(proposal.get("field") or "")
        field = field_map.get(raw_field, raw_field)
        value = proposal.get("value")
        if field in {"country_code", "sector_code", "offering_code", "client_key"}:
            value = _canonical_filter_value(field, value, config)
        if behavior == "HARD_CANDIDATE":
            existing = hard_filters.get(field)
            if field == "year_from":
                hard_filters[field] = max(int(existing), int(value)) if existing is not None else int(value)
            elif field == "year_to":
                hard_filters[field] = min(int(existing), int(value)) if existing is not None else int(value)
            elif existing is None:
                hard_filters[field] = value
            elif isinstance(existing, list):
                if value not in existing:
                    existing.append(value)
            elif existing != value:
                hard_filters[field] = [existing, value]
        elif behavior in {"SOFT_PREFERENCE", "CONTEXT_ONLY"}:
            soft_preferences.append({"field": field, "value": value})
        elif behavior == "EXCLUSION_CANDIDATE":
            exclusions.append({"field": field, "operator": "eq", "value": value})
    if any(policy["policy"] == "SUPPORTING_EVIDENCE_REQUIRED" for policy in policies):
        exclusions.append(
            {"field": "evidence_available", "operator": "is_false", "value": False}
        )
    return {
        "requirements": content,
        "policy_requirements": policies,
        "hard_filters": hard_filters,
        "soft_preferences": soft_preferences,
        "exclusions": exclusions,
    }


def _stable_top(
    scores: np.ndarray,
    mask: np.ndarray,
    tie_ids: Sequence[str],
    limit: int,
) -> list[int]:
    candidates = [
        index
        for index in np.flatnonzero(mask)
        if np.isfinite(float(scores[index]))
    ]
    candidates.sort(key=lambda index: (-float(scores[index]), str(tie_ids[index])))
    return candidates[:limit]


@dataclass
class HardenedMatchingEngine:
    chunks: pd.DataFrame
    bm25: BM25Index
    embeddings: np.ndarray
    reference_contract: pd.DataFrame
    config: dict
    embedding_adapter: Any | None = None

    def __post_init__(self) -> None:
        self.chunks = self.chunks.reset_index(drop=True)
        self.embeddings = np.asarray(self.embeddings, dtype=np.float32)
        if self.bm25.document_count != len(self.chunks):
            raise AssertionError("BM25 and chunk lookup row counts differ")
        if len(self.embeddings) != len(self.chunks):
            raise AssertionError("Embedding and chunk lookup row counts differ")
        if self.chunks["vector_row"].astype(int).tolist() != list(range(len(self.chunks))):
            raise AssertionError("Chunk vector rows are not stable and contiguous")
        self.tie_ids = self.chunks["chunk_id"].astype(str).tolist()
        self.document_to_references = (
            self.reference_contract.groupby("canonical_document_id")["reference_id"]
            .apply(lambda values: sorted(set(values.astype(str))))
            .to_dict()
        )
        self.last_search_audit: dict[str, Any] = {}

    @classmethod
    def load(
        cls,
        phase5_root: Path,
        reference_contract: pd.DataFrame,
        config: dict,
        embedding_adapter: Any | None = None,
    ) -> "HardenedMatchingEngine":
        return cls(
            chunks=pd.read_parquet(phase5_root / "chunk_lookup.parquet"),
            bm25=BM25Index.load(
                phase5_root / "bm25_index.npz",
                phase5_root / "bm25_vocabulary.json",
            ),
            embeddings=np.load(phase5_root / "embeddings.npy", mmap_mode="r"),
            reference_contract=reference_contract,
            config=config,
            embedding_adapter=embedding_adapter,
        )

    def _query_vector(self, query: str, provided: np.ndarray | None) -> np.ndarray:
        if provided is None:
            if self.embedding_adapter is None:
                raise ValueError("Dense/hybrid matching requires a local query embedding adapter")
            provided = self.embedding_adapter.encode_queries([query])[0]
        vector = np.asarray(provided, dtype=np.float32).reshape(-1)
        if vector.shape != (self.embeddings.shape[1],):
            raise ValueError("Query embedding has the wrong dimension")
        norm = float(np.linalg.norm(vector))
        if not np.isfinite(vector).all() or not math.isclose(norm, 1.0, abs_tol=2e-3):
            raise ValueError("Query embedding must be a finite unit vector")
        return vector

    def search_chunks(
        self,
        query: str,
        *,
        allowed_security_classifications: Iterable[str],
        hard_filters: dict[str, Any] | None = None,
        exclusions: Sequence[dict[str, Any]] | None = None,
        mode: str | None = None,
        top_k: int | None = None,
        query_vector: np.ndarray | None = None,
    ) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
        mode = str(mode or self.config["retrieval"]["default_mode"])
        if mode not in set(self.config["retrieval"]["supported_modes"]):
            raise ValueError(f"Unsupported retrieval mode: {mode}")
        query = validate_query(query, self.config)
        eligible, filter_audit = filter_reference_contract(
            self.reference_contract,
            allowed_security_classifications=allowed_security_classifications,
            hard_filters=hard_filters,
            exclusions=exclusions,
            config=self.config,
            for_shortlist=True,
        )
        allowed_documents = set(eligible["canonical_document_id"].astype(str))
        allowed_classes = {
            str(value).upper()
            for value in allowed_security_classifications
            if str(value).strip()
        }
        authorization_mask = (
            self.chunks["security_classification"].astype(str).str.upper().isin(allowed_classes)
            & self.chunks["document_id"].astype(str).isin(allowed_documents)
        ).to_numpy(dtype=bool)
        filter_audit["authorized_chunk_rows_before_scoring"] = int(authorization_mask.sum())
        filter_audit["total_chunk_rows"] = int(len(authorization_mask))
        if not authorization_mask.any():
            self.last_search_audit = filter_audit
            return pd.DataFrame(), eligible, filter_audit
        fusion = self.config["retrieval"]["fusion"]
        requested_top_k = int(top_k or self.config["retrieval"]["candidate_chunks_per_requirement"])
        candidate_pool = max(
            requested_top_k,
            int(fusion["candidate_pool_per_retriever"]),
        )
        lexical_scores = self.bm25.score(query, allowed_mask=authorization_mask)
        dense_scores = np.full(len(self.chunks), -np.inf, dtype=np.float32)
        if mode in {"dense", "hybrid"}:
            vector = self._query_vector(query, query_vector)
            rows = np.flatnonzero(authorization_mask)
            dense_scores[rows] = self.embeddings[rows] @ vector
        lexical_rows = _stable_top(
            lexical_scores, authorization_mask, self.tie_ids, candidate_pool
        )
        dense_rows = _stable_top(
            dense_scores, authorization_mask, self.tie_ids, candidate_pool
        )
        lexical_rank = {row: rank for rank, row in enumerate(lexical_rows, start=1)}
        dense_rank = {row: rank for rank, row in enumerate(dense_rows, start=1)}
        if mode == "bm25":
            scores = {row: float(lexical_scores[row]) for row in lexical_rows}
        elif mode == "dense":
            scores = {row: float(dense_scores[row]) for row in dense_rows}
        else:
            rrf_k = float(fusion["rrf_k"])
            lexical_weight = float(fusion["lexical_weight"])
            dense_weight = float(fusion["dense_weight"])
            scores: dict[int, float] = {}
            for row in set(lexical_rows) | set(dense_rows):
                score = 0.0
                if row in lexical_rank:
                    score += lexical_weight / (rrf_k + lexical_rank[row])
                if row in dense_rank:
                    score += dense_weight / (rrf_k + dense_rank[row])
                scores[row] = score
        ordered = sorted(scores, key=lambda row: (-scores[row], self.tie_ids[row]))[
            :requested_top_k
        ]
        records: list[dict[str, Any]] = []
        for rank, row_index in enumerate(ordered, start=1):
            record = self.chunks.iloc[row_index].to_dict()
            record.update(
                {
                    "retrieval_mode": mode,
                    "retrieval_rank": rank,
                    "retrieval_score": float(scores[row_index]),
                    "bm25_score": float(lexical_scores[row_index]),
                    "dense_score": (
                        float(dense_scores[row_index])
                        if np.isfinite(dense_scores[row_index])
                        else None
                    ),
                    "bm25_rank": lexical_rank.get(row_index),
                    "dense_rank": dense_rank.get(row_index),
                }
            )
            records.append(record)
        result = pd.DataFrame(records)
        filter_audit.update(
            {
                "retrieval_mode": mode,
                "scored_chunk_rows": int(len(ordered)),
                "all_scored_rows_authorized": bool(
                    all(authorization_mask[int(row)] for row in result["vector_row"])
                )
                if not result.empty
                else True,
                "dense_query_embedding_calls": int(mode in {"dense", "hybrid"}),
                "external_embedding_api_calls": 0,
            }
        )
        self.last_search_audit = filter_audit
        return result, eligible, filter_audit

    def match(
        self,
        requirements: Sequence[dict[str, Any]],
        *,
        allowed_security_classifications: Iterable[str],
        hard_filters: dict[str, Any] | None = None,
        soft_preferences: Sequence[dict[str, Any]] | None = None,
        exclusions: Sequence[dict[str, Any]] | None = None,
        mode: str | None = None,
        query_vectors: dict[str, np.ndarray] | None = None,
        policy_requirements: Sequence[dict[str, Any]] | None = None,
    ) -> MatchResult:
        if not requirements:
            raise MatchingHardeningError("No content requirements remain after policy separation")
        all_evidence: list[dict[str, Any]] = []
        combined_audit: dict[str, Any] = {
            "authorization_applied_before_scoring": True,
            "requirements_searched": len(requirements),
            "dense_query_embedding_calls": 0,
            "external_embedding_api_calls": 0,
        }
        eligible_contract: pd.DataFrame | None = None
        for requirement in requirements:
            requirement_id = str(requirement["requirement_id"])
            chunks, eligible, audit = self.search_chunks(
                requirement["requirement_text"],
                allowed_security_classifications=allowed_security_classifications,
                hard_filters=hard_filters,
                exclusions=exclusions,
                mode=mode,
                query_vector=(query_vectors or {}).get(requirement_id),
            )
            eligible_contract = eligible
            combined_audit["dense_query_embedding_calls"] += int(
                audit["dense_query_embedding_calls"]
            )
            combined_audit["all_scored_rows_authorized"] = bool(
                combined_audit.get("all_scored_rows_authorized", True)
                and audit.get("all_scored_rows_authorized", True)
            )
            if chunks.empty:
                continue
            maximum = max(float(chunks["retrieval_score"].max()), 1e-12)
            by_reference: dict[str, list[dict[str, Any]]] = {}
            eligible_ids = set(eligible["reference_id"].astype(str))
            for chunk in chunks.to_dict(orient="records"):
                retrieval_normalized = min(
                    1.0, max(0.0, float(chunk["retrieval_score"]) / maximum)
                )
                support = term_coverage(
                    requirement["requirement_text"], chunk["chunk_text"]
                )
                rerank = (
                    float(
                        self.config["retrieval"]["reranking"]["retrieval_weight"]
                    )
                    * retrieval_normalized
                    + float(
                        self.config["retrieval"]["reranking"][
                            "term_coverage_weight"
                        ]
                    )
                    * support
                )
                for reference_id in self.document_to_references.get(
                    str(chunk["document_id"]), []
                ):
                    if reference_id not in eligible_ids:
                        continue
                    by_reference.setdefault(reference_id, []).append(
                        {
                            "reference_id": reference_id,
                            "requirement_id": requirement_id,
                            "requirement_classification": str(
                                requirement["classification"]
                            ),
                            "requirement_text": str(
                                requirement["requirement_text"]
                            ),
                            "document_id": str(chunk["document_id"]),
                            "document_type": str(chunk["document_type"]),
                            "data_quality_status": str(
                                chunk["data_quality_status"]
                            ),
                            "security_classification": str(
                                chunk["security_classification"]
                            ),
                            "chunk_id": str(chunk["chunk_id"]),
                            "chunk_text": str(chunk["chunk_text"]),
                            "chunk_text_sha256": str(
                                chunk["chunk_text_sha256"]
                            ),
                            "source_sha256": str(chunk["source_sha256"]),
                            "source_file_name": str(
                                chunk["source_file_name"]
                            ),
                            "page_number_1_based": int(
                                chunk["page_number_1_based"]
                            ),
                            "citation_label": str(
                                chunk["citation_label"]
                            ),
                            "citation_uri": str(chunk["citation_uri"]),
                            "retrieval_mode": str(chunk["retrieval_mode"]),
                            "retrieval_rank": int(chunk["retrieval_rank"]),
                            "retrieval_score": float(
                                chunk["retrieval_score"]
                            ),
                            "retrieval_score_normalized": retrieval_normalized,
                            "term_coverage": support,
                            "rerank_score": rerank,
                        }
                    )
            limit = int(
                self.config["retrieval"][
                    "evidence_chunks_per_reference_requirement"
                ]
            )
            for rows in by_reference.values():
                rows.sort(
                    key=lambda row: (
                        -row["rerank_score"],
                        row["retrieval_rank"],
                        row["chunk_id"],
                    )
                )
                all_evidence.extend(rows[:limit])
        if eligible_contract is None:
            eligible_contract, audit = filter_reference_contract(
                self.reference_contract,
                allowed_security_classifications=allowed_security_classifications,
                hard_filters=hard_filters,
                exclusions=exclusions,
                config=self.config,
            )
            combined_audit.update(audit)
        evidence = pd.DataFrame(all_evidence)
        if not evidence.empty:
            evidence = evidence.sort_values(
                ["requirement_id", "rerank_score", "reference_id", "chunk_id"],
                ascending=[True, False, True, True],
            ).reset_index(drop=True)
        recommendations, ineligible, coverage = score_matches(
            requirements=requirements,
            evidence=evidence,
            eligible_contract=eligible_contract,
            soft_preferences=list(soft_preferences or []),
            config=self.config,
        )
        metrics = citation_metrics(evidence)
        return MatchResult(
            recommendations=recommendations,
            ineligible_candidates=ineligible,
            evidence=evidence,
            coverage=coverage,
            facets=facet_counts(eligible_contract),
            filter_audit=combined_audit,
            citation_metrics=metrics,
            policy_requirements=list(policy_requirements or []),
        )


def _soft_preference_fit(
    reference: dict[str, Any],
    preferences: Sequence[dict[str, Any]],
    config: dict,
) -> float:
    if not preferences:
        return 1.0
    matches = 0
    applicable = 0
    for preference in preferences:
        field = str(preference.get("field") or "")
        value = preference.get("value")
        if field not in reference:
            continue
        applicable += 1
        canonical = _canonical_filter_value(field, value, config)
        actual = reference.get(field)
        if field in {"region", "subregion", "client_key"}:
            matches += int(normalize_key(actual) == normalize_key(canonical))
        else:
            matches += int(str(actual) == str(canonical))
    return matches / applicable if applicable else 1.0


def _recency_score(year_end: Any, reference_year: int) -> float:
    if year_end is None or pd.isna(year_end):
        return 0.25
    latest = min(int(year_end), reference_year)
    return max(0.0, 1.0 - (reference_year - latest) / 20.0)


def score_matches(
    *,
    requirements: Sequence[dict[str, Any]],
    evidence: pd.DataFrame,
    eligible_contract: pd.DataFrame,
    soft_preferences: Sequence[dict[str, Any]],
    config: dict,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if evidence.empty:
        empty = pd.DataFrame()
        return empty, empty, empty
    weights = config["scoring"]["requirement_weights"]
    components = config["scoring"]["component_weights"]
    threshold = float(config["scoring"]["coverage_threshold"])
    total_weight = sum(float(weights[row["classification"]]) for row in requirements)
    contract_map = eligible_contract.set_index("reference_id").to_dict(orient="index")
    best_evidence = (
        evidence.sort_values(
            ["reference_id", "requirement_id", "rerank_score", "chunk_id"],
            ascending=[True, True, False, True],
        )
        .drop_duplicates(["reference_id", "requirement_id"], keep="first")
    )
    lookup = {
        (str(row["reference_id"]), str(row["requirement_id"])): row
        for row in best_evidence.to_dict(orient="records")
    }
    candidates = sorted(set(evidence["reference_id"].astype(str)))
    recommendation_rows: list[dict[str, Any]] = []
    coverage_rows: list[dict[str, Any]] = []
    for reference_id in candidates:
        metadata = {"reference_id": reference_id, **dict(contract_map[reference_id])}
        relevance_sum = 0.0
        covered_weight = 0.0
        must_total = 0
        must_covered = 0
        for requirement in requirements:
            requirement_id = str(requirement["requirement_id"])
            weight = float(weights[requirement["classification"]])
            hit = lookup.get((reference_id, requirement_id))
            relevance = float(hit["rerank_score"]) if hit else 0.0
            covered = bool(hit) and relevance >= threshold
            relevance_sum += weight * relevance
            covered_weight += weight * int(covered)
            if requirement["classification"] == "MUST":
                must_total += 1
                must_covered += int(covered)
            coverage_rows.append(
                {
                    "reference_id": reference_id,
                    "requirement_id": requirement_id,
                    "classification": requirement["classification"],
                    "requirement_text": requirement["requirement_text"],
                    "covered": covered,
                    "relevance_score": relevance,
                    "citation_uri": hit["citation_uri"] if hit else "",
                    "chunk_id": hit["chunk_id"] if hit else "",
                }
            )
        relevance_score = relevance_sum / max(total_weight, 1e-12)
        coverage_score = covered_weight / max(total_weight, 1e-12)
        evidence_quality = float(metadata["evidence_strength"])
        recency = _recency_score(
            metadata.get("year_end"), int(config["scoring"]["reference_year"])
        )
        soft_fit = _soft_preference_fit(metadata, soft_preferences, config)
        final_score = (
            float(components["relevance"]) * relevance_score
            + float(components["coverage"]) * coverage_score
            + float(components["evidence_quality"]) * evidence_quality
            + float(components["recency"]) * recency
            + float(components["soft_preference_fit"]) * soft_fit
        )
        reasons: list[str] = []
        if config["eligibility"]["must_requirement_gate"] and must_covered < must_total:
            reasons.append(f"MISSING_MUST:{must_covered}/{must_total}")
        if not bool(metadata["base_shortlist_eligible"]):
            reasons.append("BASE_ELIGIBILITY_FAILED")
        recommendation_rows.append(
            {
                **metadata,
                "final_score": round(final_score, 8),
                "weighted_relevance": round(relevance_score, 8),
                "weighted_coverage": round(coverage_score, 8),
                "must_covered": must_covered,
                "must_total": must_total,
                "recency_score": round(recency, 8),
                "soft_preference_fit": round(soft_fit, 8),
                "eligible_for_shortlist": not reasons,
                "eligibility_reasons": ";".join(reasons),
                "recommendation_basis": (
                    f"Evidence-backed {evidence.iloc[0]['retrieval_mode']} score; "
                    f"{must_covered}/{must_total} content MUST requirements covered."
                ),
            }
        )
    ranked = pd.DataFrame(recommendation_rows).sort_values(
        ["final_score", "reference_id"], ascending=[False, True]
    )
    eligible = ranked.loc[ranked["eligible_for_shortlist"].eq(True)].copy()
    ineligible = ranked.loc[ranked["eligible_for_shortlist"].eq(False)].copy()
    if config["eligibility"]["maximum_one_per_duplicate_group"] and not eligible.empty:
        duplicate = eligible.duplicated("duplicate_group_id", keep="first")
        suppressed = eligible.loc[duplicate].copy()
        if not suppressed.empty:
            suppressed["eligible_for_shortlist"] = False
            suppressed["eligibility_reasons"] = "DUPLICATE_GROUP_SUPPRESSED"
            ineligible = pd.concat([ineligible, suppressed], ignore_index=True)
        eligible = eligible.loc[~duplicate].copy()
    maximum = int(config["scoring"]["maximum_recommendations"])
    eligible = eligible.head(maximum).reset_index(drop=True)
    eligible["final_rank"] = np.arange(1, len(eligible) + 1)
    return eligible, ineligible.reset_index(drop=True), pd.DataFrame(coverage_rows)


def citation_metrics(evidence: pd.DataFrame) -> dict[str, Any]:
    if evidence.empty:
        return {
            "citation_completeness": 0.0,
            "citation_integrity": 0.0,
            "citation_support_proxy": 0.0,
            "citation_correctness_status": "PENDING_HUMAN_AUDIT",
        }
    required = [
        "reference_id",
        "requirement_id",
        "document_id",
        "chunk_id",
        "chunk_text",
        "chunk_text_sha256",
        "source_sha256",
        "page_number_1_based",
        "citation_label",
        "citation_uri",
    ]
    missing = set(required) - set(evidence.columns)
    if missing:
        raise AssertionError(f"Citation fields missing: {sorted(missing)}")
    complete = evidence[required].apply(
        lambda column: column.map(
            lambda value: value not in (None, "") and not pd.isna(value)
        )
    ).all(axis=1)
    integrity = evidence.apply(
        lambda row: (
            sha256_text(str(row["chunk_text"])) == str(row["chunk_text_sha256"])
            and bool(HASH_RE.fullmatch(str(row["source_sha256"])))
            and int(row["page_number_1_based"]) > 0
            and str(row["citation_uri"]).startswith("https://")
        ),
        axis=1,
    )
    return {
        "citation_completeness": float(complete.mean()),
        "citation_integrity": float(integrity.mean()),
        "citation_support_proxy": float(
            evidence["term_coverage"].astype(float).mean()
        ),
        "citation_correctness_status": "PENDING_HUMAN_AUDIT",
        "citation_rows": int(len(evidence)),
    }


def create_citation_audit_workbook(
    path: Path,
    evidence: pd.DataFrame,
    sample_size: int,
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = openpyxl.Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    for row in (
        ("Purpose", "Human verification that cited evidence supports each requirement."),
        ("Decision", "Mark SUPPORTS, PARTIAL, DOES_NOT_SUPPORT, or UNREADABLE."),
        ("Boundary", "Automated completeness/integrity checks are not semantic correctness."),
    ):
        instructions.append([safe_spreadsheet_value(value) for value in row])
    audit = workbook.create_sheet("Citation Audit")
    headers = [
        "reference_id",
        "requirement_id",
        "requirement_classification",
        "requirement_text",
        "source_file_name",
        "page_number_1_based",
        "citation_uri",
        "chunk_id",
        "chunk_text",
        "term_coverage",
        "human_support_decision",
        "reviewer_notes",
    ]
    audit.append(headers)
    if not evidence.empty:
        sampled = evidence.assign(
            _sample_key=evidence.apply(
                lambda row: sha256_text(
                    f"{row['reference_id']}|{row['requirement_id']}|{row['chunk_id']}"
                ),
                axis=1,
            )
        ).sort_values(
            ["requirement_classification", "_sample_key"], kind="stable"
        ).head(sample_size)
        for row in sampled.to_dict(orient="records"):
            values = [row.get(header, "") for header in headers[:-2]]
            audit.append(
                [safe_spreadsheet_value(value) for value in values]
                + ["PENDING", ""]
            )
    validation = DataValidation(
        type="list",
        formula1='"PENDING,SUPPORTS,PARTIAL,DOES_NOT_SUPPORT,UNREADABLE"',
    )
    audit.add_data_validation(validation)
    validation.add(f"K2:K{max(2, audit.max_row)}")
    widths = {
        "Instructions": [24, 105],
        "Citation Audit": [24, 22, 17, 75, 36, 14, 65, 28, 90, 16, 24, 55],
    }
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B2C83")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, width in enumerate(widths[sheet.title], start=1):
            sheet.column_dimensions[
                openpyxl.utils.get_column_letter(index)
            ].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def load_phase5_2_config(path: Path) -> dict:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    if str(config.get("phase")) != "5.2":
        raise ValueError("Expected Phase 5.2 configuration")
    security = config.get("security", {})
    blocked = (
        "external_llm_enabled",
        "external_embedding_api_enabled",
        "raw_offer_logging_allowed",
        "security_filters_disableable",
        "source_mutation_allowed",
    )
    if any(security.get(field) for field in blocked):
        raise ValueError("Phase 5.2 security defaults were weakened")
    if not security.get("authorization_before_scoring_required"):
        raise ValueError("Authorization must precede retrieval scoring")
    fusion = config["retrieval"]["fusion"]
    if not math.isclose(
        float(fusion["lexical_weight"]) + float(fusion["dense_weight"]), 1.0
    ):
        raise ValueError("Hybrid fusion weights must sum to one")
    reranking = config["retrieval"]["reranking"]
    if not math.isclose(
        float(reranking["retrieval_weight"])
        + float(reranking["term_coverage_weight"]),
        1.0,
    ):
        raise ValueError("Deterministic reranking weights must sum to one")
    if reranking.get("cross_encoder_enabled"):
        raise ValueError("Cross-encoder promotion remains blocked pending expert evaluation")
    if not config["evaluation"]["expert_gold_set_required_for_promotion"]:
        raise ValueError("Expert evaluation must remain a production gate")
    return config


def verify_pinned_inputs(project_root: Path, config: dict) -> dict[str, Path]:
    settings = config["input"]
    phase4_root = (
        project_root
        / "data"
        / "canonical"
        / settings["snapshot_id"]
        / settings["phase4_run_name"]
    )
    phase5_root = (
        project_root
        / "data"
        / "indexes"
        / settings["snapshot_id"]
        / settings["phase5_run_name"]
    )
    paths = {
        "phase4_root": phase4_root,
        "phase5_root": phase5_root,
        "phase4_manifest": phase4_root / "PHASE_4_MANIFEST.json",
        "references": phase4_root / "reference_catalog.parquet",
        "documents": phase4_root / "documents_catalog.parquet",
        "phase5_manifest": phase5_root / "PHASE_5_MANIFEST.json",
        "phase5_sums": phase5_root / "SHA256SUMS.txt",
        "chunk_lookup": phase5_root / "chunk_lookup.parquet",
        "bm25_index": phase5_root / "bm25_index.npz",
        "bm25_vocabulary": phase5_root / "bm25_vocabulary.json",
        "embeddings": phase5_root / "embeddings.npy",
    }
    for path in paths.values():
        if not path.exists():
            raise FileNotFoundError(path)
    expected = {
        "phase4_manifest": "expected_phase4_manifest_sha256",
        "references": "expected_phase4_references_sha256",
        "documents": "expected_phase4_documents_sha256",
        "phase5_manifest": "expected_phase5_manifest_sha256",
        "phase5_sums": "expected_phase5_sha256sums_sha256",
        "chunk_lookup": "expected_chunk_lookup_sha256",
        "bm25_index": "expected_bm25_index_sha256",
        "bm25_vocabulary": "expected_bm25_vocabulary_sha256",
        "embeddings": "expected_embeddings_sha256",
    }
    for name, key in expected.items():
        if sha256_file(paths[name]) != settings[key]:
            raise AssertionError(f"Pinned input changed: {name}")
    phase5_manifest = json.loads(paths["phase5_manifest"].read_text(encoding="utf-8"))
    if phase5_manifest.get("status") != "TECHNICAL_PASS":
        raise AssertionError("Phase 5 technical status is not PASS")
    if int(phase5_manifest.get("source_snapshot_mutation_calls", -1)) != 0:
        raise AssertionError("Phase 5 source-mutation boundary failed")
    return paths


def evaluate_fusion_sweep(
    *,
    phase5_chunks: pd.DataFrame,
    bm25: BM25Index,
    embeddings: np.ndarray,
    phase5_config: dict,
    phase5_2_config: dict,
    probes: pd.DataFrame,
    query_vectors: np.ndarray,
) -> pd.DataFrame:
    from .phase5_retrieval import HybridSearchEngine, evaluate_bootstrap

    records = []
    for candidate in phase5_2_config["retrieval"]["fusion_sweep"]:
        candidate_config = copy.deepcopy(phase5_config)
        candidate_config["hybrid"]["lexical_weight"] = float(
            candidate["lexical_weight"]
        )
        candidate_config["hybrid"]["dense_weight"] = float(
            candidate["dense_weight"]
        )
        engine = HybridSearchEngine(
            chunks=phase5_chunks,
            retrieval_texts=[""] * len(phase5_chunks),
            bm25=bm25,
            embeddings=embeddings,
            config=candidate_config,
        )
        _, metrics = evaluate_bootstrap(
            probes=probes,
            query_vectors=query_vectors,
            engine=engine,
        )
        hybrid = metrics["retrieval_metrics"]["hybrid"]
        records.append(
            {
                "candidate": candidate["name"],
                "lexical_weight": float(candidate["lexical_weight"]),
                "dense_weight": float(candidate["dense_weight"]),
                **hybrid,
                "label_source": metrics["label_source"],
                "expert_labeled": False,
                "production_selection_allowed": False,
            }
        )
    return pd.DataFrame(records).sort_values(
        ["ndcg_at_10", "recall_at_10", "lexical_weight"],
        ascending=[False, False, False],
    ).reset_index(drop=True)


def compare_v1_v2(
    baseline_path: Path,
    baseline_manifest_path: Path,
    result: MatchResult,
    mode: str,
) -> dict[str, Any]:
    baseline = (
        json.loads(baseline_path.read_text(encoding="utf-8"))
        if baseline_path.exists()
        else []
    )
    baseline_manifest = (
        json.loads(baseline_manifest_path.read_text(encoding="utf-8"))
        if baseline_manifest_path.exists()
        else {}
    )
    baseline_ids = [str(row.get("reference_id")) for row in baseline]
    v2_ids = (
        result.recommendations["reference_id"].astype(str).tolist()
        if not result.recommendations.empty
        else []
    )
    baseline_missing_must = sum(
        int(row.get("must_covered", 0)) < int(row.get("must_total", 0))
        for row in baseline
    )
    overlap = len(set(baseline_ids) & set(v2_ids))
    union = len(set(baseline_ids) | set(v2_ids))
    return {
        "comparison_scope": "SYNTHETIC_SAMPLE_TECHNICAL_ONLY",
        "expert_labeled": False,
        "production_quality_claim_allowed": False,
        "v1": {
            "retrieval_mode": baseline_manifest.get(
                "retrieval_mode", "bm25_secure_baseline"
            ),
            "recommendations": len(baseline_ids),
            "recommendations_missing_must": baseline_missing_must,
            "dense_query_embedding_calls": int(
                baseline_manifest.get("dense_query_embedding_calls", 0)
            ),
        },
        "v2": {
            "retrieval_mode": mode,
            "recommendations": len(v2_ids),
            "recommendations_missing_must": int(
                (
                    result.recommendations["must_covered"]
                    < result.recommendations["must_total"]
                ).sum()
            )
            if not result.recommendations.empty
            else 0,
            "dense_query_embedding_calls": int(
                result.filter_audit["dense_query_embedding_calls"]
            ),
            "policy_requirements_separated": len(result.policy_requirements),
            "citation_completeness": result.citation_metrics[
                "citation_completeness"
            ],
            "citation_integrity": result.citation_metrics["citation_integrity"],
            "citation_correctness_status": result.citation_metrics[
                "citation_correctness_status"
            ],
        },
        "top_reference_overlap": overlap,
        "top_reference_jaccard": overlap / union if union else 0.0,
        "decision": "KEEP_V1_AS_BASELINE_AND_USE_V2_FOR_CONTROLLED_DEVELOPMENT",
        "promotion_blocker": "PHASE_5_1_EXPERT_GOLD_SET",
    }


def _quality_gate(
    contract: pd.DataFrame,
    result: MatchResult,
    config: dict,
) -> dict[str, Any]:
    thresholds = config["evaluation"]["technical_gates"]
    nonempty_country = contract["country_raw"].ne("")
    nonempty_year = contract["project_year_raw"].ne("")
    metrics = {
        "reference_id_completeness": float(contract["reference_id"].ne("").mean()),
        "nonempty_country_mapping": float(
            contract.loc[nonempty_country, "country_code"].ne("").mean()
        )
        if nonempty_country.any()
        else 1.0,
        "nonempty_year_parse_rate": float(
            contract.loc[nonempty_year, "year_start"].notna().mean()
        )
        if nonempty_year.any()
        else 1.0,
        "access_control_prefilter": float(
            bool(
                result.filter_audit.get("authorization_applied_before_scoring")
                and result.filter_audit.get("all_scored_rows_authorized", True)
            )
        ),
        "exclusions_enforced": float(
            result.recommendations["evidence_available"].eq(True).all()
        )
        if not result.recommendations.empty
        else 0.0,
        "must_gate_enforced": float(
            (
                result.recommendations["must_covered"]
                == result.recommendations["must_total"]
            ).all()
        )
        if not result.recommendations.empty
        else 0.0,
        "spreadsheet_serialization_safe": float(
            all(
                str(safe_spreadsheet_value(value)).startswith("'")
                for value in ("=1+1", "+CMD", "-2+3", "@SUM(A1:A2)", " =HYPERLINK()")
            )
        ),
        "citation_completeness": float(
            result.citation_metrics["citation_completeness"]
        ),
        "citation_integrity": float(result.citation_metrics["citation_integrity"]),
        "duplicate_shortlist_control": float(
            not result.recommendations["duplicate_group_id"].duplicated().any()
        )
        if not result.recommendations.empty
        else 0.0,
    }
    checks = {
        name: {
            "value": value,
            "threshold": float(thresholds[name]),
            "passed": value >= float(thresholds[name]),
        }
        for name, value in metrics.items()
    }
    passed = all(check["passed"] for check in checks.values())
    return {
        "schema_version": 1,
        "phase": "5.2",
        "pipeline_version": config["pipeline_version"],
        "technical_gate": "PASS" if passed else "FAIL",
        "checks": checks,
        "citation_support_proxy": result.citation_metrics[
            "citation_support_proxy"
        ],
        "citation_correctness_status": result.citation_metrics[
            "citation_correctness_status"
        ],
        "production_promotion": "BLOCKED_PENDING_PHASE_5_1_EXPERT_EVALUATION",
    }


def _report(
    manifest: dict[str, Any],
    quality_gate: dict[str, Any],
    comparison: dict[str, Any],
) -> str:
    checks = "\n".join(
        f"- {name}: **{'PASS' if value['passed'] else 'FAIL'}** "
        f"({value['value']:.3f}, threshold {value['threshold']:.3f})"
        for name, value in quality_gate["checks"].items()
    )
    return f"""# Phase 5.2 — Matching engine and data-contract hardening

**Technical status:** {manifest['status']}
**Quality gate:** {quality_gate['technical_gate']}
**Production promotion:** {manifest['production_promotion_status']}

## Additive normalized contract

- Source reference rows: **{manifest['source_references']}**
- Base shortlist-eligible references: **{manifest['base_shortlist_eligible_references']}**
- Normalized countries: **{manifest['normalized_country_codes']}**
- Duplicate groups: **{manifest['duplicate_groups']}**
- Data-quality findings retained: **{manifest['data_quality_issues']}**

Raw Phase 4 values were not modified. Stable display IDs, ISO country codes,
date intervals, evidence types, controlled tags, and duplicate groups were
written to the Phase 5.2 output only.

## Controlled sample matching

- Retrieval mode: **{manifest['retrieval_mode']}**
- Content requirements searched: **{manifest['content_requirements']}**
- Eligibility policies separated: **{manifest['policy_requirements']}**
- Shortlist-eligible recommendations: **{manifest['recommendations']}**
- Diagnostic ineligible candidates: **{manifest['ineligible_candidates']}**
- Evidence rows: **{manifest['evidence_rows']}**
- Citation completeness: **{manifest['citation_completeness']:.3f}**
- Citation integrity: **{manifest['citation_integrity']:.3f}**
- Citation correctness: **PENDING HUMAN AUDIT**

## Quality gate

{checks}

## V1 versus V2 boundary

- V1 retrieval: `{comparison['v1']['retrieval_mode']}`
- V2 retrieval: `{comparison['v2']['retrieval_mode']}`
- V1 shortlisted rows missing a MUST: **{comparison['v1']['recommendations_missing_must']}**
- V2 shortlisted rows missing a MUST: **{comparison['v2']['recommendations_missing_must']}**

This sample and the bootstrap fusion sweep are technical diagnostics, not
expert relevance evidence. Phase 5.1 remains mandatory before selecting a
production retrieval configuration or claiming that V2 is more relevant.
"""


def _hash_outputs(run_root: Path) -> Path:
    excluded = {"SHA256SUMS.txt", "_SUCCESS.json"}
    paths = sorted(
        (
            path
            for path in run_root.rglob("*")
            if path.is_file() and path.name not in excluded
        ),
        key=lambda path: str(path.relative_to(run_root)),
    )
    text = "".join(
        f"{sha256_file(path)}  {path.relative_to(run_root)}\n" for path in paths
    )
    output = run_root / "SHA256SUMS.txt"
    _atomic_text(output, text)
    return output


def _load_sample_controls(project_root: Path, config: dict) -> dict[str, Any]:
    root = (
        project_root
        / "data"
        / "opportunities"
        / config["input"]["sample_opportunity_id"]
        / config["input"]["phase6_run_name"]
    )
    requirements_path = root / "requirements.jsonl"
    proposals_path = root / "filter_proposals.json"
    if not requirements_path.exists() or not proposals_path.exists():
        raise FileNotFoundError(
            "The signed Phase 6 redacted sample is required for technical validation"
        )
    requirements = [
        json.loads(line)
        for line in requirements_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    proposals = json.loads(proposals_path.read_text(encoding="utf-8"))
    return compile_phase6_controls(requirements, proposals, config)


def run_phase5_2(
    project_root: Path,
    config_path: Path,
    *,
    phase5_config: dict,
    embedding_adapter: Any,
    progress=print,
) -> dict[str, Any]:
    config = load_phase5_2_config(config_path)
    pinned = verify_pinned_inputs(project_root, config)
    run_root = (
        project_root
        / config["output"]["root"]
        / config["input"]["snapshot_id"]
        / config["output"]["run_name"]
    )
    success_path = run_root / config["output"]["success_marker"]
    if success_path.exists():
        verified = verify_phase5_2(run_root, config)
        progress("Existing Phase 5.2 run verified; nothing was rebuilt.")
        return verified | {"run_root": str(run_root), "resumed": True}
    run_root.mkdir(parents=True, exist_ok=True)
    reports_root = run_root / "reports"
    reports_root.mkdir(parents=True, exist_ok=True)

    progress("Building additive normalized reference data contract...")
    references = pd.read_parquet(pinned["references"])
    documents = pd.read_parquet(pinned["documents"])
    if len(references) != int(config["input"]["expected_references"]):
        raise AssertionError("Reference count changed")
    if len(documents) != int(config["input"]["expected_documents"]):
        raise AssertionError("Document count changed")
    contract, issues = build_reference_contract(references, documents, config)
    contract.to_parquet(run_root / "reference_contract.parquet", index=False)
    contract.to_json(
        run_root / "reference_contract.jsonl",
        orient="records",
        lines=True,
        force_ascii=False,
    )
    safe_spreadsheet_frame(issues).to_csv(
        run_root / "data_quality_issues.csv",
        index=False,
        encoding="utf-8-sig",
    )
    _atomic_json(run_root / "facet_values.json", facet_counts(contract))

    progress("Loading the Phase 6 redacted sample and separating eligibility policies...")
    controls = _load_sample_controls(project_root, config)
    engine = HardenedMatchingEngine.load(
        pinned["phase5_root"],
        contract,
        config,
        embedding_adapter=embedding_adapter,
    )
    result = engine.match(
        controls["requirements"],
        allowed_security_classifications=config["security"][
            "validation_allowed_classifications"
        ],
        hard_filters=controls["hard_filters"],
        soft_preferences=controls["soft_preferences"],
        exclusions=controls["exclusions"],
        mode=config["retrieval"]["default_mode"],
        policy_requirements=controls["policy_requirements"],
    )
    if result.recommendations.empty:
        raise MatchingHardeningError(
            "No shortlist-eligible sample references survived the Phase 5.2 controls"
        )
    result.recommendations.to_parquet(
        run_root / "sample_recommendations.parquet", index=False
    )
    result.recommendations.to_json(
        run_root / "sample_recommendations.json",
        orient="records",
        force_ascii=False,
        indent=2,
    )
    result.ineligible_candidates.to_parquet(
        run_root / "sample_ineligible_candidates.parquet", index=False
    )
    result.evidence.to_parquet(run_root / "sample_evidence.parquet", index=False)
    result.evidence.to_json(
        run_root / "sample_evidence.jsonl",
        orient="records",
        lines=True,
        force_ascii=False,
    )
    safe_spreadsheet_frame(result.coverage).to_csv(
        run_root / "sample_requirement_coverage.csv",
        index=False,
        encoding="utf-8-sig",
    )
    _atomic_json(run_root / "sample_filter_audit.json", result.filter_audit)
    _atomic_json(run_root / "sample_facet_counts.json", result.facets)
    _atomic_json(
        run_root / "sample_policy_requirements.json", result.policy_requirements
    )
    create_citation_audit_workbook(
        run_root / "CITATION_AUDIT_SAMPLE.xlsx",
        result.evidence,
        int(config["evaluation"]["citation_audit_sample_size"]),
    )

    progress("Running the bootstrap fusion-weight sweep as a non-expert diagnostic...")
    probes_path = pinned["phase5_root"] / "evaluation" / "bootstrap_queries.parquet"
    if probes_path.exists():
        probes = pd.read_parquet(probes_path).reset_index(drop=True)
        vectors = np.asarray(
            embedding_adapter.encode_queries(probes["query_text"].tolist()),
            dtype=np.float32,
        )
        sweep = evaluate_fusion_sweep(
            phase5_chunks=engine.chunks,
            bm25=engine.bm25,
            embeddings=engine.embeddings,
            phase5_config=phase5_config,
            phase5_2_config=config,
            probes=probes,
            query_vectors=vectors,
        )
    else:
        sweep = pd.DataFrame(
            [
                {
                    "candidate": "NOT_RUN",
                    "label_source": config["evaluation"]["label_source"],
                    "expert_labeled": False,
                    "production_selection_allowed": False,
                    "reason": "Phase 5 bootstrap queries not found",
                }
            ]
        )
    safe_spreadsheet_frame(sweep).to_csv(
        run_root / "fusion_sweep.csv", index=False, encoding="utf-8-sig"
    )

    baseline_root = (
        project_root
        / "data"
        / "recommendations"
        / config["input"]["sample_opportunity_id"]
        / config["input"]["phase7_baseline_run_name"]
    )
    comparison = compare_v1_v2(
        baseline_root / "recommendations.json",
        baseline_root / "PHASE_7_MANIFEST.json",
        result,
        config["retrieval"]["default_mode"],
    )
    _atomic_json(run_root / "v1_v2_comparison.json", comparison)
    quality_gate = _quality_gate(contract, result, config)
    _atomic_json(
        run_root / config["output"]["quality_gate_name"], quality_gate
    )
    if quality_gate["technical_gate"] != "PASS":
        failed = [
            name
            for name, check in quality_gate["checks"].items()
            if not check["passed"]
        ]
        raise AssertionError(f"Phase 5.2 quality gate failed: {failed}")

    manifest = {
        "schema_version": 1,
        "phase": "5.2",
        "pipeline_version": config["pipeline_version"],
        "completed_at_utc": utc_now(),
        "status": "TECHNICAL_PASS_SAMPLE_ONLY",
        "qa_gate": quality_gate["technical_gate"],
        "snapshot_id": config["input"]["snapshot_id"],
        "source_references": int(len(contract)),
        "base_shortlist_eligible_references": int(
            contract["base_shortlist_eligible"].sum()
        ),
        "normalized_country_codes": int(
            contract["country_code"].replace("", np.nan).nunique()
        ),
        "duplicate_groups": int(contract["duplicate_group_id"].nunique()),
        "data_quality_issues": int(len(issues)),
        "content_requirements": int(len(controls["requirements"])),
        "policy_requirements": int(len(controls["policy_requirements"])),
        "approved_hard_filters": int(len(controls["hard_filters"])),
        "enforced_exclusions": int(len(controls["exclusions"])),
        "retrieval_mode": config["retrieval"]["default_mode"],
        "recommendations": int(len(result.recommendations)),
        "ineligible_candidates": int(len(result.ineligible_candidates)),
        "evidence_rows": int(len(result.evidence)),
        "citation_completeness": result.citation_metrics[
            "citation_completeness"
        ],
        "citation_integrity": result.citation_metrics["citation_integrity"],
        "citation_support_proxy": result.citation_metrics[
            "citation_support_proxy"
        ],
        "citation_correctness_status": result.citation_metrics[
            "citation_correctness_status"
        ],
        "authorization_applied_before_scoring": True,
        "hard_filters_applied_before_scoring": True,
        "exclusions_enforced": True,
        "must_gate_enforced": True,
        "spreadsheet_formula_serialization_enabled": True,
        "external_llm_calls": 0,
        "external_embedding_api_calls": 0,
        "local_query_embedding_calls": int(
            result.filter_audit["dense_query_embedding_calls"]
            + (len(probes) if probes_path.exists() else 0)
        ),
        "cross_encoder_calls": 0,
        "source_mutation_calls": 0,
        "production_quality_claim_allowed": False,
        "production_promotion_status": (
            "BLOCKED_PENDING_PHASE_5_1_EXPERT_EVALUATION"
        ),
        "v1_baseline_preserved": True,
    }
    manifest_path = run_root / config["output"]["manifest_name"]
    _atomic_json(manifest_path, manifest)
    _atomic_text(
        reports_root / config["output"]["report_name"],
        _report(manifest, quality_gate, comparison),
    )
    sums_path = _hash_outputs(run_root)
    _atomic_json(
        success_path,
        {
            "status": "COMPLETE_REPRODUCIBLE_PHASE_5_2",
            "created_at_utc": utc_now(),
            "pipeline_version": config["pipeline_version"],
            "manifest_sha256": sha256_file(manifest_path),
            "sha256sums_sha256": sha256_file(sums_path),
        },
    )
    verified = verify_phase5_2(run_root, config)
    return verified | {"run_root": str(run_root), "resumed": False}


def verify_phase5_2(run_root: Path, config: dict) -> dict[str, Any]:
    success_path = run_root / config["output"]["success_marker"]
    manifest_path = run_root / config["output"]["manifest_name"]
    quality_path = run_root / config["output"]["quality_gate_name"]
    sums_path = run_root / "SHA256SUMS.txt"
    for path in (success_path, manifest_path, quality_path, sums_path):
        if not path.exists():
            raise FileNotFoundError(path)
    success = json.loads(success_path.read_text(encoding="utf-8"))
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    quality = json.loads(quality_path.read_text(encoding="utf-8"))
    if success["manifest_sha256"] != sha256_file(manifest_path):
        raise AssertionError("Phase 5.2 manifest hash mismatch")
    if success["sha256sums_sha256"] != sha256_file(sums_path):
        raise AssertionError("Phase 5.2 checksum-file hash mismatch")
    for line in sums_path.read_text(encoding="utf-8").splitlines():
        expected, name = line.split("  ", 1)
        if sha256_file(run_root / name) != expected:
            raise AssertionError(f"Phase 5.2 output changed: {name}")
    if quality["technical_gate"] != "PASS" or manifest["qa_gate"] != "PASS":
        raise AssertionError("Phase 5.2 technical gate is incomplete")
    if any(
        int(manifest.get(field, -1)) != 0
        for field in (
            "external_llm_calls",
            "external_embedding_api_calls",
            "cross_encoder_calls",
            "source_mutation_calls",
        )
    ):
        raise AssertionError("Phase 5.2 provider/source boundary failed")
    if not manifest["authorization_applied_before_scoring"]:
        raise AssertionError("Authorization order failed")
    if not manifest["must_gate_enforced"] or not manifest["exclusions_enforced"]:
        raise AssertionError("Mandatory matching controls failed")
    if manifest["production_quality_claim_allowed"]:
        raise AssertionError("Sample diagnostics cannot become a production claim")
    if (
        manifest["production_promotion_status"]
        != "BLOCKED_PENDING_PHASE_5_1_EXPERT_EVALUATION"
    ):
        raise AssertionError("Expert evaluation promotion gate was bypassed")
    return {
        "success": success,
        "manifest": manifest,
        "quality_gate": quality,
    }
