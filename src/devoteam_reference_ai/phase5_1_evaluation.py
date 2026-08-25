from __future__ import annotations

import hashlib
import json
import math
import os
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
import yaml

from .phase5_bm25 import BM25Index, normalize_search_text
from .phase5_retrieval import (
    E5EmbeddingAdapter,
    HybridSearchEngine,
    build_retrieval_texts,
    load_phase5_config,
    verify_phase5,
)


QUERY_COLUMNS = [
    "query_id",
    "query_text",
    "language",
    "business_context",
    "mandatory_filters_json",
    "query_type",
    "origin",
    "not_derived_from_reference_corpus",
    "approved_for_evaluation",
    "notes",
]

LABEL_COLUMNS = [
    "query_id",
    "candidate_id",
    "query_text",
    "language",
    "business_context",
    "mandatory_filters_json",
    "source_file_name",
    "document_type",
    "document_language",
    "candidate_metadata",
    "evidence_excerpt",
    "citation_label",
    "citation_uri",
    "relevance_0_1_2",
    "failure_category",
    "evidence_notes",
]

GOVERNANCE_ROLES = [
    "evaluation_owner",
    "labeler_1_name",
    "labeler_2_name",
    "adjudicator_name",
    "supervisor_name",
]


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _atomic_json(path: Path, value: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(value, indent=2, ensure_ascii=False, sort_keys=True), encoding="utf-8"
    )
    os.replace(temporary, path)


def _write_state(run_root: Path, status: str, **details: Any) -> dict:
    state = {
        "schema_version": 1,
        "pipeline_version": "phase5_1_expert_evaluation_v1",
        "status": status,
        "updated_at_utc": _now(),
        **details,
    }
    _atomic_json(run_root / "PHASE_5_1_STATE.json", state)
    return state | {"run_root": str(run_root)}


def load_phase5_1_config(path: Path) -> dict:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    if config.get("pipeline_version") != "phase5_1_expert_evaluation_v1":
        raise AssertionError("Unexpected Phase 5.1 pipeline version")
    if config["evaluation"]["gold_set_purpose"] != "FROZEN_TEST_ONLY_NO_TUNING":
        raise AssertionError("Gold-set anti-leakage policy is missing")
    if not config["evaluation"]["no_automatic_production_promotion"]:
        raise AssertionError("Phase 5.1 may not auto-promote a production model")
    return config


def _phase5_root(project_root: Path, config: dict) -> Path:
    return (
        project_root
        / "data"
        / "indexes"
        / config["input"]["snapshot_id"]
        / config["input"]["phase5_run_name"]
    )


def _run_root(project_root: Path, config: dict) -> Path:
    return (
        project_root
        / config["output"]["root"]
        / config["input"]["snapshot_id"]
        / config["output"]["run_name"]
    )


def _verify_signed_phase5(project_root: Path, config: dict) -> tuple[Path, dict]:
    phase5_root = _phase5_root(project_root, config)
    manifest = verify_phase5(phase5_root)
    manifest_path = phase5_root / "PHASE_5_MANIFEST.json"
    sums_path = phase5_root / "SHA256SUMS.txt"
    if sha256_file(manifest_path) != config["input"]["expected_phase5_manifest_sha256"]:
        raise AssertionError("Pinned Phase 5 manifest hash changed")
    if sha256_file(sums_path) != config["input"]["expected_phase5_sha256sums_sha256"]:
        raise AssertionError("Pinned Phase 5 checksum-list hash changed")
    if manifest["chunks_indexed"] != config["input"]["expected_chunks"]:
        raise AssertionError("Pinned Phase 5 chunk count changed")
    if manifest["embedding_dimensions"] != config["input"]["expected_embedding_dimensions"]:
        raise AssertionError("Pinned Phase 5 embedding dimension changed")
    return phase5_root, manifest


def _style_header(worksheet, row: int = 1) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for cell in worksheet[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6E2A8D")
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _set_widths(worksheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def create_query_intake_workbook(path: Path, config: dict) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instruction_rows = [
        ("Purpose", "Frozen expert query set for the preregistered BM25/dense/hybrid comparison."),
        ("No leakage", "Queries must not be copied or constructed from the reference corpus or Phase 5 bootstrap probes."),
        ("Query count", "Complete all 50 rows. Do not add, delete, or reorder query IDs."),
        ("Languages", "Minimum coverage: 30 French, 5 English, and 5 Arabic queries."),
        ("Difficult cases", "Include at least 5 ACRONYM_HEAVY, 5 SPARSE, and 5 AMBIGUOUS queries."),
        ("Filters", "mandatory_filters_json must be a JSON object using supported visible business filters; use {} when none apply."),
        ("Approval", "Use only approved real opportunities or approved realistic scenarios; set approved_for_evaluation to YES."),
        ("Security", "Do not paste confidential opportunity text unless Devoteam explicitly approved its use in this INTERNAL evaluation."),
        ("Next step", "Save this workbook and rerun the Phase 5.1 notebook. It will validate the intake before model download or retrieval."),
    ]
    for row in instruction_rows:
        instructions.append(row)
    _style_header(instructions)
    instructions["A1"] = "Control"
    instructions["B1"] = "Rule"
    _set_widths(instructions, {"A": 24, "B": 110})
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    governance = workbook.create_sheet("Governance")
    governance.append(["role", "name", "confirmed", "notes"])
    for role in GOVERNANCE_ROLES:
        governance.append([role, None, "NO", None])
    _style_header(governance)
    yes_no = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    governance.add_data_validation(yes_no)
    yes_no.add(f"C2:C{governance.max_row}")
    _set_widths(governance, {"A": 28, "B": 32, "C": 14, "D": 60})
    governance.freeze_panes = "A2"

    queries = workbook.create_sheet("Queries")
    queries.append(QUERY_COLUMNS)
    for index in range(1, int(config["queries"]["target_count"]) + 1):
        queries.append(
            [
                f"EXP-{index:03d}",
                None,
                "fr",
                None,
                "{}",
                "STANDARD",
                "REAL_OPPORTUNITY",
                "NO",
                "NO",
                None,
            ]
        )
    _style_header(queries)
    queries.freeze_panes = "A2"
    queries.auto_filter.ref = f"A1:J{queries.max_row}"
    _set_widths(
        queries,
        {
            "A": 14,
            "B": 62,
            "C": 12,
            "D": 55,
            "E": 38,
            "F": 22,
            "G": 24,
            "H": 30,
            "I": 24,
            "J": 45,
        },
    )
    language_validation = DataValidation(type="list", formula1='"fr,en,ar"', allow_blank=False)
    query_type_validation = DataValidation(
        type="list", formula1='"STANDARD,ACRONYM_HEAVY,SPARSE,AMBIGUOUS"', allow_blank=False
    )
    origin_validation = DataValidation(
        type="list", formula1='"REAL_OPPORTUNITY,APPROVED_REALISTIC"', allow_blank=False
    )
    yes_no_queries = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    for validation in (
        language_validation,
        query_type_validation,
        origin_validation,
        yes_no_queries,
    ):
        queries.add_data_validation(validation)
    language_validation.add(f"C2:C{queries.max_row}")
    query_type_validation.add(f"F2:F{queries.max_row}")
    origin_validation.add(f"G2:G{queries.max_row}")
    yes_no_queries.add(f"H2:I{queries.max_row}")
    for row in queries.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    queries.protection.sheet = True
    queries.protection.password = "devoteam"
    for row in range(2, queries.max_row + 1):
        for column in range(2, 11):
            queries.cell(row, column).protection = openpyxl.styles.Protection(locked=False)
    queries.sheet_properties.pageSetUpPr.fitToPage = True

    manifest = workbook.create_sheet("TemplateManifest")
    manifest.append(["key", "value"])
    manifest.append(["pipeline_version", config["pipeline_version"]])
    manifest.append(["snapshot_id", config["input"]["snapshot_id"]])
    manifest.append(["gold_set_purpose", config["evaluation"]["gold_set_purpose"]])
    manifest.append(["created_at_utc", _now()])
    _style_header(manifest)
    manifest.sheet_state = "hidden"
    workbook.save(path)


def _read_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name, dtype=object, keep_default_na=False)


def validate_query_intake(path: Path, config: dict) -> tuple[pd.DataFrame, dict[str, str], list[str]]:
    issues: list[str] = []
    try:
        queries = _read_sheet(path, "Queries")
        governance = _read_sheet(path, "Governance")
    except Exception as exc:
        return pd.DataFrame(), {}, [f"Workbook structure cannot be read: {exc}"]
    if list(queries.columns) != QUERY_COLUMNS:
        return queries, {}, ["Queries sheet columns were changed; restore the controlled template."]
    target = int(config["queries"]["target_count"])
    if len(queries) != target:
        issues.append(f"Exactly {target} query rows are required; found {len(queries)}.")
    expected_ids = [f"EXP-{index:03d}" for index in range(1, target + 1)]
    if queries["query_id"].astype(str).tolist() != expected_ids:
        issues.append("Query IDs must remain EXP-001 through EXP-050 in their original order.")
    query_texts = queries["query_text"].astype(str).map(lambda value: " ".join(value.split()))
    minimum = int(config["queries"]["min_normalized_characters"])
    maximum = int(config["queries"]["max_characters"])
    for index, value in enumerate(query_texts, start=1):
        if len(value) < minimum:
            issues.append(f"EXP-{index:03d}: query_text needs at least {minimum} normalized characters.")
        if len(value) > maximum:
            issues.append(f"EXP-{index:03d}: query_text exceeds {maximum} characters.")
    normalized = query_texts.map(normalize_search_text)
    duplicate_ids = queries.loc[normalized.duplicated(keep=False), "query_id"].astype(str).tolist()
    if duplicate_ids:
        issues.append(f"Duplicate normalized queries: {', '.join(duplicate_ids)}.")
    if queries["business_context"].astype(str).str.strip().str.len().lt(10).any():
        bad = queries.loc[
            queries["business_context"].astype(str).str.strip().str.len().lt(10), "query_id"
        ].astype(str).tolist()
        issues.append(f"Business context needs at least 10 characters: {', '.join(bad)}.")
    allowed_languages = set(config["queries"]["allowed_languages"])
    languages = queries["language"].astype(str).str.lower().str.strip()
    if not set(languages).issubset(allowed_languages):
        issues.append("Languages must be one of fr, en, or ar.")
    language_counts = languages.value_counts().to_dict()
    for language, required in config["queries"]["language_minimums"].items():
        if int(language_counts.get(language, 0)) < int(required):
            issues.append(f"Language {language} requires at least {required} queries.")
    query_types = queries["query_type"].astype(str).str.upper().str.strip()
    allowed_types = {"STANDARD", *config["queries"]["required_query_types"].keys()}
    if not set(query_types).issubset(allowed_types):
        issues.append("query_type contains an unsupported value.")
    query_type_counts = query_types.value_counts().to_dict()
    for query_type, required in config["queries"]["required_query_types"].items():
        if int(query_type_counts.get(query_type, 0)) < int(required):
            issues.append(f"query_type {query_type} requires at least {required} queries.")
    origins = set(queries["origin"].astype(str).str.upper().str.strip())
    if not origins.issubset(set(config["queries"]["allowed_origins"])):
        issues.append("origin must be REAL_OPPORTUNITY or APPROVED_REALISTIC.")
    if not queries["not_derived_from_reference_corpus"].astype(str).str.upper().eq("YES").all():
        issues.append("Every query must confirm it was not derived from the reference corpus.")
    if not queries["approved_for_evaluation"].astype(str).str.upper().eq("YES").all():
        issues.append("Every query must be explicitly approved for evaluation.")
    supported_filters = {
        "country",
        "business_unit",
        "client",
        "sector",
        "service_nature",
        "offering",
        "project_year",
        "attestation_available",
        "document_type",
        "data_quality_status",
        "year_before",
        "year_after",
    }
    for _, row in queries.iterrows():
        try:
            filters = json.loads(str(row["mandatory_filters_json"]).strip() or "{}")
            if not isinstance(filters, dict):
                raise ValueError("must be a JSON object")
            unknown = sorted(set(filters) - supported_filters)
            if unknown:
                raise ValueError(f"unsupported keys: {unknown}")
        except Exception as exc:
            issues.append(f"{row['query_id']}: mandatory_filters_json is invalid: {exc}.")
    governance_map = {
        str(row["role"]).strip(): str(row["name"]).strip()
        for _, row in governance.iterrows()
        if str(row.get("role", "")).strip()
    }
    confirmations = {
        str(row["role"]).strip(): str(row["confirmed"]).strip().upper()
        for _, row in governance.iterrows()
        if str(row.get("role", "")).strip()
    }
    for role in GOVERNANCE_ROLES:
        if not governance_map.get(role):
            issues.append(f"Governance name is missing: {role}.")
        if confirmations.get(role) != "YES":
            issues.append(f"Governance confirmation must be YES: {role}.")
    if governance_map.get("labeler_1_name") == governance_map.get("labeler_2_name"):
        issues.append("Labeler 1 and labeler 2 must be two different people.")
    if governance_map.get("adjudicator_name") in {
        governance_map.get("labeler_1_name"),
        governance_map.get("labeler_2_name"),
    }:
        issues.append("The adjudicator must be independent from both labelers.")
    clean = queries.copy()
    clean["query_text"] = query_texts
    clean["language"] = languages
    clean["query_type"] = query_types
    clean["origin"] = clean["origin"].astype(str).str.upper().str.strip()
    clean["mandatory_filters_json"] = clean["mandatory_filters_json"].map(
        lambda value: json.dumps(json.loads(str(value) or "{}"), ensure_ascii=False, sort_keys=True)
    )
    return clean, governance_map, issues


def _validation_report(title: str, issues: Iterable[str]) -> str:
    values = list(issues)
    body = "\n".join(f"- {value}" for value in values) if values else "- No issues."
    return f"# {title}\n\n{body}\n"


def _extract_json_values(row: pd.Series, field: str) -> list[str]:
    try:
        value = json.loads(str(row.get(field, "[]")) or "[]")
    except json.JSONDecodeError:
        return []
    if isinstance(value, list):
        return [str(item) for item in value if str(item).strip()]
    return []


def _candidate_metadata(row: pd.Series) -> str:
    fields = [
        ("Client", "client_values_json"),
        ("Country", "country_values_json"),
        ("Sector", "sector_values_json"),
        ("Offering", "offering_values_json"),
        ("Service", "service_nature_values_json"),
        ("Year", "project_year_values_json"),
    ]
    parts = []
    for label, field in fields:
        values = _extract_json_values(row, field)
        if values:
            parts.append(f"{label}: {', '.join(values[:4])}")
    return " | ".join(parts)[:1000]


def build_candidate_pool(
    *,
    queries: pd.DataFrame,
    engine: HybridSearchEngine,
    query_vectors: np.ndarray,
    config: dict,
    progress=print,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    if query_vectors.shape[0] != len(queries):
        raise AssertionError("Expert query-vector alignment failed")
    systems = list(config["retrieval"]["systems"])
    depth = int(config["retrieval"]["judging_depth_per_system"])
    allowed = list(config["retrieval"]["allowed_security_classifications"])
    rankings: list[dict] = []
    filter_checks: list[bool] = []
    citation_checks: list[bool] = []
    for query_index, query in queries.iterrows():
        hard_filters = json.loads(query["mandatory_filters_json"])
        for mode in systems:
            results = engine.search_documents(
                query["query_text"],
                allowed_security_classifications=allowed,
                hard_filters=hard_filters,
                mode=mode,
                top_k=depth,
                query_vector=query_vectors[query_index],
            )
            for record in results.to_dict(orient="records"):
                rankings.append(
                    {
                        "query_id": query["query_id"],
                        "retrieval_mode": mode,
                        "document_id": str(record["document_id"]),
                        "rank": int(record["rank"]),
                        "document_score": float(record["document_score"]),
                        "best_chunk_id": str(record["best_chunk_id"]),
                        "citation_labels_json": record["citation_labels_json"],
                        "citation_uris_json": record["citation_uris_json"],
                    }
                )
                citation_checks.append(bool(json.loads(record["citation_uris_json"])))
        mask = engine.filter_engine.mask(
            allowed_security_classifications=allowed, hard_filters=hard_filters
        )
        filtered = engine.search_chunks(
            query["query_text"],
            allowed_security_classifications=allowed,
            hard_filters=hard_filters,
            mode="hybrid",
            top_k=depth,
            query_vector=query_vectors[query_index],
        )
        filter_checks.append(
            filtered.empty
            or all(mask[int(row)] for row in filtered["vector_row"].astype(int).tolist())
        )
        progress(f"Candidate retrieval: {query_index + 1}/{len(queries)} queries.")
    ranking_frame = pd.DataFrame(rankings)
    if ranking_frame.empty:
        raise AssertionError("No candidates were retrieved for the expert query set")
    if ranking_frame.duplicated(["query_id", "retrieval_mode", "document_id"]).any():
        raise AssertionError("Duplicate document inside one retrieval ranking")
    chunk_lookup = engine.chunks.set_index("chunk_id", drop=False)
    query_lookup = queries.set_index("query_id", drop=False)
    candidates: list[dict] = []
    for (query_id, document_id), group in ranking_frame.groupby(
        ["query_id", "document_id"], sort=True
    ):
        best = group.sort_values(["rank", "retrieval_mode"]).iloc[0]
        chunk = chunk_lookup.loc[best["best_chunk_id"]]
        query = query_lookup.loc[query_id]
        labels = json.loads(best["citation_labels_json"])
        uris = json.loads(best["citation_uris_json"])
        candidates.append(
            {
                "query_id": query_id,
                "candidate_id": "CAND-" + _sha256_text(f"{query_id}|{document_id}")[:16],
                "document_id": document_id,
                "query_text": query["query_text"],
                "language": query["language"],
                "business_context": query["business_context"],
                "mandatory_filters_json": query["mandatory_filters_json"],
                "source_file_name": chunk["source_file_name"],
                "document_type": chunk["document_type"],
                "document_language": chunk["document_language"],
                "candidate_metadata": _candidate_metadata(chunk),
                "evidence_excerpt": " ".join(str(chunk["chunk_text"]).split())[:1200],
                "citation_label": " | ".join(str(value) for value in labels)[:1000],
                "citation_uri": " | ".join(str(value) for value in uris)[:2000],
                "retrieval_modes_hidden_json": json.dumps(sorted(group["retrieval_mode"].tolist())),
                "ranks_hidden_json": json.dumps(
                    {
                        row["retrieval_mode"]: int(row["rank"])
                        for _, row in group.sort_values("retrieval_mode").iterrows()
                    },
                    sort_keys=True,
                ),
            }
        )
    candidate_frame = pd.DataFrame(candidates)
    if not candidate_frame["candidate_id"].is_unique:
        raise AssertionError("Candidate identifiers are not unique")
    diagnostics = {
        "query_count": int(len(queries)),
        "ranking_rows": int(len(ranking_frame)),
        "candidate_judgments_per_labeler": int(len(candidate_frame)),
        "maximum_candidates_per_query": int(candidate_frame.groupby("query_id").size().max()),
        "filter_correctness": float(np.mean(filter_checks)),
        "citation_coverage": float(np.mean(citation_checks)) if citation_checks else 0.0,
    }
    if diagnostics["filter_correctness"] != 1.0:
        raise AssertionError("Mandatory filter verification failed")
    if diagnostics["citation_coverage"] != 1.0:
        raise AssertionError("Candidate citation coverage failed")
    return ranking_frame, candidate_frame, diagnostics


def _blinded_order(frame: pd.DataFrame, seed: str) -> pd.DataFrame:
    ordered = frame.copy()
    ordered["_blind_order"] = ordered.apply(
        lambda row: _sha256_text(f"{seed}|{row['query_id']}|{row['candidate_id']}"), axis=1
    )
    return ordered.sort_values(["query_id", "_blind_order"]).drop(columns="_blind_order")


def create_labeler_packet(path: Path, candidates: pd.DataFrame, config: dict, labeler_number: int) -> None:
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = openpyxl.Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    rows = [
        ("Role", f"Independent Devoteam relevance labeler {labeler_number}."),
        ("Blinding", "Retrieval system, score, and rank are intentionally hidden."),
        ("0", "Irrelevant: the reference does not address the opportunity."),
        ("1", "Partially relevant: useful overlap, but not a strong primary reference."),
        ("2", "Highly relevant: strong evidence for material opportunity requirements."),
        ("Independence", "Do not discuss labels with the other labeler before both packets are complete."),
        ("Required", "Complete relevance for every row. Use a failure category for relevance 0."),
        ("Evidence", "Open citations when necessary. Do not change candidate or query columns."),
    ]
    for row in rows:
        instructions.append(row)
    instructions["A1"] = "Control"
    instructions["B1"] = "Instruction"
    _style_header(instructions)
    _set_widths(instructions, {"A": 22, "B": 105})
    labels = workbook.create_sheet("Labels")
    labels.append(LABEL_COLUMNS)
    blinded = _blinded_order(
        candidates,
        f"{config['retrieval']['deterministic_blinding_seed']}|labeler-{labeler_number}",
    )
    for _, row in blinded.iterrows():
        labels.append([row.get(column, "") if column not in {"relevance_0_1_2", "failure_category", "evidence_notes"} else "" for column in LABEL_COLUMNS])
    _style_header(labels)
    labels.freeze_panes = "A2"
    labels.auto_filter.ref = f"A1:P{labels.max_row}"
    _set_widths(
        labels,
        {
            "A": 14,
            "B": 24,
            "C": 55,
            "D": 10,
            "E": 45,
            "F": 34,
            "G": 35,
            "H": 22,
            "I": 16,
            "J": 55,
            "K": 75,
            "L": 48,
            "M": 48,
            "N": 18,
            "O": 28,
            "P": 45,
        },
    )
    relevance = DataValidation(type="list", formula1='"0,1,2"', allow_blank=False)
    categories = ",".join(config["labeling"]["failure_categories"])
    category = DataValidation(type="list", formula1=f'"{categories}"', allow_blank=True)
    labels.add_data_validation(relevance)
    labels.add_data_validation(category)
    relevance.add(f"N2:N{labels.max_row}")
    category.add(f"O2:O{labels.max_row}")
    for row in labels.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column <= 13:
                cell.protection = openpyxl.styles.Protection(locked=True)
            else:
                cell.protection = openpyxl.styles.Protection(locked=False)
    labels.protection.sheet = True
    labels.protection.password = "devoteam"
    audit = workbook.create_sheet("PacketManifest")
    audit.append(["key", "value"])
    audit.append(["pipeline_version", config["pipeline_version"]])
    audit.append(["snapshot_id", config["input"]["snapshot_id"]])
    audit.append(["labeler_number", labeler_number])
    audit.append(["candidate_set_sha256", _frame_hash(candidates, ["query_id", "candidate_id", "document_id"])])
    audit.append(["created_at_utc", _now()])
    _style_header(audit)
    audit.sheet_state = "hidden"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _frame_hash(frame: pd.DataFrame, columns: list[str]) -> str:
    values = frame[columns].fillna("").astype(str).sort_values(columns).to_dict(orient="records")
    return _sha256_text(json.dumps(values, ensure_ascii=False, sort_keys=True, separators=(",", ":")))


def read_completed_labeler_packet(
    path: Path, candidates: pd.DataFrame, config: dict
) -> tuple[pd.DataFrame, list[str]]:
    if not path.exists():
        return pd.DataFrame(), [f"Missing file: {path.name}."]
    try:
        labels = _read_sheet(path, "Labels")
    except Exception as exc:
        return pd.DataFrame(), [f"{path.name} cannot be read: {exc}."]
    issues: list[str] = []
    if list(labels.columns) != LABEL_COLUMNS:
        return labels, [f"{path.name}: controlled Labels columns changed."]
    expected = set(candidates["candidate_id"].astype(str))
    observed = set(labels["candidate_id"].astype(str))
    if expected != observed or len(labels) != len(candidates):
        issues.append(f"{path.name}: candidate rows were added, deleted, or changed.")
    if labels["candidate_id"].astype(str).duplicated().any():
        issues.append(f"{path.name}: duplicate candidate IDs.")
    relevance = pd.to_numeric(labels["relevance_0_1_2"], errors="coerce")
    if relevance.isna().any() or not relevance.dropna().isin(config["labeling"]["relevance_values"]).all():
        missing = labels.loc[relevance.isna(), "candidate_id"].astype(str).tolist()[:10]
        issues.append(f"{path.name}: every row needs relevance 0, 1, or 2; examples: {missing}.")
    zero_mask = relevance.eq(0)
    categories = labels["failure_category"].astype(str).str.strip().str.upper()
    allowed_categories = set(config["labeling"]["failure_categories"])
    if not set(categories[categories.ne("")]).issubset(allowed_categories):
        issues.append(f"{path.name}: unsupported failure category.")
    if (zero_mask & categories.isin({"", "NONE"})).any():
        bad = labels.loc[zero_mask & categories.isin({"", "NONE"}), "candidate_id"].astype(str).tolist()[:10]
        issues.append(f"{path.name}: relevance 0 requires a failure category; examples: {bad}.")
    clean = labels[["query_id", "candidate_id", "relevance_0_1_2", "failure_category", "evidence_notes"]].copy()
    clean["relevance"] = relevance
    return clean, issues


def weighted_kappa(first: Iterable[int], second: Iterable[int], categories: int = 3) -> float:
    first_values = np.asarray(list(first), dtype=int)
    second_values = np.asarray(list(second), dtype=int)
    if first_values.shape != second_values.shape or first_values.size == 0:
        raise ValueError("Kappa arrays must be non-empty and aligned")
    observed = np.zeros((categories, categories), dtype=float)
    for left, right in zip(first_values, second_values):
        observed[left, right] += 1.0
    observed /= observed.sum()
    first_marginal = observed.sum(axis=1)
    second_marginal = observed.sum(axis=0)
    expected = np.outer(first_marginal, second_marginal)
    weights = np.fromfunction(
        lambda i, j: ((i - j) / max(categories - 1, 1)) ** 2,
        (categories, categories),
        dtype=float,
    )
    numerator = float((weights * observed).sum())
    denominator = float((weights * expected).sum())
    return 1.0 - numerator / denominator if denominator else 1.0


def label_agreement(first: pd.DataFrame, second: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    merged = first.merge(second, on=["query_id", "candidate_id"], suffixes=("_1", "_2"), validate="one_to_one")
    merged["agreement"] = merged["relevance_1"].astype(int).eq(merged["relevance_2"].astype(int))
    metrics = {
        "judgments": int(len(merged)),
        "raw_agreement": round(float(merged["agreement"].mean()), 6),
        "quadratic_weighted_kappa": round(
            weighted_kappa(merged["relevance_1"].astype(int), merged["relevance_2"].astype(int)), 6
        ),
        "disagreements": int((~merged["agreement"]).sum()),
    }
    return merged, metrics


def _citation_sample(candidates: pd.DataFrame, target: int, seed: str) -> pd.DataFrame:
    sample = candidates.copy()
    sample["_order"] = sample.apply(
        lambda row: _sha256_text(f"{seed}|{row['query_id']}|{row['candidate_id']}"), axis=1
    )
    sample = sample.sort_values(["language", "_order"])
    selected: list[pd.DataFrame] = []
    languages = sorted(sample["language"].unique())
    base = target // max(len(languages), 1)
    for language in languages:
        selected.append(sample.loc[sample["language"].eq(language)].head(base))
    chosen = pd.concat(selected, ignore_index=True) if selected else sample.head(0)
    remaining = sample.loc[~sample["candidate_id"].isin(chosen["candidate_id"])].head(target - len(chosen))
    chosen = pd.concat([chosen, remaining], ignore_index=True).head(target)
    return chosen.drop(columns="_order")


def create_adjudication_workbook(
    path: Path,
    merged_labels: pd.DataFrame,
    candidates: pd.DataFrame,
    agreement: dict,
    config: dict,
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    joined = candidates.merge(
        merged_labels[["query_id", "candidate_id", "relevance_1", "relevance_2", "agreement"]],
        on=["query_id", "candidate_id"],
        validate="one_to_one",
    )
    workbook = openpyxl.Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    for row in [
        ("Purpose", "Independent adjudication of relevance disagreements and citation audit."),
        ("Agreement", f"Raw agreement={agreement['raw_agreement']:.3f}; weighted kappa={agreement['quadratic_weighted_kappa']:.3f}."),
        ("Adjudication", "Agreements are prefilled. For disagreements, choose 0, 1, or 2 and add a concise rationale."),
        ("Citation audit", "Open each sampled citation and mark 1 only when the link and cited evidence are correct."),
        ("Independence", "The adjudicator must be different from both original labelers."),
    ]:
        instructions.append(row)
    instructions["A1"] = "Control"
    instructions["B1"] = "Instruction"
    _style_header(instructions)
    _set_widths(instructions, {"A": 24, "B": 105})

    sheet = workbook.create_sheet("Adjudication")
    columns = [
        "query_id",
        "candidate_id",
        "query_text",
        "source_file_name",
        "candidate_metadata",
        "evidence_excerpt",
        "citation_uri",
        "labeler_1_relevance",
        "labeler_2_relevance",
        "adjudicated_relevance",
        "adjudication_notes",
    ]
    sheet.append(columns)
    for _, row in joined.sort_values(["query_id", "candidate_id"]).iterrows():
        agreed = int(row["relevance_1"]) if bool(row["agreement"]) else ""
        sheet.append(
            [
                row["query_id"],
                row["candidate_id"],
                row["query_text"],
                row["source_file_name"],
                row["candidate_metadata"],
                row["evidence_excerpt"],
                row["citation_uri"],
                int(row["relevance_1"]),
                int(row["relevance_2"]),
                agreed,
                "" if not bool(row["agreement"]) else "AUTO_AGREEMENT",
            ]
        )
    _style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{sheet.max_row}"
    _set_widths(sheet, {"A": 14, "B": 24, "C": 55, "D": 35, "E": 55, "F": 75, "G": 48, "H": 18, "I": 18, "J": 22, "K": 55})
    relevance = DataValidation(type="list", formula1='"0,1,2"', allow_blank=False)
    sheet.add_data_validation(relevance)
    relevance.add(f"J2:J{sheet.max_row}")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.protection = openpyxl.styles.Protection(locked=cell.column not in {10, 11})
    sheet.protection.sheet = True
    sheet.protection.password = "devoteam"

    citation_sheet = workbook.create_sheet("CitationAudit")
    citation_columns = [
        "query_id",
        "candidate_id",
        "source_file_name",
        "citation_label",
        "citation_uri",
        "citation_correct_0_1",
        "citation_issue_notes",
    ]
    citation_sheet.append(citation_columns)
    sample = _citation_sample(
        candidates,
        min(int(config["citation_audit"]["target_sample"]), len(candidates)),
        config["citation_audit"]["deterministic_seed"],
    )
    for _, row in sample.iterrows():
        citation_sheet.append([row[column] for column in citation_columns[:5]] + ["", ""])
    _style_header(citation_sheet)
    citation_sheet.freeze_panes = "A2"
    citation_sheet.auto_filter.ref = f"A1:G{citation_sheet.max_row}"
    _set_widths(citation_sheet, {"A": 14, "B": 24, "C": 36, "D": 55, "E": 55, "F": 22, "G": 60})
    citation_validation = DataValidation(type="list", formula1='"0,1"', allow_blank=False)
    citation_sheet.add_data_validation(citation_validation)
    citation_validation.add(f"F2:F{citation_sheet.max_row}")
    for row in citation_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.protection = openpyxl.styles.Protection(locked=cell.column not in {6, 7})
    citation_sheet.protection.sheet = True
    citation_sheet.protection.password = "devoteam"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def read_completed_adjudication(
    path: Path, candidates: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    if not path.exists():
        return pd.DataFrame(), pd.DataFrame(), [f"Missing file: {path.name}."]
    try:
        adjudication = _read_sheet(path, "Adjudication")
        citations = _read_sheet(path, "CitationAudit")
    except Exception as exc:
        return pd.DataFrame(), pd.DataFrame(), [f"{path.name} cannot be read: {exc}."]
    issues: list[str] = []
    expected = set(candidates["candidate_id"].astype(str))
    observed = set(adjudication["candidate_id"].astype(str))
    if expected != observed or len(adjudication) != len(candidates):
        issues.append("Adjudication candidate rows were added, deleted, or changed.")
    values = pd.to_numeric(adjudication["adjudicated_relevance"], errors="coerce")
    if values.isna().any() or not values.dropna().isin([0, 1, 2]).all():
        issues.append("Every adjudication row requires relevance 0, 1, or 2.")
    disagreements = adjudication["labeler_1_relevance"].astype(str).ne(
        adjudication["labeler_2_relevance"].astype(str)
    )
    notes = adjudication["adjudication_notes"].astype(str).str.strip()
    if (disagreements & notes.eq("")).any():
        issues.append("Every labeler disagreement requires an adjudication rationale.")
    citation_values = pd.to_numeric(citations["citation_correct_0_1"], errors="coerce")
    if citation_values.isna().any() or not citation_values.dropna().isin([0, 1]).all():
        issues.append("Every citation-audit row requires 0 or 1.")
    citation_notes = citations["citation_issue_notes"].astype(str).str.strip()
    if (citation_values.eq(0) & citation_notes.eq("")).any():
        issues.append("Every incorrect citation requires issue notes.")
    clean = adjudication[["query_id", "candidate_id"]].copy()
    clean["adjudicated_relevance"] = values
    citation_clean = citations[["query_id", "candidate_id"]].copy()
    citation_clean["citation_correct"] = citation_values
    citation_clean["citation_issue_notes"] = citation_notes
    return clean, citation_clean, issues


def _metric_row(grades: dict[str, int], ranked: list[str]) -> dict[str, float]:
    relevant = {document_id for document_id, grade in grades.items() if int(grade) >= 1}
    top10 = ranked[:10]
    top5 = ranked[:5]
    if not relevant:
        return {
            "answerable": 0.0,
            "recall_at_10": math.nan,
            "precision_at_5": 0.0,
            "mrr": math.nan,
            "ndcg_at_10": math.nan,
        }
    recall = len(relevant & set(top10)) / len(relevant)
    precision = len(relevant & set(top5)) / 5.0
    reciprocal_rank = next(
        (1.0 / rank for rank, document_id in enumerate(ranked, start=1) if document_id in relevant),
        0.0,
    )
    gains = [(2 ** int(grades.get(document_id, 0)) - 1) for document_id in top10]
    dcg = sum(gain / math.log2(rank + 1) for rank, gain in enumerate(gains, start=1))
    ideal_gains = sorted((2 ** int(value) - 1 for value in grades.values()), reverse=True)[:10]
    ideal = sum(gain / math.log2(rank + 1) for rank, gain in enumerate(ideal_gains, start=1))
    return {
        "answerable": 1.0,
        "recall_at_10": recall,
        "precision_at_5": precision,
        "mrr": reciprocal_rank,
        "ndcg_at_10": dcg / ideal if ideal else math.nan,
    }


def evaluate_rankings(
    rankings: pd.DataFrame,
    candidates: pd.DataFrame,
    adjudicated: pd.DataFrame,
    queries: pd.DataFrame,
    systems: Iterable[str],
) -> tuple[pd.DataFrame, dict]:
    labels = candidates[["query_id", "candidate_id", "document_id"]].merge(
        adjudicated, on=["query_id", "candidate_id"], validate="one_to_one"
    )
    grade_lookup = {
        query_id: {
            str(row["document_id"]): int(row["adjudicated_relevance"])
            for _, row in group.iterrows()
        }
        for query_id, group in labels.groupby("query_id")
    }
    query_lookup = queries.set_index("query_id")
    rows: list[dict] = []
    for query_id in queries["query_id"]:
        grades = grade_lookup[query_id]
        for system in systems:
            ranked = (
                rankings.loc[
                    rankings["query_id"].eq(query_id)
                    & rankings["retrieval_mode"].eq(system)
                ]
                .sort_values("rank")["document_id"]
                .astype(str)
                .tolist()
            )
            rows.append(
                {
                    "query_id": query_id,
                    "language": query_lookup.loc[query_id, "language"],
                    "query_type": query_lookup.loc[query_id, "query_type"],
                    "retrieval_mode": system,
                    "relevant_documents_in_judged_pool": sum(value >= 1 for value in grades.values()),
                    **_metric_row(grades, ranked),
                }
            )
    per_query = pd.DataFrame(rows)
    aggregate: dict[str, dict] = {}
    metrics = ["recall_at_10", "precision_at_5", "mrr", "ndcg_at_10"]
    for system, group in per_query.groupby("retrieval_mode"):
        answerable = group.loc[group["answerable"].eq(1.0)]
        aggregate[system] = {
            "queries": int(len(group)),
            "answerable_queries": int(len(answerable)),
            "no_relevant_queries": int(len(group) - len(answerable)),
            **{
                metric: round(float(answerable[metric].mean()), 6) if len(answerable) else None
                for metric in metrics
            },
            "precision_at_5_all_queries": round(float(group["precision_at_5"].mean()), 6),
        }
    return per_query, aggregate


def paired_bootstrap(
    per_query: pd.DataFrame,
    systems: Iterable[str],
    metrics: Iterable[str],
    samples: int,
    seed: int,
) -> dict:
    rng = np.random.default_rng(seed)
    systems = list(systems)
    result: dict[str, dict] = {}
    for left_index, left in enumerate(systems):
        for right in systems[left_index + 1 :]:
            key = f"{left}_minus_{right}"
            result[key] = {}
            for metric in metrics:
                pivot = per_query.pivot(index="query_id", columns="retrieval_mode", values=metric)[[left, right]].dropna()
                differences = (pivot[left] - pivot[right]).to_numpy(dtype=float)
                if differences.size == 0:
                    result[key][metric] = {"mean_difference": None, "ci95": [None, None], "queries": 0}
                    continue
                draws = rng.choice(differences, size=(samples, len(differences)), replace=True).mean(axis=1)
                result[key][metric] = {
                    "mean_difference": round(float(differences.mean()), 6),
                    "ci95": [round(float(np.percentile(draws, 2.5)), 6), round(float(np.percentile(draws, 97.5)), 6)],
                    "queries": int(len(differences)),
                }
    return result


def _subgroup_metrics(per_query: pd.DataFrame) -> dict:
    output: dict[str, dict] = {}
    for dimension in ("language", "query_type"):
        output[dimension] = {}
        for (value, mode), group in per_query.groupby([dimension, "retrieval_mode"]):
            answerable = group.loc[group["answerable"].eq(1.0)]
            output[dimension][f"{value}|{mode}"] = {
                "queries": int(len(group)),
                "answerable_queries": int(len(answerable)),
                "recall_at_10": round(float(answerable["recall_at_10"].mean()), 6) if len(answerable) else None,
                "mrr": round(float(answerable["mrr"].mean()), 6) if len(answerable) else None,
                "ndcg_at_10": round(float(answerable["ndcg_at_10"].mean()), 6) if len(answerable) else None,
            }
    return output


def _select_candidate(metrics: dict, config: dict) -> tuple[str | None, str]:
    gate = float(config["evaluation"]["pilot_gates"]["recall_at_10"])
    eligible = [mode for mode, values in metrics.items() if values.get("recall_at_10") is not None and values["recall_at_10"] >= gate]
    if not eligible:
        return None, "BLOCKED_PRIMARY_RECALL_GATE"
    eligible.sort(
        key=lambda mode: (
            -metrics[mode]["recall_at_10"],
            -metrics[mode]["ndcg_at_10"],
            -metrics[mode]["mrr"],
            -metrics[mode]["precision_at_5"],
            mode,
        )
    )
    return eligible[0], "ELIGIBLE_FOR_SUPERVISOR_REVIEW"


def _hash_outputs(run_root: Path, paths: Iterable[Path]) -> Path:
    lines = [f"{sha256_file(path)}  {path.relative_to(run_root).as_posix()}" for path in sorted(paths)]
    output = run_root / "SHA256SUMS.txt"
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return output


def _final_report(manifest: dict, metrics: dict, agreement: dict) -> str:
    lines = [
        "# Phase 5.1 — Frozen expert retrieval evaluation",
        "",
        f"**Decision status:** {manifest['decision_status']}  ",
        f"**Recommended baseline for review:** {manifest.get('recommended_candidate') or 'NONE'}  ",
        "**Automatic production promotion:** PROHIBITED",
        "",
        "## Evidence quality",
        "",
        f"- Queries: **{manifest['expert_query_count']}**",
        f"- Candidate judgments per labeler: **{manifest['candidate_judgments_per_labeler']}**",
        f"- Raw label agreement: **{agreement['raw_agreement']:.3f}**",
        f"- Quadratic weighted kappa: **{agreement['quadratic_weighted_kappa']:.3f}**",
        f"- Citation correctness: **{manifest['citation_correctness']:.3f}**",
        "",
        "## Retrieval metrics",
        "",
        "| System | Recall@10 | Precision@5 | MRR | nDCG@10 |",
        "|---|---:|---:|---:|---:|",
    ]
    for mode in ("bm25", "dense", "hybrid"):
        value = metrics[mode]
        lines.append(
            f"| {mode} | {value['recall_at_10']:.3f} | {value['precision_at_5']:.3f} | {value['mrr']:.3f} | {value['ndcg_at_10']:.3f} |"
        )
    lines.extend(
        [
            "",
            "## Decision boundary",
            "",
            "This preregistered frozen set supports one baseline comparison only. It must not be reused to tune RRF weights, prompts, embeddings, or rerankers. The named supervisor and security/governance owners must approve any pilot promotion.",
            "",
        ]
    )
    return "\n".join(lines)


def _finalize(
    *,
    run_root: Path,
    phase5_manifest: dict,
    queries: pd.DataFrame,
    governance: dict[str, str],
    rankings: pd.DataFrame,
    candidates: pd.DataFrame,
    first: pd.DataFrame,
    second: pd.DataFrame,
    merged: pd.DataFrame,
    agreement: dict,
    adjudicated: pd.DataFrame,
    citation_audit: pd.DataFrame,
    diagnostics: dict,
    config: dict,
) -> dict:
    final_root = run_root / "final"
    reports_root = run_root / "reports"
    final_root.mkdir(parents=True, exist_ok=True)
    reports_root.mkdir(parents=True, exist_ok=True)
    final_labels = candidates[["query_id", "candidate_id", "document_id"]].merge(
        adjudicated, on=["query_id", "candidate_id"], validate="one_to_one"
    )
    final_labels.to_parquet(final_root / "expert_gold_labels.parquet", index=False)
    queries.to_parquet(final_root / "expert_queries.parquet", index=False)
    rankings.to_parquet(final_root / "frozen_rankings.parquet", index=False)
    per_query, system_metrics = evaluate_rankings(
        rankings, candidates, adjudicated, queries, config["retrieval"]["systems"]
    )
    per_query.to_parquet(final_root / "per_query_metrics.parquet", index=False)
    bootstrap = paired_bootstrap(
        per_query,
        config["retrieval"]["systems"],
        config["evaluation"]["metrics"],
        int(config["evaluation"]["bootstrap_samples"]),
        int(config["evaluation"]["bootstrap_seed"]),
    )
    subgroups = _subgroup_metrics(per_query)
    citation_correctness = float(citation_audit["citation_correct"].mean())
    recommended, decision_status = _select_candidate(system_metrics, config)
    reliability_pass = agreement["quadratic_weighted_kappa"] >= float(
        config["labeling"]["minimum_weighted_kappa"]
    )
    citation_pass = citation_correctness >= float(
        config["evaluation"]["pilot_gates"]["citation_correctness"]
    )
    if not reliability_pass:
        decision_status = "BLOCKED_LOW_LABEL_AGREEMENT"
        recommended = None
    if not citation_pass:
        decision_status = "BLOCKED_CITATION_CORRECTNESS_GATE"
        recommended = None
    metrics_payload = {
        "schema_version": 1,
        "gold_set_purpose": config["evaluation"]["gold_set_purpose"],
        "expert_labeled": True,
        "system_metrics": system_metrics,
        "subgroup_metrics": subgroups,
        "paired_bootstrap": bootstrap,
        "agreement": agreement,
        "citation_correctness": round(citation_correctness, 6),
        "deterministic_filter_correctness": diagnostics["filter_correctness"],
        "indicative_secondary_targets": config["evaluation"]["indicative_secondary_targets"],
        "production_quality_claim_allowed": decision_status == "ELIGIBLE_FOR_SUPERVISOR_REVIEW",
    }
    metrics_path = final_root / "expert_metrics.json"
    _atomic_json(metrics_path, metrics_payload)
    manifest = {
        "schema_version": 1,
        "phase": "5.1",
        "pipeline_version": config["pipeline_version"],
        "snapshot_id": config["input"]["snapshot_id"],
        "completed_at_utc": _now(),
        "status": "EXPERT_EVALUATION_COMPLETE",
        "decision_status": decision_status,
        "recommended_candidate": recommended,
        "automatic_production_promotion": False,
        "supervisor_review_required": True,
        "expert_query_count": int(len(queries)),
        "candidate_judgments_per_labeler": int(len(candidates)),
        "labelers": [governance["labeler_1_name"], governance["labeler_2_name"]],
        "adjudicator": governance["adjudicator_name"],
        "supervisor": governance["supervisor_name"],
        "weighted_kappa": agreement["quadratic_weighted_kappa"],
        "citation_correctness": round(citation_correctness, 6),
        "filter_correctness": diagnostics["filter_correctness"],
        "phase5_manifest_sha256": config["input"]["expected_phase5_manifest_sha256"],
        "phase5_output_mutation_calls": 0,
        "phase4_output_mutation_calls": 0,
        "source_snapshot_mutation_calls": 0,
        "external_embedding_api_calls": 0,
        "external_llm_calls": 0,
        "reranker_calls": 0,
        "raw_query_operational_log_rows": 0,
        "query_intake_sha256": sha256_file(run_root / "human_inputs" / "PHASE_5_1_QUERY_INTAKE.xlsx"),
        "labeler_1_packet_sha256": sha256_file(run_root / "human_inputs" / "LABELER_1.xlsx"),
        "labeler_2_packet_sha256": sha256_file(run_root / "human_inputs" / "LABELER_2.xlsx"),
        "adjudication_packet_sha256": sha256_file(run_root / "human_inputs" / "ADJUDICATION_AND_CITATION_AUDIT.xlsx"),
    }
    manifest_path = run_root / "PHASE_5_1_MANIFEST.json"
    _atomic_json(manifest_path, manifest)
    report_path = reports_root / "PHASE_5_1_EXPERT_EVALUATION_REPORT.md"
    report_path.write_text(_final_report(manifest, system_metrics, agreement), encoding="utf-8")
    outputs = [
        manifest_path,
        metrics_path,
        final_root / "expert_gold_labels.parquet",
        final_root / "expert_queries.parquet",
        final_root / "frozen_rankings.parquet",
        final_root / "per_query_metrics.parquet",
        report_path,
    ]
    sums_path = _hash_outputs(run_root, outputs)
    _atomic_json(
        run_root / "_SUCCESS.json",
        {
            "schema_version": 1,
            "pipeline_version": config["pipeline_version"],
            "snapshot_id": config["input"]["snapshot_id"],
            "status": "EXPERT_EVALUATION_COMPLETE",
            "decision_status": decision_status,
            "manifest_sha256": sha256_file(manifest_path),
            "sha256sums_sha256": sha256_file(sums_path),
            "created_at_utc": _now(),
        },
    )
    return _write_state(
        run_root,
        "EXPERT_EVALUATION_COMPLETE",
        decision_status=decision_status,
        recommended_candidate=recommended,
        metrics=system_metrics,
        weighted_kappa=agreement["quadratic_weighted_kappa"],
        citation_correctness=round(citation_correctness, 6),
    )


def advance_phase5_1(project_root: Path, config_path: Path, progress=print) -> dict:
    project_root = Path(project_root).resolve()
    config = load_phase5_1_config(Path(config_path))
    phase5_root, phase5_manifest = _verify_signed_phase5(project_root, config)
    phase5_config = load_phase5_config(project_root / "config" / "phase5_retrieval.yaml")
    run_root = _run_root(project_root, config)
    human_root = run_root / "human_inputs"
    audit_root = run_root / "audit"
    run_root.mkdir(parents=True, exist_ok=True)
    human_root.mkdir(parents=True, exist_ok=True)
    audit_root.mkdir(parents=True, exist_ok=True)
    intake_path = human_root / "PHASE_5_1_QUERY_INTAKE.xlsx"
    if not intake_path.exists():
        create_query_intake_workbook(intake_path, config)
        progress("Created controlled query-intake workbook. No model was downloaded.")
        return _write_state(
            run_root,
            "AWAITING_QUERY_INTAKE",
            next_action="Complete Governance and all 50 Queries rows, then rerun this notebook.",
            query_intake_path=str(intake_path),
        )
    queries, governance, query_issues = validate_query_intake(intake_path, config)
    validation_path = audit_root / "QUERY_INTAKE_VALIDATION.md"
    validation_path.write_text(_validation_report("Query intake validation", query_issues), encoding="utf-8")
    if query_issues:
        progress(f"Query intake is incomplete: {len(query_issues)} issue(s). No model was downloaded.")
        return _write_state(
            run_root,
            "AWAITING_QUERY_INTAKE",
            next_action="Fix the query-intake validation issues and rerun this notebook.",
            issue_count=len(query_issues),
            validation_report=str(validation_path),
            query_intake_path=str(intake_path),
        )

    rankings_path = audit_root / "FROZEN_SYSTEM_RANKINGS.parquet"
    candidates_path = audit_root / "BLINDED_CANDIDATE_POOL.parquet"
    queries_path = audit_root / "FROZEN_EXPERT_QUERIES.parquet"
    diagnostics_path = audit_root / "CANDIDATE_DIAGNOSTICS.json"
    query_hash_path = audit_root / "QUERY_INTAKE_SHA256.txt"
    frozen_hash = query_hash_path.read_text(encoding="utf-8").strip() if query_hash_path.exists() else None
    current_hash = sha256_file(intake_path)
    if rankings_path.exists() and frozen_hash != current_hash:
        raise AssertionError(
            "Query intake changed after candidate packets were frozen. Preserve this run and start a new evaluation version."
        )
    if not rankings_path.exists():
        chunks = pd.read_parquet(phase5_root / "chunk_lookup.parquet")
        embeddings = np.load(phase5_root / "embeddings.npy", mmap_mode="r")
        bm25 = BM25Index.load(phase5_root / "bm25_index.npz", phase5_root / "bm25_vocabulary.json")
        retrieval_texts = build_retrieval_texts(chunks)
        progress("Loading the pinned multilingual E5 model locally for 50 expert queries.")
        adapter = E5EmbeddingAdapter(phase5_config)
        query_vectors = adapter.encode_queries(queries["query_text"].tolist())
        engine = HybridSearchEngine(
            chunks=chunks,
            retrieval_texts=retrieval_texts,
            bm25=bm25,
            embeddings=embeddings,
            config=phase5_config,
            embedding_adapter=adapter,
        )
        rankings, candidates, diagnostics = build_candidate_pool(
            queries=queries,
            engine=engine,
            query_vectors=query_vectors,
            config=config,
            progress=progress,
        )
        rankings.to_parquet(rankings_path, index=False)
        candidates.to_parquet(candidates_path, index=False)
        queries.to_parquet(queries_path, index=False)
        _atomic_json(diagnostics_path, diagnostics)
        query_hash_path.write_text(current_hash + "\n", encoding="utf-8")
        create_labeler_packet(human_root / "LABELER_1.xlsx", candidates, config, 1)
        create_labeler_packet(human_root / "LABELER_2.xlsx", candidates, config, 2)
        progress("Created two blinded independent-labeling packets.")
        return _write_state(
            run_root,
            "AWAITING_INDEPENDENT_LABELS",
            next_action="Two different experts complete LABELER_1.xlsx and LABELER_2.xlsx independently, then rerun.",
            candidate_judgments_per_labeler=int(len(candidates)),
            labeler_1_path=str(human_root / "LABELER_1.xlsx"),
            labeler_2_path=str(human_root / "LABELER_2.xlsx"),
        )

    rankings = pd.read_parquet(rankings_path)
    candidates = pd.read_parquet(candidates_path)
    queries = pd.read_parquet(queries_path)
    diagnostics = json.loads(diagnostics_path.read_text(encoding="utf-8"))
    first, first_issues = read_completed_labeler_packet(
        human_root / "LABELER_1.xlsx", candidates, config
    )
    second, second_issues = read_completed_labeler_packet(
        human_root / "LABELER_2.xlsx", candidates, config
    )
    label_issues = first_issues + second_issues
    label_validation_path = audit_root / "LABEL_PACKET_VALIDATION.md"
    label_validation_path.write_text(
        _validation_report("Independent label packet validation", label_issues), encoding="utf-8"
    )
    if label_issues:
        return _write_state(
            run_root,
            "AWAITING_INDEPENDENT_LABELS",
            next_action="Complete both independent labeler packets and rerun this notebook.",
            issue_count=len(label_issues),
            validation_report=str(label_validation_path),
        )

    merged, agreement = label_agreement(first, second)
    _atomic_json(audit_root / "LABEL_AGREEMENT.json", agreement)
    adjudication_path = human_root / "ADJUDICATION_AND_CITATION_AUDIT.xlsx"
    if not adjudication_path.exists():
        create_adjudication_workbook(
            adjudication_path, merged, candidates, agreement, config
        )
        return _write_state(
            run_root,
            "AWAITING_ADJUDICATION_AND_CITATION_AUDIT",
            next_action="The independent adjudicator resolves disagreements and completes CitationAudit, then rerun.",
            disagreements=agreement["disagreements"],
            weighted_kappa=agreement["quadratic_weighted_kappa"],
            adjudication_path=str(adjudication_path),
        )

    adjudicated, citation_audit, adjudication_issues = read_completed_adjudication(
        adjudication_path, candidates
    )
    adjudication_validation_path = audit_root / "ADJUDICATION_VALIDATION.md"
    adjudication_validation_path.write_text(
        _validation_report("Adjudication and citation-audit validation", adjudication_issues),
        encoding="utf-8",
    )
    if adjudication_issues:
        return _write_state(
            run_root,
            "AWAITING_ADJUDICATION_AND_CITATION_AUDIT",
            next_action="Complete adjudication and citation audit, then rerun this notebook.",
            issue_count=len(adjudication_issues),
            validation_report=str(adjudication_validation_path),
        )

    return _finalize(
        run_root=run_root,
        phase5_manifest=phase5_manifest,
        queries=queries,
        governance=governance,
        rankings=rankings,
        candidates=candidates,
        first=first,
        second=second,
        merged=merged,
        agreement=agreement,
        adjudicated=adjudicated,
        citation_audit=citation_audit,
        diagnostics=diagnostics,
        config=config,
    )


def verify_phase5_1(run_root: Path) -> dict:
    run_root = Path(run_root)
    state_path = run_root / "PHASE_5_1_STATE.json"
    if not state_path.exists():
        raise AssertionError("Phase 5.1 state is missing")
    state = json.loads(state_path.read_text(encoding="utf-8"))
    if state["status"] != "EXPERT_EVALUATION_COMPLETE":
        return state
    manifest_path = run_root / "PHASE_5_1_MANIFEST.json"
    sums_path = run_root / "SHA256SUMS.txt"
    success_path = run_root / "_SUCCESS.json"
    success = json.loads(success_path.read_text(encoding="utf-8"))
    if sha256_file(manifest_path) != success["manifest_sha256"]:
        raise AssertionError("Phase 5.1 manifest hash mismatch")
    if sha256_file(sums_path) != success["sha256sums_sha256"]:
        raise AssertionError("Phase 5.1 checksum-list hash mismatch")
    for line in sums_path.read_text(encoding="utf-8").splitlines():
        expected, relative = line.split("  ", 1)
        if sha256_file(run_root / relative) != expected:
            raise AssertionError(f"Phase 5.1 output hash mismatch: {relative}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest["automatic_production_promotion"] is not False:
        raise AssertionError("Phase 5.1 improperly enabled automatic promotion")
    if any(
        manifest[field] != 0
        for field in (
            "phase5_output_mutation_calls",
            "phase4_output_mutation_calls",
            "source_snapshot_mutation_calls",
            "external_embedding_api_calls",
            "external_llm_calls",
            "reranker_calls",
            "raw_query_operational_log_rows",
        )
    ):
        raise AssertionError("Phase 5.1 security invariant failed")
    return manifest
